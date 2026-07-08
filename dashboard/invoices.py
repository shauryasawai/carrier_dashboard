"""Carrier invoice parser + cost aggregation.

Handles the heterogeneous carrier invoice/billing files Frido receives and
normalizes every billed line into a common record:
    {carrier, awb, order_id, sku, sku_name, product, category,
     weight_kg, zone, amount, shipments}

Supported inputs
----------------
* Frido "Working" billing template (Invoice Amt (₹) + Order ID/AWB) — used by
  ElasticRun, SPS, Swift, BlueDart pre-billing, etc. One generic adapter.
* Urban Bolt billing (Invoice Value + Awb Number).
* BlueDart B2B (NET + SKUs + Sub_Cat).
* SkyAir (Total + SKU Name/Code).
* Frido Prime forward CSV (total_charges + product_cat).
* Master files (Arcatron Weights, Required wts, Swift SKU master) — no charges;
  used to enrich SKU/category onto invoices by AWB / SKU code.

Carrier is read from a Courier column when present, else inferred from the file
name. Reads .xlsx (openpyxl), .xlsb (pyxlsb) and .csv/.tsv (stdlib csv).
"""
from __future__ import annotations

import csv as _csv
import io as _io
import math
import re
from functools import lru_cache
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Category resolution
# --------------------------------------------------------------------------
CATEGORY_RULES = [
    ("Maternity & Baby Care", ["pregnancy", "maternity", "baby", "infant", "nursing", "feeding pillow"]),
    ("Barefoot", ["barefoot", "sock shoe", "skinners"]),
    ("Mobility Devices", ["wheelchair", "commode", "scooter", "walker", "rollator", "transfer lift",
                          "bed rail", "grab bar", "guardrail", "safety rail", "ramp", "crutch",
                          "bath mat", "shower stool", "shower chair", "mobility"]),
    ("Chairs", ["ergo chair", "ergonomic chair", "executive chair", "gaming chair", "office chair",
                "ergoluxe", "ergo luxe", "posture plus chair", "study chair", "recliner", "chair"]),
    ("Workspace", ["standing desk", "desk converter", "desk", "laptop table", "study table",
                   "work table", "laptop stand", "monitor stand", "monitor arm", "monitor mount",
                   "laptop mount", "laptop holder", "monitor", "workstation", "footrest", "foot rest"]),
    ("Insoles", ["insole", "arch support", "arch cushion", "shoe insert", "foot insert", "arch sports"]),
    ("Socks", ["sock"]),
    ("Masks", ["face mask", "n95", "n-95", "anti pollution", "anti-pollution", "pollution mask", "surgical mask"]),
    ("Mattress Topper Protector", ["mattress topper", "mattress protector", "mattress", "topper"]),
    ("Covers", ["cuddle cover", "wedge cover", "pillow cover", "cushion cover", "seat cover",
                "replacement cover", "cover"]),
    ("Pillows", ["neck pillow", "cervical", "wedge", "travel pillow", "memory foam pillow",
                 "sleep pillow", "cozy pillow", "pillow"]),
    ("Cushions", ["seat cushion", "coccyx", "donut", "backrest", "back rest", "lumbar", "cushion", "seat"]),
    ("Footwear", ["sandal", "slipper", "chappal", "flip flop", "flipflop", "clog", "sneaker", "footwear", "shoe"]),
    ("Orthotics", ["posture", "orthotic", "knee", "ankle", "elbow", "wrist", "shoulder",
                   "brace", "wrap", "support belt", "lumbo", "sacral", "bunion", "heel",
                   "plantar", "toe", "compression", "belt", "support"]),
    ("Personal Care", ["eye mask", "sleep mask", "mask", "therapy", "heating pad", "heat pad",
                       "nasal", "nose", "massager", "massage", "roller", "pain relief",
                       "kinesiology", "tape", "glove"]),
    ("Accessories", ["cap", "pouch", "bag", "strap", "wallet", "card holder", "cardholder",
                     "spare part", "sparepart", "castor", "joystick", "accessor", "combo"]),
]

# Maps an EXPLICIT category label (e.g. the item-master's own "Category" column
# value, lower-cased) straight to a Frido website category. Keys cover the exact
# labels used in item_master.xlsx so SKU-master categories are authoritative.
# Deliberately NOT mapped (so the product-NAME keyword rules decide instead):
#   "combo"          — bundles span categories; classify by the product name
#   "ergo furniture" — split into Chairs vs Workspace by the product name
EXPLICIT_MAP = {
    "orthotics": "Orthotics", "orthotic": "Orthotics",
    "footwear": "Footwear", "footwears": "Footwear", "shoes": "Footwear",
    "insole": "Insoles", "insoles": "Insoles",
    "pillows": "Pillows", "pillow": "Pillows",
    "cushion": "Cushions", "cushions": "Cushions",
    "mattress topper": "Mattress Topper Protector", "mattress": "Mattress Topper Protector",
    "topper": "Mattress Topper Protector",
    "socks": "Socks", "sock": "Socks",
    "covers": "Covers", "cover": "Covers", "cuddle covers": "Covers", "wedge covers": "Covers",
    "cap": "Accessories", "accessories": "Accessories", "accessory": "Accessories",
    "sparepart (chair & mobility)": "Accessories", "sparepart": "Accessories",
    "gloves": "Accessories", "furnishing": "Accessories",
    "mask": "Masks", "masks": "Masks",
    "eye mask": "Personal Care", "personal care": "Personal Care",
    "maternity": "Maternity & Baby Care", "baby": "Maternity & Baby Care",
    "barefoot": "Barefoot",
    "chairs": "Chairs", "chair": "Chairs",
    "workspace": "Workspace",
    "mobility": "Mobility Devices", "wheelchairs": "Mobility Devices", "wheelchair": "Mobility Devices",
}


# Precompiled: one alternation regex per category (any keyword -> that category),
# checked in category order. Far fewer operations than the nested keyword loop.
_CATEGORY_RES = [(cat, re.compile("|".join(re.escape(k) for k in kws)))
                 for cat, kws in CATEGORY_RULES]


@lru_cache(maxsize=20000)
def resolve_category(explicit, name_text):
    # Memoized: invoice lines repeat the same SKU/product names heavily, so the
    # cache turns most calls into a dict hit.
    ex = (explicit or "").strip()
    if ex:
        first = re.split(r"[|/]", ex)[0].strip().lower()
        if first in EXPLICIT_MAP:
            return EXPLICIT_MAP[first]
    text = (name_text or "").lower()
    for cat, rx in _CATEGORY_RES:
        if rx.search(text):
            return cat
    if ex:
        low = ex.lower()
        for cat, rx in _CATEGORY_RES:
            if rx.search(low):
                return cat
    return "Others"


# --------------------------------------------------------------------------
# value helpers
# --------------------------------------------------------------------------
_NUM_CLEAN = re.compile(r"[,₹\s]")


def _f(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):   # already numeric: skip string/regex work
        return float(v)
    try:
        return float(_NUM_CLEAN.sub("", str(v)))
    except (ValueError, TypeError):
        return None


def _s(v):
    return "" if v is None else str(v).strip()


_JUNK = {"", "#n/a", "n/a", "na", "nan", "0", "none", "null", "-", "0.0"}


def _clean_name(v):
    s = _s(v)
    return "" if s.lower() in _JUNK else s


def _norm_awb(v):
    s = _s(v).upper()
    s = re.sub(r"\.0$", "", s)
    return s


def _norm_zone(z):
    z = _s(z)
    return z.title() if z else ""


def _unwrap_xl(v):
    """Strip Excel text-formula wrapping used by Delhivery CSVs, e.g.
    `="54484610119195"` -> `54484610119195`. Leaves normal values untouched."""
    s = _s(v)
    if s.startswith("="):
        s = s[1:]
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1]
    return s.strip().strip('"').strip()


def _nh(h):
    """Normalize a header cell for matching."""
    s = "" if h is None else str(h)
    s = s.replace("\n", " ").lower()
    s = s.replace("₹", " ")
    s = re.sub(r"\(.*?\)", " ", s)   # drop "(₹)", "(kg)", "(incl gst)" etc.
    s = re.sub(r"\s+", " ", s).strip()
    return s


# --------------------------------------------------------------------------
# low-level readers
# --------------------------------------------------------------------------
def _xlsx_sheets(data):
    wb = load_workbook(filename=_io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        out.append((name, [r for r in ws.iter_rows(values_only=True)]))
    return out


def _xlsb_sheets(data):
    from pyxlsb import open_workbook
    out = []
    with open_workbook(_io.BytesIO(data)) as wb:
        for name in wb.sheets:
            rows = []
            with wb.get_sheet(name) as sheet:
                for row in sheet.rows():
                    rows.append([c.v for c in row])
            out.append((name, rows))
    return out


def _csv_rows(data):
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1", "replace")
    return list(_csv.reader(_io.StringIO(text)))


def _read_sheets(data, filename=""):
    name = (filename or "").lower()
    if name.endswith(".xlsb"):
        return _xlsb_sheets(data)
    if data[:2] == b"PK":
        return _xlsx_sheets(data)
    return [("csv", _csv_rows(data))]


# --------------------------------------------------------------------------
# header location + field maps
# --------------------------------------------------------------------------
AMOUNT_KEYS = ["invoice amt", "invoice value", "net amount", "net",
               "grand total", "total amount", "cost", "total freight", "total"]
ORDER_KEYS = ["order id", "order number", "order_id", "client_order_id", "reference number"]
AWB_KEYS = ["awb number", "awb no.", "awbno", "awb", "tracking_id", "global_tracking_id",
            "awb_number", "waybill number", "transaction_id", "swift id", "cawbno", "wbn"]
ZONE_KEYS = ["zone charged", "zone", "shipment zone", "pricing zone"]
WEIGHT_KEYS = ["weight", "charge weight", "round weight", "billedwt", "frido final wt in kg"]
SKU_KEYS = ["skus", "sku code", "sku codes", "product sku codes", "sku", "sku_list", "product code"]
NAME_KEYS = ["sku name", "product name", "product_desc", "sub_cat", "item names", "sub_category"]
CARRIER_COL_KEYS = ["courier", "courier name", "carrier", "carrier partner name"]


def _find_header(rows, need_amount=True, scan=8):
    """Return (idx, colmap) for the first row that has an amount column (when
    need_amount) plus at least one AWB/order id column."""
    for i in range(min(scan, len(rows))):
        cells = [_nh(c) for c in (rows[i] or [])]
        cset = set(cells)
        has_amount = any(k in cset for k in AMOUNT_KEYS)
        has_id = any(k in cset for k in (AWB_KEYS + ORDER_KEYS))
        if (has_amount or not need_amount) and has_id:
            cmap = {}
            for j, c in enumerate(cells):
                if c and c not in cmap:
                    cmap[c] = j
            return i, cmap
    return None, None


def _first(cmap, keys):
    for k in keys:
        if k in cmap:
            return cmap[k]
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# --------------------------------------------------------------------------
# carrier inference
# --------------------------------------------------------------------------
_CARRIER_FILE = [
    ("skyair", "SkyAir"), ("sky air", "SkyAir"),
    ("elasticrun", "ElasticRun"), ("elastic", "ElasticRun"), ("er_", "ElasticRun"), ("er ", "ElasticRun"),
    ("urbanbolt", "Urban Bolt"), ("urbane", "Urban Bolt"), ("ub_", "Urban Bolt"), ("ub ", "Urban Bolt"),
    ("swift", "Swift"), ("frido invoice summary", "Swift"),
    ("safexpress", "Safexpress"), ("safe express", "Safexpress"),
    ("sps", "SPS"),
    ("ekart", "Ekart"),
    ("prime large", "Frido Prime Large"), ("prime small", "Frido Prime Small"), ("prime", "Frido Prime"),
    ("bd b2b", "BlueDart"), ("bd_", "BlueDart"), ("bluedart", "BlueDart"), ("blue dart", "BlueDart"),
    ("prebilling", "BlueDart"), ("delhivery", "Delhivery"),
]


def carrier_from_filename(filename):
    n = (filename or "").lower()
    for token, label in _CARRIER_FILE:
        if token in n:
            return label
    return "Unknown carrier"


# --------------------------------------------------------------------------
# month inference (each invoice FILE is treated as one billing period)
# --------------------------------------------------------------------------
_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]


def month_from_filename(filename):
    """Derive a (sort_key, label) billing period from an invoice file name.

    Each uploaded invoice file is one period; files whose names resolve to the
    same month/year are merged into one bucket (so several carriers' invoices
    for May 2025 compare side by side). Recognises:
        2025-05 / 2025_05 / 05-2025, "May 2025", "May-25", "May'25", "may2025".
    When no month can be parsed the cleaned file name itself becomes the label
    (sorted after all dated periods).
    """
    base = re.sub(r"\.(xlsx|xlsm|xlsb|csv|tsv)$", "", _s(filename), flags=re.I)
    low = base.lower()
    month = year = None

    # Numeric YYYY-MM (year first)
    m = re.search(r"(20\d{2})[-_/.](0?[1-9]|1[0-2])(?!\d)", low)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
    # Numeric MM-YYYY (month first)
    if month is None:
        m = re.search(r"(?<!\d)(0?[1-9]|1[0-2])[-_/.](20\d{2})", low)
        if m:
            month, year = int(m.group(1)), int(m.group(2))
    # Month name (full or abbreviated) + optional trailing year
    if month is None:
        m = re.search(r"(?:^|[^a-z])(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*"
                      r"[\s\-_'’]*((?:20)?\d{2})?", low)
        if m:
            month = _MONTHS.get(m.group(1))
            yr = m.group(2)
            if yr:
                yr = int(yr)
                year = yr if yr >= 1900 else (2000 + yr if yr <= 79 else 1900 + yr)
    # A bare 4-digit year elsewhere in the name (e.g. "ElasticRun May FY2025")
    if month is not None and year is None:
        ym = re.search(r"(20\d{2})", low)
        if ym:
            year = int(ym.group(1))

    if month:
        if year:
            return ("%04d-%02d" % (year, month), "%s %d" % (_MONTH_NAMES[month], year))
        return ("0000-%02d" % month, _MONTH_NAMES[month])
    return ("zzzz-" + low, base or "Unknown period")


def _clean_courier(v):
    s = _s(v)
    if not s:
        return ""
    token = re.split(r"[ _\-/]", s)[0]
    return token.title() if token else s


# --------------------------------------------------------------------------
# invoice date / service month (BlueDart B2C carries an explicit invoice date)
# --------------------------------------------------------------------------
from datetime import datetime as _dt, timedelta as _td  # noqa: E402

_EXCEL_EPOCH = _dt(1899, 12, 30)


def _excel_to_date(v):
    """Convert a value to a date. Accepts an Excel serial number (what .xlsb
    stores for date cells), a datetime, or a parseable date string. Returns a
    datetime.date or None."""
    if v is None or v == "":
        return None
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, (int, float)):
        try:
            return (_EXCEL_EPOCH + _td(days=float(v))).date()
        except (ValueError, OverflowError):
            return None
    s = _s(v)
    # An ISO date embedded in a longer string (e.g. "2026-05-19 16:59:58" or
    # "2026-05-26+05:30") — take the date part.
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return _dt(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _invoice_period(v):
    """From a raw invoice-date cell return (date_label, month_label, month_key).

    date_label  -> "31 May 2026"   (display)
    month_label -> "May 2026"      (service month, display)
    month_key   -> "2026-05"       (sort key)
    Returns ("", "", "") when the date can't be parsed."""
    d = _excel_to_date(v)
    if d is None:
        return "", "", ""
    return (d.strftime("%d %b %Y"),
            "%s %d" % (_MONTH_NAMES[d.month], d.year),
            "%04d-%02d" % (d.year, d.month))


def _billing_month(filename, rows, idx, date_i):
    """Billing (service) month for carriers whose files carry no per-row invoice
    date — SkyAir, Shadowfax/Prime. Prefer the month parsed from the file name;
    if the name has no recognisable month, fall back to the most common date in
    the given column so it works for any month regardless of naming.
    Returns (month_key, month_label)."""
    mkey, mlabel = month_from_filename(filename)
    if not mkey.startswith("zzzz"):
        return mkey, mlabel
    counts = {}
    if date_i is not None:
        for row in rows[idx + 1:]:
            d = _excel_to_date(_cell(row, date_i))
            if d:
                counts[(d.year, d.month)] = counts.get((d.year, d.month), 0) + 1
    if counts:
        yy, mm = max(counts, key=counts.get)
        return "%04d-%02d" % (yy, mm), "%s %d" % (_MONTH_NAMES[mm], yy)
    return mkey, mlabel


# --------------------------------------------------------------------------
# adapters
# --------------------------------------------------------------------------
def _parse_frido_prime(sheets, filename):
    """Shadowfax 'Frido Prime' forward billing (Prime Large / Prime Small). The
    file has product_desc / product_cat / weight but no per-row invoice number
    and no GST column, so the invoice is the tier (from client_name) + service
    month (from picked_date), and Amount with GST is derived at +18%."""
    GST = 1.18
    for name, rows in sheets:
        idx, cmap = None, None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if "awb_number" in cells and "product_cat" in cells and "total_charges" in cells:
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = cmap.get("awb_number")
        amt_i = cmap.get("total_charges")
        client_i = cmap.get("client_name")
        # No per-row invoice number/date — treat the whole file as one billing
        # period, taken from the file name (e.g. "...Prime Large_FORWARD_May26"),
        # falling back to the most common picked_date month when the name has none.
        date_i = _first(cmap, ["picked_date", "received_at_hub_date", "last_updated"])
        mkey, mlabel = _billing_month(filename, rows, idx, date_i)
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            amt = _f(_cell(row, amt_i))
            if not awb or amt is None:
                continue
            desc = _clean_name(_cell(row, cmap.get("product_desc")))
            cat = _s(_cell(row, cmap.get("product_cat")))
            tier = re.sub(r"(?i)^my\s*frido\s*-\s*", "", _s(_cell(row, client_i))).strip() or "Prime"
            inv_no = (tier + " " + mlabel).strip() if mlabel else tier
            out.append({
                "carrier": "Shadowfax", "awb": awb,
                "order_id": _s(_cell(row, cmap.get("client_order_id"))),
                "sku": "", "sku_name": desc, "product": desc,
                "category": resolve_category(cat, desc),
                "weight_kg": _f(_cell(row, cmap.get("weight"))),
                "zone": _norm_zone(_cell(row, cmap.get("zone"))),
                "amount": round(amt * GST, 2), "shipments": 1,
                "amount_ex_gst": round(amt, 2),
                "invoice_number": inv_no,
                "invoice_date": "",
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_bluedart_b2b(sheets, filename):
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if "net" in cells and "skus" in cells and ("awb no." in cells or "cawbno" in cells):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, cmap.get("awb no.")) or _cell(row, cmap.get("cawbno")))
            amt = _f(_cell(row, cmap.get("net"))) or _f(_cell(row, cmap.get("total ")))
            if not awb:
                continue
            nm = _clean_name(_cell(row, cmap.get("sub_cat")))
            out.append({
                "carrier": "BlueDart", "awb": awb, "order_id": "",
                "sku": _s(_cell(row, cmap.get("skus"))), "sku_name": nm, "product": nm,
                "category": resolve_category("", nm),
                "weight_kg": _f(_cell(row, cmap.get("nchrgwt"))),
                "zone": _norm_zone(_cell(row, cmap.get("zone"))),
                "amount": amt or 0.0, "shipments": 1,
            })
        if out:
            return out
    return None


def _parse_bluedart_b2c(sheets, filename):
    """BlueDart B2C raw billing dump (one row per AWB) with explicit invoice
    columns: CINVOICENBR, DINVDATE and GROSS TOTAL (amount incl. GST). Captures
    invoice number / date / service month so the management reconciliation table
    can group billed lines by invoice and compute the final payable.

    Reads EVERY sheet that matches the B2C layout (real exports can split lines
    across multiple tabs), accumulating all billed lines so nothing is dropped.
    The header row can sit a few rows down (a banner row precedes it), so we
    scan the first several rows of each sheet for it."""
    out = []
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(15, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if ("cinvoicenbr" in cells and "cawbno" in cells
                    and ("gross total" in cells or "with gst" in cells)):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        amt_i = _first(cmap, ["gross total", "with gst"])
        exgst_i = _first(cmap, ["total", "ex gst", "ntotalamt"])
        awb_i = cmap.get("cawbno")
        inv_i = cmap.get("cinvoicenbr")
        date_i = cmap.get("dinvdate")
        wt_i = _first(cmap, ["nchrgwt", "final wt", "nactwgt"])
        zone_i = cmap.get("zone")
        dir_i = cmap.get("fwd/reverse")
        for row in rows[idx + 1:]:
            if not row:
                continue
            inv_no = _s(_cell(row, inv_i))
            amt = _f(_cell(row, amt_i))
            awb = _norm_awb(_cell(row, awb_i))
            # A line needs an invoice number plus either an AWB or an amount.
            if not inv_no or (amt is None and not awb):
                continue
            dlabel, mlabel, mkey = _invoice_period(_cell(row, date_i))
            out.append({
                "carrier": "BlueDart", "awb": awb, "order_id": "",
                "sku": "", "sku_name": "", "product": "",
                "category": "Others",
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": _f(_cell(row, exgst_i)) or 0.0,
                "invoice_number": inv_no,
                "invoice_date": dlabel,
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": _s(_cell(row, dir_i)).upper(),
            })
    return out or None


def _parse_skyair(sheets, filename):
    """SkyAir hyperlocal billing ('Sheet 1'). Columns: AWB / Pickup Date /
    Round Weight (kg) / Before Tax (ex GST) / GST / Total (with GST); some
    exports also carry SKU Name. No invoice-number column, so the file is treated
    as one invoice for the billing month taken from the file name."""
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if ("awb" in cells and "total" in cells
                    and ("before tax" in cells or "round weight" in cells or "sku name" in cells)):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = cmap.get("awb")
        amt_i = cmap.get("total")
        exgst_i = _first(cmap, ["before tax", "ex gst"])
        wt_i = _first(cmap, ["round weight", "charge weight", "sky air weight", "weight"])
        name_i = _first(cmap, ["sku name", "product name"])
        sku_i = _first(cmap, ["sku code", "sku codes", "sku"])
        zone_i = _first(cmap, ["city", "zone"])
        date_i = _first(cmap, ["pickup date", "receiving date", "order last status date"])
        mkey, mlabel = _billing_month(filename, rows, idx, date_i)
        inv_no = ("SkyAir " + mlabel).strip() if mlabel else "SkyAir"
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            amt = _f(_cell(row, amt_i))
            if not awb or amt is None:
                continue
            nm = _clean_name(_cell(row, name_i))
            ex = _f(_cell(row, exgst_i))
            out.append({
                "carrier": "SkyAir", "awb": awb, "order_id": "",
                "sku": _s(_cell(row, sku_i)), "sku_name": nm, "product": nm,
                "category": resolve_category("", nm),
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": ex if ex is not None else round((amt or 0.0) / 1.18, 2),
                "invoice_number": inv_no,
                "invoice_date": "",
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_swift(sheets, filename):
    """Swift B2C billing detail. Columns: Swift Id / AWB / Cost (incl GST) /
    Weight (in gms) / Billing Date / Direction / Product Description. The invoice
    number lives in the SHEET NAME (e.g. 'Swift Invoice #SWT26001442 Details'),
    so every line on the sheet shares that invoice number. SKU(s) and product
    name are parsed from the Product Description: '[sku] {name} {qty}'."""
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(8, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if "swift id" in cells and "awb" in cells and ("cost" in cells or "billing date" in cells):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        amt_i = _first(cmap, ["cost", "cost incl gst", "freight", "shipping cost"])
        awb_i = cmap.get("awb")
        date_i = _first(cmap, ["billing date", "invoice date", "date"])
        wt_i = _first(cmap, ["weight", "weight in gms", "charge weight"])
        dir_i = cmap.get("direction")
        desc_i = _first(cmap, ["product description", "item names", "product name"])
        car_i = _first(cmap, ["courier name", "courier", "carrier"])
        ord_i = _first(cmap, ["order number", "order id", "order_id"])
        zone_i = cmap.get("zone")
        # Invoice number from the sheet name: token after '#', else first
        # alphanumeric code that contains a digit.
        m = re.search(r"#\s*([A-Za-z0-9][\w-]*)", name or "") \
            or re.search(r"([A-Za-z]{2,}\d[\w-]*)", name or "")
        inv_no = m.group(1) if m else (name or filename)
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            amt = _f(_cell(row, amt_i))
            if not awb or amt is None:
                continue
            draw = _s(_cell(row, date_i))
            dm = re.search(r"(\d{4}-\d{2}-\d{2})", draw)
            dlabel, mlabel, mkey = _invoice_period(dm.group(1) if dm else _cell(row, date_i))
            desc = _s(_cell(row, desc_i))
            skus = re.findall(r"\[([^\]]+)\]", desc)
            sku = ", ".join(s.strip().upper() for s in skus)
            names = re.findall(r"\{([^}]+)\}", desc)
            pname = _clean_name(names[0]) if names else ""
            wt_g = _f(_cell(row, wt_i))
            carrier = _clean_courier(_cell(row, car_i)) or "Swift"
            out.append({
                "carrier": carrier or "Swift", "awb": awb,
                "order_id": _s(_cell(row, ord_i)),
                "sku": sku, "sku_name": pname, "product": pname,
                "category": "Others",
                "weight_kg": (wt_g / 1000.0) if wt_g else None,
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": round(amt / 1.18, 2) if amt else 0.0,
                "invoice_number": inv_no,
                "invoice_date": dlabel,
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": _s(_cell(row, dir_i)).upper(),
            })
        if out:
            return out
    return None


def _parse_delhivery(sheets, filename):
    """Delhivery billing files (CSV or XLSX). Columns: waybill_num / serial_number
    / total_amount (incl GST) / gross_amount (ex GST) / charged_weight (grams) /
    zone / order_id / item_shipped|product_description / pickup_date. CSVs wrap
    text fields in Excel formulas (`="..."`) which are unwrapped. The invoice
    number is the serial_number (one per file)."""
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(6, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if "waybill_num" in cells and "total_amount" in cells and "gross_amount" in cells:
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = cmap.get("waybill_num")
        amt_i = cmap.get("total_amount")
        exgst_i = cmap.get("gross_amount")
        serial_i = cmap.get("serial_number")
        date_i = _first(cmap, ["pickup_date", "status_date", "invoice date"])
        wt_i = cmap.get("charged_weight")
        zone_i = cmap.get("zone")
        ord_i = cmap.get("order_id")
        name_i = _first(cmap, ["item_shipped", "product_description", "product_desc"])
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_unwrap_xl(_cell(row, awb_i)))
            amt = _f(_unwrap_xl(_cell(row, amt_i)))
            if not awb or amt is None:
                continue
            inv_no = _unwrap_xl(_cell(row, serial_i)) or _s(filename)
            draw = _cell(row, date_i)
            if isinstance(draw, (int, float)):
                dlabel, mlabel, mkey = _invoice_period(draw)
            else:
                ds = _unwrap_xl(draw)
                dm = re.search(r"(\d{4}-\d{2}-\d{2})", ds)
                dlabel, mlabel, mkey = _invoice_period(dm.group(1) if dm else ds)
            wt_g = _f(_unwrap_xl(_cell(row, wt_i)))   # charged weight in grams
            nm = _clean_name(_unwrap_xl(_cell(row, name_i)).rstrip(","))
            out.append({
                "carrier": "Delhivery", "awb": awb,
                "order_id": _unwrap_xl(_cell(row, ord_i)),
                "sku": "", "sku_name": nm, "product": nm,
                "category": "Others",
                "weight_kg": (wt_g / 1000.0) if wt_g else None,
                "zone": _norm_zone(_unwrap_xl(_cell(row, zone_i))),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": _f(_unwrap_xl(_cell(row, exgst_i))) or 0.0,
                "invoice_number": inv_no,
                "invoice_date": dlabel,
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_urbanbolt(sheets, filename):
    """Urban Bolt billing summary. Columns: AWB No / Shipment Date /
    Chg. Weight (kg) / Freight Subtotal (ex GST) / GST Amount / Total Amount
    (with GST). No invoice-number column, so the file is one invoice for the
    billing month (file name, falling back to the Shipment Date)."""
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if ("awb no" in cells and "total amount" in cells
                    and ("freight subtotal" in cells or "gst amount" in cells
                         or "chg. weight" in cells or "chg weight" in cells)):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = _first(cmap, ["awb no", "awb number", "awb"])
        amt_i = cmap.get("total amount")
        exgst_i = _first(cmap, ["freight subtotal", "base charge"])
        wt_i = _first(cmap, ["chg. weight", "chg weight", "charge weight", "weight"])
        date_i = _first(cmap, ["shipment date", "pickup date", "date"])
        zone_i = _first(cmap, ["lane / zone", "zone", "lane"])
        mkey, mlabel = _billing_month(filename, rows, idx, date_i)
        inv_no = ("Urban Bolt " + mlabel).strip() if mlabel else "Urban Bolt"
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            amt = _f(_cell(row, amt_i))
            if not awb or amt is None:
                continue
            ex = _f(_cell(row, exgst_i))
            out.append({
                "carrier": "Urban Bolt", "awb": awb, "order_id": "",
                "sku": "", "sku_name": "", "product": "",
                "category": "Others",
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": ex if ex is not None else round((amt or 0.0) / 1.18, 2),
                "invoice_number": inv_no,
                "invoice_date": "",
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_safexpress(sheets, filename):
    """Safexpress B2B billing ('Data' sheet). Columns: Bill Number (invoice) /
    Waybill Number (AWB) / Pickup Date / Charge Weight (kg) / Total Freight
    (ex GST) / GST Amount / Grand Total (with GST). One bill per file."""
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if ("waybill number" in cells and "grand total" in cells
                    and ("bill number" in cells or "total freight" in cells)):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = cmap.get("waybill number")
        amt_i = cmap.get("grand total")
        exgst_i = _first(cmap, ["total freight", "total amount"])
        inv_i = cmap.get("bill number")
        date_i = _first(cmap, ["pickup date", "invoice date", "booking date"])
        wt_i = _first(cmap, ["charge weight", "charged weight", "chargeable weight", "weight"])
        zone_i = _first(cmap, ["waybill destination", "destination", "zone"])
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            amt = _f(_cell(row, amt_i))
            inv_no = _s(_cell(row, inv_i))
            if not awb or amt is None:
                continue
            dlabel, mlabel, mkey = _invoice_period(_cell(row, date_i))
            out.append({
                "carrier": "Safexpress", "awb": awb, "order_id": "",
                "sku": "", "sku_name": "", "product": "",
                "category": "Others",
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt or 0.0, "shipments": 1,
                "amount_ex_gst": _f(_cell(row, exgst_i)) or 0.0,
                "invoice_number": inv_no or _s(filename),
                "invoice_date": dlabel,
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_elasticrun(sheets, filename):
    """ElasticRun billing ('Data' sheet). Columns: transaction_id (AWB) /
    client_invoice_number / client_invoice_date / work_complete_time /
    chargeable_weight / total_charge. The file carries the pre-tax charge only
    (no GST column), so Amount with GST is derived at +18%. Invoice number is the
    client_invoice_number; service month is taken from work_complete_time (when
    the shipment was actually done), while the displayed invoice date is the
    client_invoice_date."""
    GST = 1.18
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(4, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            if ("transaction_id" in cells and "total_charge" in cells
                    and "client_invoice_number" in cells):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = cmap.get("transaction_id")
        charge_i = cmap.get("total_charge")
        inv_i = cmap.get("client_invoice_number")
        invdate_i = cmap.get("client_invoice_date")
        svc_i = _first(cmap, ["work_complete_time", "client_invoice_date"])
        wt_i = _first(cmap, ["chargeable_weight", "weight"])
        zone_i = _first(cmap, ["zone", "state"])
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            awb = _norm_awb(_cell(row, awb_i))
            charge = _f(_cell(row, charge_i))
            inv_no = _s(_cell(row, inv_i))
            if not awb or charge is None or not inv_no:
                continue
            _, _, mkey = _invoice_period(_cell(row, svc_i))
            _, mlabel, _ = _invoice_period(_cell(row, svc_i))
            dlabel, _, _ = _invoice_period(_cell(row, invdate_i))
            out.append({
                "carrier": "ElasticRun", "awb": awb, "order_id": "",
                "sku": "", "sku_name": "", "product": "",
                "category": "Others",
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": round(charge * GST, 2), "shipments": 1,
                "amount_ex_gst": round(charge, 2),
                "invoice_number": inv_no,
                "invoice_date": dlabel,
                "service_month": mlabel,
                "service_month_key": mkey,
                "direction": "",
            })
        if out:
            return out
    return None


def _parse_generic(sheets, filename):
    """Frido 'Working' billing template + Urban Bolt: any sheet with an amount
    column (Invoice Amt (₹) / Invoice Value / ...) plus an AWB or Order id."""
    carrier_fb = carrier_from_filename(filename)
    for name, rows in sheets:
        idx, cmap = _find_header(rows, need_amount=True)
        if idx is None:
            continue
        amt_i = _first(cmap, AMOUNT_KEYS)
        awb_i = _first(cmap, AWB_KEYS)
        ord_i = _first(cmap, ORDER_KEYS)
        zone_i = _first(cmap, ZONE_KEYS)
        wt_i = _first(cmap, WEIGHT_KEYS)
        sku_i = _first(cmap, SKU_KEYS)
        name_i = _first(cmap, NAME_KEYS)
        car_i = _first(cmap, CARRIER_COL_KEYS)
        out = []
        for row in rows[idx + 1:]:
            if not row:
                continue
            amt = _f(_cell(row, amt_i))
            awb = _norm_awb(_cell(row, awb_i))
            order = _s(_cell(row, ord_i))
            if amt is None or (not awb and not order):
                continue
            carrier = _clean_courier(_cell(row, car_i)) or carrier_fb
            nm = _clean_name(_cell(row, name_i))
            out.append({
                "carrier": carrier or "Unknown carrier", "awb": awb, "order_id": order,
                "sku": _s(_cell(row, sku_i)), "sku_name": nm, "product": nm,
                "category": resolve_category(nm, nm),
                "weight_kg": _f(_cell(row, wt_i)),
                "zone": _norm_zone(_cell(row, zone_i)),
                "amount": amt, "shipments": 1,
            })
        if out:
            return out
    return None


_ADAPTERS = [_parse_frido_prime, _parse_bluedart_b2c, _parse_bluedart_b2b,
             _parse_skyair, _parse_swift, _parse_delhivery, _parse_elasticrun,
             _parse_safexpress, _parse_urbanbolt, _parse_generic]


# --------------------------------------------------------------------------
# master files (no charges) -> AWB / SKU enrichment
# --------------------------------------------------------------------------
# Header keywords for a per-SKU volumetric / billable weight column (an item
# master used to detect carrier weight over-charges).
VOL_WEIGHT_KEYS = ["vol. weight", "vol. wt", "volumetric weight", "volumetric wt",
                   "vol weight", "vol wt", "volumetric", "billable weight", "billable wt",
                   "charge weight", "chargeable weight", "expected weight", "weight kg",
                   "weight"]


def _parse_master(sheets):
    """Detect a SKU/weight master and return enrichment maps, or None if it isn't
    a master:
        {'awb2cat': {awb: (cat, product, sku)},
         'sku2cat': {sku: (cat, product)},
         'sku2vol': {sku: volumetric_weight_kg}}
    sku2vol is populated when the file carries a per-SKU volumetric / billable
    weight column (used to flag carrier weight over-charges)."""
    awb2cat, sku2cat, sku2vol = {}, {}, {}
    found = False
    for name, rows in sheets:
        idx = cmap = None
        for i in range(min(6, len(rows))):
            cells = set(_nh(c) for c in (rows[i] or []))
            has_cat = ("category" in cells)
            has_awb = any(k in cells for k in AWB_KEYS)
            has_sku = any(k in cells for k in ["sku_list", "sku code", "product code", "skus", "sku"])
            # Reached only after every invoice adapter declined, so we don't need
            # to exclude files that merely carry an (empty) amount column.
            if (has_cat or has_sku) and (has_awb or has_sku):
                idx = i
                cmap = {_nh(c): j for j, c in enumerate(rows[i] or []) if _nh(c)}
                break
        if idx is None:
            continue
        awb_i = _first(cmap, AWB_KEYS)
        cat_i = cmap.get("category")
        sub_i = cmap.get("sub_category")
        sku_i = _first(cmap, ["sku_list", "sku code", "product code", "skus", "sku"])
        nm_i = _first(cmap, ["sku name", "product name", "product_desc", "sub_cat"])
        vol_i = _first(cmap, VOL_WEIGHT_KEYS)   # per-SKU volumetric / billable weight
        for row in rows[idx + 1:]:
            if not row:
                continue
            cat_raw = _s(_cell(row, cat_i))
            nm = _clean_name(_cell(row, nm_i)) or _clean_name(_cell(row, sub_i))
            sku = _s(_cell(row, sku_i))
            if not (cat_raw or nm or sku):
                continue
            cat = resolve_category(cat_raw, nm)
            awb = _norm_awb(_cell(row, awb_i)) if awb_i is not None else ""
            if awb:
                awb2cat[awb] = (cat, nm or sku, sku)
                found = True
            if sku and (nm or cat != "Others"):
                sku2cat[sku.upper()] = (cat, nm or sku)
                found = True
            if sku and vol_i is not None:
                vol = _f(_cell(row, vol_i))
                if vol is not None and vol > 0:
                    sku2vol[sku.upper()] = vol
                    found = True
    return ({"awb2cat": awb2cat, "sku2cat": sku2cat, "sku2vol": sku2vol}
            if found else None)


def ingest(data, filename=""):
    """Classify and parse a file. Returns ('invoice', items) or ('master', maps)."""
    sheets = _read_sheets(data, filename)
    for adapter in _ADAPTERS:
        try:
            items = adapter(list(sheets), filename)
        except Exception:
            items = None
        if items:
            return ("invoice", items)
    master = _parse_master(sheets)
    if master:
        return ("master", master)
    raise ValueError("Unrecognised file format: " + (filename or "file"))


def parse_invoice(data, filename=""):
    kind, payload = ingest(data, filename)
    if kind != "invoice":
        raise ValueError("This looks like a master/reference file, not an invoice: " + (filename or "file"))
    return payload


# --------------------------------------------------------------------------
# Frido category images (CDN) + icon fallback
# --------------------------------------------------------------------------
_CDN = "https://cdn.shopify.com/s/files/1/0553/0419/2034/files/"
CATEGORY_IMAGES = {
    "Orthotics": _CDN + "Category_1_2_1.jpg?v=1769695217&width=400",
    "Insoles": _CDN + "Everyday_Insole_Combo_55a98dda-d383-4a27-9f57-39c4a5e8bcab.png?v=1770808742&width=400",
    "Footwear": _CDN + "WS-01_8b6068d3-4e48-4b4c-8d37-07fd0338d7cb.jpg?v=1753101892&width=400",
    "Pillows": _CDN + "CNP-Black-01B_4997708e-d0e7-49d3-af3f-fe7d51f06423.jpg?v=1739886430&width=400",
    "Cushions": _CDN + "Wedge-plus_color-picker-tab_2edb45fe-649c-4966-b706-0d148f3de35d.png?v=1747912121&width=400",
    "Mattress Topper Protector": _CDN + "UMT_2.jpg?v=1720264311&width=400",
    "Chairs": _CDN + "Ergo-chair-01_f909b535-6d3d-4bc9-b4bd-0c3ff7ecdeee.jpg?v=1729333804&width=400",
}
CATEGORY_ICONS = {
    "Orthotics": "🩺", "Insoles": "👣", "Footwear": "🥿", "Pillows": "🛏️",
    "Cushions": "🪑", "Mattress Topper Protector": "🛌", "Mobility Devices": "♿", "Socks": "🧦",
    "Maternity & Baby Care": "🤰", "Personal Care": "💆", "Accessories": "🎒",
    "Workspace": "🖥️", "Chairs": "🪑", "Barefoot": "🦶", "Masks": "😷",
    "Covers": "🛡️", "Others": "📦",
}


def _avg(spend, n):
    return round(spend / n, 1) if n else None


def _enrich(items, awb2cat, sku2cat):
    """Fill category / product / sku for weak rows using the master maps. A
    master category only overrides when it is a real (non-'Others') value."""
    filled = 0
    for it in items:
        weak = it["category"] in ("Others", "Unknown", "")
        if not (weak or not it["product"]):
            continue
        ent = awb2cat.get(it["awb"]) if it["awb"] else None
        if ent:
            cat, prod, sku = ent
            if weak and cat and cat != "Others":
                it["category"] = cat; weak = False
            if not it["product"] and prod:
                it["product"] = prod
            if not it["sku"] and sku:
                it["sku"] = sku
        skey = (it["sku"] or "").upper()
        if weak and skey and skey in sku2cat:
            cat, prod = sku2cat[skey]
            if cat and cat != "Others":
                it["category"] = cat; weak = False
            if not it["product"] and prod:
                it["product"] = prod
        if not weak:
            filled += 1
    return filled


def _attach_lanes(items, awb2lane, order2lane):
    """Tag each invoice line with its pickup/drop pincode by joining to the
    loaded shipment data on AWB (falling back to order id). Lines whose AWB
    isn't in the shipment set get empty pins and drop out of the lane table."""
    awb2lane = awb2lane or {}
    order2lane = order2lane or {}
    matched = 0
    for it in items:
        lane = None
        awb = it.get("awb")
        if awb and awb in awb2lane:
            lane = awb2lane[awb]
        if lane is None:
            oid = (it.get("order_id") or "").strip().upper()
            if oid and oid in order2lane:
                lane = order2lane[oid]
        it["pickup_pin"], it["drop_pin"] = lane if lane else ("", "")
        if lane:
            matched += 1
    return matched


def _attach_products(items, awb2prod, order2prod):
    """Tag each invoice line with the product info (category, sub-category, SKU,
    item name) of its shipment, joined from the loaded BigQuery shipment data on
    AWB (falling back to order id). Lines whose AWB isn't in the shipment set
    keep an empty sub-category and are bucketed as "Unmatched" in the analysis.
    Returns the number of lines that matched."""
    awb2prod = awb2prod or {}
    order2prod = order2prod or {}
    matched = 0
    for it in items:
        prod = None
        awb = it.get("awb")
        if awb and awb in awb2prod:
            prod = awb2prod[awb]
        if prod is None:
            oid = (it.get("order_id") or "").strip().upper()
            if oid and oid in order2prod:
                prod = order2prod[oid]
        if prod:
            cat, sub, sku, name = prod
            it["subcategory"] = sub or it.get("subcategory") or ""
            if sku and not it.get("sku"):
                it["sku"] = sku
            if name and not it.get("product"):
                it["product"] = name
                it["sku_name"] = it.get("sku_name") or name
            if cat and cat not in ("Others", "", "Unknown"):
                it["category"] = cat
            it["prod_matched"] = True
            matched += 1
        else:
            it.setdefault("subcategory", "")
            it["prod_matched"] = bool(it.get("prod_matched"))
    return matched


def _attach_values(items, awb2value, order2value):
    """Tag each invoice line with its shipment's declared order value (selling
    price), joined on AWB (falling back to order id). Lines with no match keep
    order_value = None and are excluded from the freight-to-value ratio."""
    awb2value = awb2value or {}
    order2value = order2value or {}
    for it in items:
        val = None
        awb = it.get("awb")
        if awb and awb in awb2value:
            val = awb2value[awb]
        if val is None:
            oid = (it.get("order_id") or "").strip().upper()
            if oid and oid in order2value:
                val = order2value[oid]
        it["order_value"] = val


def _build_subcategory_analysis(items):
    """Spend (with GST) and shipment counts grouped by sub-category, plus a
    matched/unmatched summary. Lines with no joined sub-category fall into the
    "Unmatched" bucket so totals always reconcile to the invoice total."""
    sub = {}
    total_spend = 0.0
    total_ship = 0
    total_weight = 0.0
    matched_ship = 0
    matched_spend = 0.0
    for i in items:
        amt = i["amount"]
        n = i["shipments"]
        wt = i.get("weight_kg") or 0.0   # NCHRGWT — the charged/billed weight
        total_spend += amt
        total_ship += n
        total_weight += wt
        is_m = bool(i.get("prod_matched"))
        if is_m:
            matched_ship += n
            matched_spend += amt
        key = (i.get("subcategory") or "").strip() or ("Unmatched" if not is_m else "Other")
        g = sub.setdefault(key, {"subcategory": key, "spend": 0.0, "shipments": 0,
                                 "weight": 0.0})
        g["spend"] += amt
        g["shipments"] += n
        g["weight"] += wt
    rows = []
    for g in sorted(sub.values(), key=lambda x: -x["spend"]):
        rows.append({
            "subcategory": g["subcategory"],
            "spend": round(g["spend"], 2),
            "shipments": g["shipments"],
            "weight": round(g["weight"], 1),   # total charged kg
            "share": round(g["spend"] / total_spend * 100, 1) if total_spend else 0,
        })
    return {
        "rows": rows,
        "matched_shipments": matched_ship,
        "matched_spend": round(matched_spend, 2),
        "total_shipments": total_ship,
        "total_spend": round(total_spend, 2),
        "total_weight": round(total_weight, 1),
        "subcategory_count": sum(1 for r in rows if r["subcategory"] not in ("Unmatched",)),
    }


def _build_product_analysis(items, top=500):
    """Spend (with GST), shipments and charged weight grouped by PRODUCT (SKU),
    across the whole invoice set. Products are keyed by their SKU string (the
    BigQuery product_sku_code; a combo line keeps its combined SKU), with the
    item name shown for readability. Lines with no SKU fall into a single
    "(no SKU / unmatched)" bucket so totals reconcile to the invoice total."""
    prod = {}
    total_spend = total_weight = total_value = 0.0
    total_ship = 0
    for i in items:
        amt = i["amount"]
        n = i["shipments"]
        wt = i.get("weight_kg") or 0.0
        val = i.get("order_value") or 0.0
        total_spend += amt
        total_ship += n
        total_weight += wt
        total_value += val
        sku = (i.get("sku") or "").strip()
        name = (i.get("product") or i.get("sku_name") or "").strip()
        key = sku.upper() or "(no SKU / unmatched)"
        g = prod.setdefault(key, {"sku": sku, "name": name, "spend": 0.0,
                                  "shipments": 0, "weight": 0.0, "sell_value": 0.0})
        g["spend"] += amt
        g["shipments"] += n
        g["weight"] += wt
        g["sell_value"] += val
        if not g["name"] and name:
            g["name"] = name
        if not g["sku"] and sku:
            g["sku"] = sku
    rows = []
    for g in sorted(prod.values(), key=lambda x: -x["spend"]):
        rows.append({
            "sku": g["sku"],
            "name": g["name"] or g["sku"] or "(no SKU / unmatched)",
            "spend": round(g["spend"], 2),
            "shipments": g["shipments"],
            "weight": round(g["weight"], 1),
            # Total declared order value (selling price) across this product's
            # matched shipments; None when nothing matched so the % blanks out.
            "sell_value": round(g["sell_value"], 2) if g["sell_value"] else None,
            "share": round(g["spend"] / total_spend * 100, 1) if total_spend else 0,
        })
    return {
        "rows": rows[:top],
        "product_total": len(rows),
        "shown": min(top, len(rows)),
        "total_spend": round(total_spend, 2),
        "total_shipments": total_ship,
        "total_weight": round(total_weight, 1),
        "total_sell_value": round(total_value, 2) if total_value else None,
    }


def _build_weight_dispute(items, sku2vol, slab=1.0, min_over=1.0):
    """Flag invoice AWB lines where BlueDart's charged weight exceeds the
    expected billable weight (from the uploaded item master) by at least one
    weight slab — i.e. likely weight over-charges worth disputing.

    Per AWB: expected = sum of the master volumetric weight (kg) for the SKU(s)
    on that AWB, rounded UP to the next `slab` (0.5 kg) = expected_slab. A line is
    flagged when charged_weight - expected_slab >= min_over (one slab). The
    estimated over-charge is excess_kg x the line's own ₹/kg (amount / charged).
    Aggregated per invoice number and per sub-category for raising disputes."""
    sku2vol = sku2vol or {}
    by_inv, by_sub, by_prod = {}, {}, {}
    lines_out = []          # per-AWB flagged detail (for the CSV export)
    checked = flagged = 0
    tot_excess = tot_est = 0.0
    tot_charged = tot_expected = 0.0
    for it in items:
        skus = [s.strip().upper() for s in re.split(r"[,/|]", it.get("sku") or "") if s.strip()]
        vols = [sku2vol[s] for s in skus if s in sku2vol]
        charged = it.get("weight_kg") or 0.0
        if not vols or charged <= 0:
            continue
        expected = sum(vols)
        expected_slab = math.ceil(round(expected / slab, 6)) * slab
        checked += 1
        inv = it.get("invoice_number") or "(no invoice)"
        bi = by_inv.setdefault(inv, {"invoice_number": inv, "checked": 0,
                                     "flagged": 0, "charged_kg": 0.0, "expected_kg": 0.0,
                                     "excess_kg": 0.0, "est_overcharge": 0.0})
        bi["checked"] += 1
        excess = round(charged - expected_slab, 3)
        if excess >= min_over:
            est = excess * (it["amount"] / charged) if charged else 0.0
            flagged += 1
            bi["flagged"] += 1
            bi["charged_kg"] += charged
            bi["expected_kg"] += expected_slab
            bi["excess_kg"] += excess
            bi["est_overcharge"] += est
            sub = (it.get("subcategory") or "").strip() or "Unmatched"
            bs = by_sub.setdefault(sub, {"subcategory": sub, "flagged": 0,
                                         "charged_kg": 0.0, "expected_kg": 0.0,
                                         "excess_kg": 0.0, "est_overcharge": 0.0})
            bs["flagged"] += 1
            bs["charged_kg"] += charged
            bs["expected_kg"] += expected_slab
            bs["excess_kg"] += excess
            bs["est_overcharge"] += est
            # product (SKU) bucket
            sku_raw = (it.get("sku") or "").strip()
            pname = (it.get("product") or it.get("sku_name") or "").strip()
            pkey = sku_raw.upper() or "(no SKU)"
            bp = by_prod.setdefault(pkey, {"sku": sku_raw, "name": pname, "flagged": 0,
                                           "charged_kg": 0.0, "expected_kg": 0.0,
                                           "excess_kg": 0.0, "est_overcharge": 0.0})
            bp["flagged"] += 1
            bp["charged_kg"] += charged
            bp["expected_kg"] += expected_slab
            bp["excess_kg"] += excess
            bp["est_overcharge"] += est
            if not bp["name"] and pname:
                bp["name"] = pname
            if not bp["sku"] and sku_raw:
                bp["sku"] = sku_raw
            # per-AWB detail row (used for the AWB-level CSV export)
            lines_out.append({
                "awb": it.get("awb") or "",
                "invoice_number": it.get("invoice_number") or "",
                "carrier": it.get("carrier") or "",
                "sku": sku_raw,
                "name": pname or sku_raw,
                "subcategory": sub,
                "charged_kg": round(charged, 2),
                "expected_kg": round(expected_slab, 2),
                "excess_kg": round(excess, 2),
                "est_overcharge": round(est, 2),
            })
            tot_excess += excess
            tot_est += est
            tot_charged += charged
            tot_expected += expected_slab

    def _round_rows(rows):
        for r in rows:
            if "charged_kg" in r:
                r["charged_kg"] = round(r["charged_kg"], 1)
            if "expected_kg" in r:
                r["expected_kg"] = round(r["expected_kg"], 1)
            r["excess_kg"] = round(r["excess_kg"], 1)
            r["est_overcharge"] = round(r["est_overcharge"], 2)
        return rows

    # Per-AWB detail, grouped by product then biggest excess first, capped so the
    # payload stays reasonable for the browser.
    lines_out.sort(key=lambda x: (x["name"] or "zzz", -x["excess_kg"]))
    inv_rows = _round_rows(sorted(by_inv.values(), key=lambda x: -x["est_overcharge"]))
    sub_rows = _round_rows(sorted(by_sub.values(), key=lambda x: -x["est_overcharge"]))
    prod_rows = _round_rows(sorted(by_prod.values(), key=lambda x: -x["est_overcharge"]))
    for r in prod_rows:
        r["name"] = r["name"] or r["sku"] or "(no SKU)"
    return {
        "by_invoice": inv_rows,
        "by_subcategory": sub_rows,
        "by_product": prod_rows[:200],
        "product_count": len(prod_rows),
        "lines": lines_out[:10000],
        "line_count": len(lines_out),
        "checked": checked, "flagged": flagged,
        "total_lines": len(items),
        "charged_kg": round(tot_charged, 1),
        "expected_kg": round(tot_expected, 1),
        "excess_kg": round(tot_excess, 1),
        "est_overcharge": round(tot_est, 2),
        "slab": slab,
        "has_master": bool(sku2vol),
    }


def _build_lane_comparison(items):
    """Like-for-like price comparison: per pickup->drop pincode lane, the avg
    billed cost of each carrier, side by side. Only lanes served by 2+ carriers
    are kept, since the whole point is to compare carriers on the SAME lane."""
    lanes = {}
    for i in items:
        pin, drop = i.get("pickup_pin"), i.get("drop_pin")
        if not (pin and drop):
            continue
        g = lanes.setdefault((pin, drop), {"pickup": pin, "drop": drop, "carriers": {}})
        cc = g["carriers"].setdefault(i["carrier"], {"spend": 0.0, "shipments": 0})
        cc["spend"] += i["amount"]
        cc["shipments"] += i["shipments"]

    rows = []
    carrier_set = set()
    for g in lanes.values():
        cs = g["carriers"]
        if len(cs) < 2:                      # need 2+ carriers to compare
            continue
        cells, avgs, ship = {}, {}, 0
        for car, v in cs.items():
            avg = round(v["spend"] / v["shipments"], 1) if v["shipments"] else None
            cells[car] = {"avg_cost": avg, "shipments": v["shipments"],
                          "spend": round(v["spend"], 1)}
            if avg is not None:
                avgs[car] = avg
            ship += v["shipments"]
            carrier_set.add(car)
        if len(avgs) < 2:                    # need 2+ priced carriers
            continue
        cheapest = min(avgs, key=avgs.get)
        priciest = max(avgs, key=avgs.get)
        mn, mx = avgs[cheapest], avgs[priciest]
        rows.append({
            "pickup": g["pickup"], "drop": g["drop"],
            "lane": g["pickup"] + " → " + g["drop"],
            "shipments": ship, "carrier_count": len(cs), "cells": cells,
            "cheapest": cheapest, "priciest": priciest,
            "min_avg": mn, "max_avg": mx,
            "save_pct": round((mx - mn) / mx * 100, 1) if mx else None,
        })
    rows.sort(key=lambda r: -r["shipments"])
    # Carrier columns ordered by how many comparable lanes each appears in.
    appear = {}
    for r in rows:
        for car in r["cells"]:
            appear[car] = appear.get(car, 0) + 1
    carriers_order = sorted(carrier_set, key=lambda c: -appear.get(c, 0))
    return {
        "carriers": carriers_order,
        "rows": rows[:400],
        "lane_total": len(rows),
    }


def build_carrier_comparison(items, tds_rate=2.0):
    """Summary per carrier (across ALL loaded invoices, ignoring the carrier
    filter) so multiple carriers can be compared side by side: invoices,
    shipments, amount with / ex GST, avg ₹/parcel, charged weight, TDS @2% and
    payable (before any disputes / CNs, which are entered per-invoice in the UI)."""
    car = {}
    months = {}                 # month_key -> month_label
    cmn = {}                    # (carrier, month_key) -> aggregates
    for i in items:
        c = i.get("carrier") or "Unknown"
        g = car.setdefault(c, {"carrier": c, "invoices": set(), "shipments": 0,
                               "amount": 0.0, "amount_ex_gst": 0.0, "weight": 0.0})
        inv = i.get("invoice_number")
        if inv:
            g["invoices"].add(inv)
        g["shipments"] += i["shipments"]
        g["amount"] += i["amount"]
        g["amount_ex_gst"] += i.get("amount_ex_gst") or 0.0
        g["weight"] += i.get("weight_kg") or 0.0
        # per carrier x month
        mk = i.get("service_month_key") or "zzzz"
        months[mk] = i.get("service_month") or "Unknown"
        cg = cmn.setdefault((c, mk), {"amount": 0.0, "amount_ex_gst": 0.0,
                                      "shipments": 0, "weight": 0.0})
        cg["amount"] += i["amount"]
        cg["amount_ex_gst"] += i.get("amount_ex_gst") or 0.0
        cg["shipments"] += i["shipments"]
        cg["weight"] += i.get("weight_kg") or 0.0

    rows = []
    tot = {"invoices": 0, "shipments": 0, "amount": 0.0, "amount_ex_gst": 0.0,
           "weight": 0.0, "tds": 0.0, "payable": 0.0}
    for g in sorted(car.values(), key=lambda x: -x["amount"]):
        amt = round(g["amount"], 2)
        tds = round(amt * tds_rate / 100.0, 2)
        rows.append({
            "carrier": g["carrier"],
            "invoices": len(g["invoices"]),
            "shipments": g["shipments"],
            "amount_with_gst": amt,
            "amount_ex_gst": round(g["amount_ex_gst"], 2),
            "per_parcel": round(amt / g["shipments"], 2) if g["shipments"] else None,
            "weight": round(g["weight"], 1),
            "tds": tds,
            "payable": round(amt - tds, 2),
        })
        tot["invoices"] += len(g["invoices"])
        tot["shipments"] += g["shipments"]
        tot["amount"] += amt
        tot["amount_ex_gst"] += g["amount_ex_gst"]
        tot["weight"] += g["weight"]
        tot["tds"] += tds
        tot["payable"] += amt - tds
    totals = {
        "invoices": tot["invoices"], "shipments": tot["shipments"],
        "amount_with_gst": round(tot["amount"], 2),
        "amount_ex_gst": round(tot["amount_ex_gst"], 2),
        "per_parcel": round(tot["amount"] / tot["shipments"], 2) if tot["shipments"] else None,
        "weight": round(tot["weight"], 1),
        "tds": round(tot["tds"], 2), "payable": round(tot["payable"], 2),
    }
    # Month-wise: carrier x month cells, for trend / comparison charts.
    month_keys = sorted(months)
    month_list = [{"key": mk, "label": months[mk]} for mk in month_keys]
    by_carrier_month = {}
    for (c, mk), cg in cmn.items():
        d = by_carrier_month.setdefault(c, {})
        d[mk] = {
            "amount_with_gst": round(cg["amount"], 2),
            "amount_ex_gst": round(cg["amount_ex_gst"], 2),
            "shipments": cg["shipments"],
            "weight": round(cg["weight"], 1),
            "per_parcel": round(cg["amount"] / cg["shipments"], 2) if cg["shipments"] else None,
            "per_kg": round(cg["amount"] / cg["weight"], 2) if cg["weight"] else None,
        }
    month_totals = {}
    for mk in month_keys:
        a = sum(cmn[(c, mk)]["amount"] for c in car if (c, mk) in cmn)
        s = sum(cmn[(c, mk)]["shipments"] for c in car if (c, mk) in cmn)
        w = sum(cmn[(c, mk)]["weight"] for c in car if (c, mk) in cmn)
        month_totals[mk] = {
            "amount_with_gst": round(a, 2), "shipments": s, "weight": round(w, 1),
            "per_parcel": round(a / s, 2) if s else None,
            "per_kg": round(a / w, 2) if w else None,
        }

    return {"rows": rows, "carrier_count": len(rows), "totals": totals,
            "tds_rate": tds_rate, "months": month_list,
            "by_carrier_month": by_carrier_month, "month_totals": month_totals}


def build_cost_report(items, files=None, awb2cat=None, sku2cat=None,
                      awb2lane=None, order2lane=None,
                      awb2prod=None, order2prod=None, sku2vol=None,
                      awb2value=None, order2value=None):
    if awb2cat or sku2cat:
        _enrich(items, awb2cat or {}, sku2cat or {})

    # Join each line to the SKU / sub-category of its shipment (from loaded
    # BigQuery data) so spend can be analysed by product sub-category.
    prod_matched = _attach_products(items, awb2prod, order2prod)
    # Join each line to its shipment's declared order value (selling price) so
    # shipping spend can be shown as a % of item value in the product table.
    _attach_values(items, awb2value, order2value)
    subcategory_analysis = _build_subcategory_analysis(items)
    subcategory_analysis["matched_lines"] = prod_matched
    subcategory_analysis["total_lines"] = len(items)

    # Product-wise spend breakdown across the whole invoice (by SKU).
    product_analysis = _build_product_analysis(items)

    # Compare BlueDart's charged weight to the item-master billable weight to
    # flag likely weight over-charges (dispute opportunities).
    weight_dispute = _build_weight_dispute(items, sku2vol or {})

    # Join each line to its pickup/drop pincode (from the loaded shipment data)
    # so carrier prices can be compared on the same lane.
    lane_matched = _attach_lanes(items, awb2lane, order2lane)
    lane_comparison = _build_lane_comparison(items)
    lane_comparison["matched"] = lane_matched
    lane_comparison["total_lines"] = len(items)

    total_spend = sum(i["amount"] for i in items)
    total_ship = sum(i["shipments"] for i in items)
    total_weight = sum((i["weight_kg"] or 0.0) for i in items)

    cb = {}
    for i in items:
        g = cb.setdefault(i["carrier"], {"carrier": i["carrier"], "spend": 0.0,
                                         "shipments": 0, "weight": 0.0})
        g["spend"] += i["amount"]; g["shipments"] += i["shipments"]
        if i["weight_kg"]:
            g["weight"] += i["weight_kg"]
    carriers = []
    for g in sorted(cb.values(), key=lambda x: -x["spend"]):
        carriers.append({"carrier": g["carrier"], "spend": round(g["spend"], 1),
                         "shipments": g["shipments"], "avg_cost": _avg(g["spend"], g["shipments"]),
                         "avg_per_kg": _avg(g["spend"], g["weight"]) if g["weight"] else None,
                         "weight": round(g["weight"], 1),
                         "share": round(g["spend"] / total_spend * 100, 1) if total_spend else 0})
    carrier_names = [c["carrier"] for c in carriers]

    cat = {}
    for i in items:
        c = i["category"] or "Others"
        g = cat.setdefault(c, {"category": c, "spend": 0.0, "shipments": 0,
                               "weight": 0.0, "carriers": {}, "skus": {}})
        g["spend"] += i["amount"]; g["shipments"] += i["shipments"]
        if i["weight_kg"]:
            g["weight"] += i["weight_kg"]
        cc = g["carriers"].setdefault(i["carrier"], {"carrier": i["carrier"], "spend": 0.0, "shipments": 0})
        cc["spend"] += i["amount"]; cc["shipments"] += i["shipments"]
        skey = i["sku"] or i["product"] or "(unmapped)"
        sk = g["skus"].setdefault(skey, {"sku": i["sku"], "name": i["product"] or i["sku_name"] or skey,
                                          "spend": 0.0, "shipments": 0})
        sk["spend"] += i["amount"]; sk["shipments"] += i["shipments"]

    categories = []
    for g in sorted(cat.values(), key=lambda x: -x["spend"]):
        clist = sorted(g["carriers"].values(), key=lambda x: -x["spend"])
        carrier_rows = [{"carrier": c["carrier"], "spend": round(c["spend"], 1),
                         "shipments": c["shipments"], "avg_cost": _avg(c["spend"], c["shipments"])}
                        for c in clist]
        skus = sorted(g["skus"].values(), key=lambda x: -x["spend"])
        top_skus = [{"sku": s["sku"], "name": s["name"], "spend": round(s["spend"], 1),
                     "shipments": s["shipments"], "avg_cost": _avg(s["spend"], s["shipments"])}
                    for s in skus[:8]]
        categories.append({
            "category": g["category"],
            "image": CATEGORY_IMAGES.get(g["category"], ""),
            "icon": CATEGORY_ICONS.get(g["category"], "📦"),
            "spend": round(g["spend"], 1), "shipments": g["shipments"],
            "avg_cost": _avg(g["spend"], g["shipments"]),
            "avg_per_kg": _avg(g["spend"], g["weight"]) if g["weight"] else None,
            "share": round(g["spend"] / total_spend * 100, 1) if total_spend else 0,
            "carriers": carrier_rows, "top_skus": top_skus, "sku_count": len(g["skus"]),
        })

    cells = {}
    for c in categories:
        cells[c["category"]] = {cr["carrier"]: {"spend": cr["spend"], "shipments": cr["shipments"],
                                                "avg_cost": cr["avg_cost"]} for cr in c["carriers"]}
    matrix = {"carriers": carrier_names, "categories": [c["category"] for c in categories], "cells": cells}

    sk = {}
    for i in items:
        skey = (i["sku"] or i["product"] or "(unmapped)")
        g = sk.setdefault(skey, {"sku": i["sku"], "name": i["product"] or i["sku_name"] or skey,
                                 "category": i["category"], "spend": 0.0, "shipments": 0, "carriers": {}})
        g["spend"] += i["amount"]; g["shipments"] += i["shipments"]
        cc = g["carriers"].setdefault(i["carrier"], {"carrier": i["carrier"], "spend": 0.0, "shipments": 0})
        cc["spend"] += i["amount"]; cc["shipments"] += i["shipments"]
    skus = []
    for g in sorted(sk.values(), key=lambda x: -x["spend"]):
        crows = sorted(g["carriers"].values(), key=lambda x: -x["spend"])
        skus.append({"sku": g["sku"], "name": g["name"], "category": g["category"],
                     "spend": round(g["spend"], 1), "shipments": g["shipments"],
                     "avg_cost": _avg(g["spend"], g["shipments"]),
                     "carriers": [{"carrier": c["carrier"], "spend": round(c["spend"], 1),
                                   "shipments": c["shipments"], "avg_cost": _avg(c["spend"], c["shipments"])}
                                  for c in crows]})
    sku_total = len(skus)
    skus = skus[:500]

    # --- month-wise (per billing period) aggregation -----------------------
    # Each invoice file is one period (see month_from_filename); files mapping
    # to the same month/year merge. We track spend / shipments plus per-carrier
    # and per-category spend so the UI can compare periods side by side and show
    # month-over-month change.
    mon = {}
    for i in items:
        mk = i.get("month") or "zzzz-unknown"
        ml = i.get("month_label") or "Unknown period"
        g = mon.setdefault(mk, {"key": mk, "label": ml, "spend": 0.0, "shipments": 0,
                                "carriers": {}, "categories": {}})
        g["spend"] += i["amount"]; g["shipments"] += i["shipments"]
        g["carriers"][i["carrier"]] = g["carriers"].get(i["carrier"], 0.0) + i["amount"]
        cat = i["category"] or "Others"
        g["categories"][cat] = g["categories"].get(cat, 0.0) + i["amount"]

    months = []
    prev = None
    for mk in sorted(mon.keys()):
        g = mon[mk]
        avg = _avg(g["spend"], g["shipments"])
        entry = {
            "key": mk, "label": g["label"],
            "spend": round(g["spend"], 1), "shipments": g["shipments"], "avg_cost": avg,
            "carriers": [{"carrier": k, "spend": round(v, 1)}
                         for k, v in sorted(g["carriers"].items(), key=lambda kv: -kv[1])],
            "categories": [{"category": k, "spend": round(v, 1)}
                           for k, v in sorted(g["categories"].items(), key=lambda kv: -kv[1])],
            "spend_delta": None, "avg_delta": None,
        }
        if prev is not None:
            if prev["spend"]:
                entry["spend_delta"] = round((entry["spend"] - prev["spend"]) / prev["spend"] * 100, 1)
            if prev["avg_cost"] and avg is not None:
                entry["avg_delta"] = round((avg - prev["avg_cost"]) / prev["avg_cost"] * 100, 1)
        months.append(entry)
        prev = entry

    # Carriers / categories present across periods, ordered by total spend, so
    # the stacked charts use a stable, meaningful series order.
    month_carrier_order = [c["carrier"] for c in carriers]
    month_category_order = [c["category"] for c in categories]

    # --- per-invoice reconciliation (management payable view) --------------
    # Group billed lines by carrier + invoice number into one row per invoice.
    # The raw invoice only gives us Amount with GST + dates; the dispute and
    # credit-note (CN) adjustments are entered by the team in the UI. TDS is
    # pre-filled at 2% of the billed amount (sec 194C) and is overridable.
    TDS_RATE = 2.0
    recon_map = {}
    for i in items:
        inv_no = i.get("invoice_number")
        if not inv_no:
            continue
        key = (i.get("carrier") or "", inv_no)
        g = recon_map.setdefault(key, {
            "carrier": i.get("carrier") or "",
            "invoice_number": inv_no,
            "invoice_date": i.get("invoice_date") or "",
            "service_month": i.get("service_month") or i.get("month_label") or "",
            "service_month_key": i.get("service_month_key") or "",
            "amount_with_gst": 0.0, "amount_ex_gst": 0.0, "shipments": 0,
        })
        g["amount_with_gst"] += i["amount"]
        g["amount_ex_gst"] += i.get("amount_ex_gst") or 0.0
        g["shipments"] += i["shipments"]
        if not g["invoice_date"] and i.get("invoice_date"):
            g["invoice_date"] = i["invoice_date"]
        if not g["service_month"] and i.get("service_month"):
            g["service_month"] = i["service_month"]

    reconciliation = None
    if recon_map:
        rrows = []
        for g in sorted(recon_map.values(),
                        key=lambda x: (x["service_month_key"] or "zzzz",
                                       x["carrier"], x["invoice_number"])):
            amt = round(g["amount_with_gst"], 2)
            rrows.append({
                "carrier": g["carrier"],
                "invoice_date": g["invoice_date"],
                "invoice_number": g["invoice_number"],
                "service_month": g["service_month"],
                "shipments": g["shipments"],
                "amount_with_gst": amt,
                "amount_ex_gst": round(g["amount_ex_gst"], 2),
                # TDS (sec 194C) is deducted on the taxable value (ex GST), not
                # on the GST-inclusive amount. Overridable in the UI.
                "tds": round(g["amount_ex_gst"] * TDS_RATE / 100.0, 2),
                # Team-entered adjustments — default 0, edited in the UI.
                "billing_dispute": 0.0, "weight_dispute": 0.0,
                "billing_cn": 0.0, "weight_cn": 0.0, "lost_cn": 0.0,
            })
        reconciliation = {
            "rows": rrows,
            "tds_rate": TDS_RATE,
            "invoice_count": len(rrows),
            "total_amount": round(sum(r["amount_with_gst"] for r in rrows), 2),
            "total_amount_ex_gst": round(sum(r["amount_ex_gst"] for r in rrows), 2),
            "total_tds": round(sum(r["tds"] for r in rrows), 2),
            "total_shipments": sum(r["shipments"] for r in rrows),
            "subcategories": subcategory_analysis,
            "products": product_analysis,
            "weight_dispute": weight_dispute,
        }

    return {
        "currency": "₹",
        "summary": {"total_spend": round(total_spend, 1), "shipments": total_ship,
                    "avg_cost": _avg(total_spend, total_ship),
                    "total_weight": round(total_weight, 1),
                    "avg_per_kg": _avg(total_spend, total_weight) if total_weight else None,
                    "categories": len(categories), "skus": sku_total, "carriers": len(carriers)},
        "carriers": carriers, "categories": categories, "matrix": matrix,
        "skus": skus, "sku_total": sku_total, "files": files or [],
        "lane_comparison": lane_comparison,
        "months": months,
        "month_carriers": month_carrier_order,
        "month_categories": month_category_order,
        "reconciliation": reconciliation,
    }
