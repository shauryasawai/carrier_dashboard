"""Profit & Loss (P&L) engine.

Builds a full unit-economics P&L for the actual sales in a report, on top of the
same shipment records the efficiency dashboard already loads. Everything is
computed from the NET order value (Order value - RTO value - Cancelled/Lost
value); the four outcome buckets in kpi.build_report reconcile to the total, so

    net_order_value = revenue - rto_value - cancelled_value
                    = delivered_value + pending_value

Per-unit costs come from two persisted masters that live next to this module:

  * data/cogs_master_sku.xlsx - the "COGS" sheet: per product the Product Cost
    (col D), Consumables (col E), Direct Expense (col F) and COGS (col G), keyed
    by SKU (col J) plus "Tally Product name" (col B) / "Product Category" (col C)
    / coarse "Category" (col A). There is NO selling price in this file.
  * data/item_master.xlsx  - per-SKU volumetric / billable weight (kg), reused
    from the invoice weight-dispute pipeline (see invoices._parse_master).

The COGS file is keyed by product name / category while shipments are keyed by
SKU, so each shipment is matched to a COGS row through a layered lookup:

    1. exact product name  (record item name  -> Tally Product name)
    2. product category    (record subcategory -> Product Category)
    3. product name        (record item name   -> Product Category)
    4. coarse category      (record category    -> COGS Category average)
    5. blended average       (mean unit cost across the whole COGS master)

The P&L build-up (all per the business model):

    Selling Price (SP)          = net order value
    Product COGS                = Product Cost               (from master)
    Consumables                 = Consumables                (from master)
    Direct labour / Others      = (Product COGS + Consumables) x 10%
    -> Total cost for Gross Profit = Product COGS + Consumables + Direct labour
    Forward Logistics Cost      = Weight (kg) x 9
    -> Total cost for CM1         = Total cost for GP + Forward Logistics
    Selling Incentives          = SP x 2%
    Travelling & Event Expenses = SP x 5%
    -> Total cost for CM2         = Total cost for CM1 + Incentives + Travel
    Employee Cost               = SP x 9%
    Office Overheads            = SP x 2%
    PG / Finance Cost           = SP x 2%
    -> Total cost for EBITDA      = Total cost for CM2 + Employee + Office + PG
    EBITDA                      = SP - Total cost for EBITDA
    EBITDA Margin % (on SP)     = EBITDA / SP           (the ACTUAL blended margin)

    Estimated Selling Price (ex-GST, per unit)
        = Total cost for CM1 (per unit)
          / (1 - 2% - 5% - 9% - 2% - 2% - EBITDA Margin%)
    Rounded SP = Estimated SP rounded to the nearest Rs.10
    GST Rate   = 5% for wheelchairs / face masks, else 18%
"""

from __future__ import annotations

import logging
import os
import re

logger = logging.getLogger(__name__)

_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
# The single COGS mapping master (SKU-tagged): Category | Tally Product name |
# Product Category | Product Cost | Consumables | Direct Expense | COGS |
# COGS Return | Return % | SKU | SKU Match Confidence | Match Basis.
_COGS_PATH = os.path.join(_DATA_DIR, "cogs_master_sku.xlsx")
_ITEM_MASTER_PATH = os.path.join(_DATA_DIR, "item_master.xlsx")

# ---- P&L model rates ------------------------------------------------------
DIRECT_LABOUR_PCT = 0.10       # of (Product COGS + Consumables)
LOGISTICS_RATE_PER_KG = 9.0    # Forward logistics = Weight (kg) x 9
INCENTIVE_PCT = 0.02           # of Selling Price
TRAVEL_PCT = 0.05              # of Selling Price
EMPLOYEE_PCT = 0.09            # of Selling Price
OFFICE_PCT = 0.02              # of Selling Price
PG_FINANCE_PCT = 0.02          # of Selling Price
# Sum of the fixed %-of-SP costs charged BELOW CM1 (incentive + travel +
# employee + office + PG/finance). Used to invert the Estimated Selling Price.
BELOW_CM1_SP_PCT = (INCENTIVE_PCT + TRAVEL_PCT + EMPLOYEE_PCT
                    + OFFICE_PCT + PG_FINANCE_PCT)

GST_LOW = 5        # wheelchairs / face masks
GST_STD = 18       # everything else

# Cap the product-wise breakdown sent to the browser (ranked by net value).
PRODUCT_LIMIT = 300


def _norm(s):
    """Lower-case, strip punctuation and collapse whitespace for fuzzy keying."""
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").strip().lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _num(v):
    """Coerce a cell to float. The COGS sheet uses ' -   ' for zero, so treat any
    non-numeric string as 0.0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


# ---------------------------------------------------------------------------
# Master loading (cached by file mtime so a re-upload is picked up live).
# ---------------------------------------------------------------------------
_COGS_CACHE = {"mtime": None, "maps": None}
_WEIGHT_CACHE = {"mtime": None, "sku2vol": None}


def _blank_maps():
    return {"by_sku": {}, "by_name": {}, "by_pcat": {}, "by_cat": {},
            "blended": None, "rows": 0, "sku_count": 0, "updated_at": None}


def load_cogs():
    """Parse data/cogs_master_sku.xlsx into layered per-unit cost lookups.

    Returns a dict with:
        by_name   {normalized Tally Product name -> unit cost dict}
        by_pcat   {normalized Product Category   -> averaged unit cost dict}
        by_cat    {normalized Category (col A)    -> averaged unit cost dict}
        blended   averaged unit cost dict across the whole file (fallback)
    A unit cost dict is {product_cost, consumables, gst}.
    """
    try:
        st = os.stat(_COGS_PATH)
    except OSError:
        return _blank_maps()
    if _COGS_CACHE["maps"] is not None and _COGS_CACHE["mtime"] == st.st_mtime:
        return _COGS_CACHE["maps"]

    maps = _blank_maps()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=_COGS_PATH, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        # Header row 1: Category | Tally Product name | Product Category |
        #   Product Cost | Consumables | Direct Expense | COGS | COGS Return |
        #   Return % | SKU | SKU Match Confidence | Match Basis
        by_name, pcat_acc, cat_acc, blend = {}, {}, {}, []
        sku_acc = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            category = (row[0] or "")
            tally = (row[1] or "")
            pcat = (row[2] or "")
            product_cost = _num(row[3])
            consumables = _num(row[4])
            sku = str(row[9]).strip().upper() if (len(row) > 9 and row[9]) else ""
            if not (tally or pcat) and product_cost == 0 and consumables == 0:
                continue
            gst = _gst_for(category, pcat, tally)
            entry = {"product_cost": product_cost, "consumables": consumables,
                     "gst": gst}
            # A SKU can appear on several product rows (family-level mapping); the
            # per-SKU cost is the average across those rows (see _avg).
            if sku and sku not in ("UNMATCHED", "NONE"):
                sku_acc.setdefault(sku, []).append(entry)
            nkey = _norm(tally)
            if nkey:
                by_name[nkey] = entry
            pk = _norm(pcat)
            if pk:
                pcat_acc.setdefault(pk, []).append(entry)
            ck = _norm(category)
            if ck:
                cat_acc.setdefault(ck, []).append(entry)
            # Only positive-cost rows feed the averages so free services (0) don't
            # drag the blended fallback down.
            if product_cost > 0 or consumables > 0:
                blend.append(entry)

        def _avg(entries):
            pos = [e for e in entries if e["product_cost"] > 0 or e["consumables"] > 0]
            use = pos or entries
            if not use:
                return None
            n = len(use)
            gsts = [e["gst"] for e in use]
            return {
                "product_cost": sum(e["product_cost"] for e in use) / n,
                "consumables": sum(e["consumables"] for e in use) / n,
                # Majority GST in the bucket.
                "gst": max(set(gsts), key=gsts.count),
            }

        maps["by_sku"] = {k: _avg(v) for k, v in sku_acc.items() if _avg(v)}
        maps["by_name"] = by_name
        maps["by_pcat"] = {k: _avg(v) for k, v in pcat_acc.items() if _avg(v)}
        maps["by_cat"] = {k: _avg(v) for k, v in cat_acc.items() if _avg(v)}
        maps["blended"] = _avg(blend)
        maps["rows"] = len(by_name)
        maps["sku_count"] = len(maps["by_sku"])
        import datetime
        maps["updated_at"] = datetime.datetime.fromtimestamp(
            st.st_mtime).strftime("%d %b %Y %H:%M")
    except Exception:  # noqa: BLE001 - a bad master must never break the report
        logger.exception("COGS master parse failed")
        maps = _blank_maps()

    _COGS_CACHE["mtime"] = st.st_mtime
    _COGS_CACHE["maps"] = maps
    return maps


def load_weights():
    """{SKU(upper) -> volumetric/billable weight kg} from the item master, reusing
    the invoice master parser. Cached by mtime."""
    try:
        st = os.stat(_ITEM_MASTER_PATH)
    except OSError:
        return {}
    if _WEIGHT_CACHE["sku2vol"] is not None and _WEIGHT_CACHE["mtime"] == st.st_mtime:
        return _WEIGHT_CACHE["sku2vol"]
    sku2vol = {}
    try:
        from . import invoices
        with open(_ITEM_MASTER_PATH, "rb") as fh:
            kind, payload = invoices.ingest(fh.read(),
                                            os.path.basename(_ITEM_MASTER_PATH))
        if kind == "master":
            sku2vol = payload.get("sku2vol", {}) or {}
    except Exception:  # noqa: BLE001
        logger.exception("Weight master parse failed")
        sku2vol = {}
    _WEIGHT_CACHE["mtime"] = st.st_mtime
    _WEIGHT_CACHE["sku2vol"] = sku2vol
    return sku2vol


# ---------------------------------------------------------------------------
# GST + category helpers.
# ---------------------------------------------------------------------------
# item-master categories that map onto the coarse COGS "Category" column so the
# category-level fallback can still fire when finer matches miss.
_CAT_ALIAS = {
    "wheelchairs": "mobility", "mobility": "mobility",
    "sparepart chair mobility": "mobility",
    "orthotics": "orthotics",
    "insole": "insoles", "insoles": "insoles",
    "footwear": "ortho footwear", "shoes": "ortho footwear",
    "socks": "ortho footwear",
    "mask": "masks", "eye mask": "masks",
    "personal care": "personal care",
    "ergo furniture": "ergo furniture",
    "pillows": "body support", "cushions": "body support",
    "mattress topper": "body support", "furnishing": "body support",
}


def _gst_for(*texts):
    """5% for wheelchairs / face masks, 18% otherwise (by keyword)."""
    blob = _norm(" ".join(str(t or "") for t in texts))
    if "wheelchair" in blob or "wheel chair" in blob or "mask" in blob:
        return GST_LOW
    return GST_STD


def _match(record, cogs):
    """Return the per-unit COGS entry for a shipment record and the match level
    used (sku / name / pcat / category / blended / none).

    SKU is the strongest key now that the COGS master carries SKUs: a shipment's
    product_sku_code is matched directly against the COGS SKU. A shipment line can
    list several SKUs, so the first one present in the master wins.
    """
    by_sku = cogs.get("by_sku") or {}
    raw_sku = (record.get("sku") or "").strip()
    if raw_sku and by_sku:
        for tok in re.split(r"[,;|/]+", raw_sku):
            hit = by_sku.get(tok.strip().upper())
            if hit:
                return hit, "sku"
    name = _norm(record.get("item_name"))
    sub = _norm(record.get("subcategory"))
    cat = _norm(record.get("category"))
    if name and name in cogs["by_name"]:
        return cogs["by_name"][name], "name"
    if sub and sub in cogs["by_pcat"]:
        return cogs["by_pcat"][sub], "pcat"
    if name and name in cogs["by_pcat"]:
        return cogs["by_pcat"][name], "pcat"
    alias = _CAT_ALIAS.get(cat, cat)
    if alias and alias in cogs["by_cat"]:
        return cogs["by_cat"][alias], "category"
    if cogs["blended"]:
        return cogs["blended"], "blended"
    return None, "none"


def _round10(v):
    if v is None:
        return None
    return int(round(v / 10.0) * 10)


# ---------------------------------------------------------------------------
# The P&L build-up.
# ---------------------------------------------------------------------------
def _pnl_from_totals(sp, units, product_cost, consumables, weight,
                     ebitda_margin_for_price=None):
    """Assemble the full P&L block from accumulated totals.

    sp                        Selling Price = net order value
    product_cost/consumables  summed per-unit COGS for the net units
    weight                    summed billable weight (kg) for the net units
    ebitda_margin_for_price   EBITDA margin (fraction) to price against for the
                              Estimated Selling Price; defaults to this block's
                              own actual EBITDA margin.
    """
    direct_labour = (product_cost + consumables) * DIRECT_LABOUR_PCT
    total_gp_cost = product_cost + consumables + direct_labour
    gross_profit = sp - total_gp_cost

    fwd_logistics = weight * LOGISTICS_RATE_PER_KG
    total_cm1_cost = total_gp_cost + fwd_logistics
    cm1 = sp - total_cm1_cost

    incentives = sp * INCENTIVE_PCT
    travel = sp * TRAVEL_PCT
    total_cm2_cost = total_cm1_cost + incentives + travel
    cm2 = sp - total_cm2_cost

    employee = sp * EMPLOYEE_PCT
    office = sp * OFFICE_PCT
    pg_finance = sp * PG_FINANCE_PCT
    total_ebitda_cost = total_cm2_cost + employee + office + pg_finance
    ebitda = sp - total_ebitda_cost

    def pct(part):
        return (part / sp * 100.0) if sp else None

    ebitda_margin = (ebitda / sp) if sp else None
    price_margin = (ebitda_margin_for_price
                    if ebitda_margin_for_price is not None else ebitda_margin)

    # Estimated Selling Price (ex-GST) per unit, inverted from the per-unit CM1
    # cost so that after the %-of-SP costs and the target EBITDA margin the price
    # lands at the intended margin.
    est_sp = rounded_sp = None
    if units and price_margin is not None:
        per_unit_cm1_cost = total_cm1_cost / units
        denom = 1.0 - BELOW_CM1_SP_PCT - price_margin
        if denom > 0.01:
            est_sp = per_unit_cm1_cost / denom
            rounded_sp = _round10(est_sp)

    return {
        "units": units,
        "selling_price": round(sp, 2),
        "product_cogs": round(product_cost, 2),
        "consumables": round(consumables, 2),
        "final_cogs": round(product_cost + consumables, 2),
        "direct_labour": round(direct_labour, 2),
        "total_gp_cost": round(total_gp_cost, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_profit_pct": round(pct(gross_profit), 2) if sp else None,
        "weight": round(weight, 3),
        "fwd_logistics": round(fwd_logistics, 2),
        "total_cm1_cost": round(total_cm1_cost, 2),
        "cm1": round(cm1, 2),
        "cm1_pct": round(pct(cm1), 2) if sp else None,
        "incentives": round(incentives, 2),
        "travel": round(travel, 2),
        "total_cm2_cost": round(total_cm2_cost, 2),
        "cm2": round(cm2, 2),
        "cm2_pct": round(pct(cm2), 2) if sp else None,
        "employee": round(employee, 2),
        "office": round(office, 2),
        "pg_finance": round(pg_finance, 2),
        "total_ebitda_cost": round(total_ebitda_cost, 2),
        "ebitda": round(ebitda, 2),
        "ebitda_margin_pct": round(ebitda_margin * 100.0, 2) if ebitda_margin is not None else None,
        "est_selling_price": round(est_sp, 2) if est_sp is not None else None,
        "rounded_sp": rounded_sp,
    }


def compute_pnl(rows, sku2vol=None):
    """Build the P&L block for a filtered record set.

    `rows` must already carry the per-row de-duplicated revenue `_rev` and
    `outcome` that kpi.build_report stamps on each record. Only NET units count:
    RTO and Cancelled/Lost orders are excluded (their value is subtracted from
    order value to get the net order value, per the model).
    """
    cogs = load_cogs()
    if sku2vol is None:
        sku2vol = load_weights()
    configured = bool(cogs["rows"]) or bool(cogs["blended"])

    # Overall accumulators + a per-PRODUCT breakdown. The P&L is built ONLY from
    # lines whose COGS is an exact product/category match (tier == "matched");
    # category-average and blended-estimate lines are excluded from the P&L
    # (but still surfaced in the reconciliation so it's clear what's left out).
    tot = {"sp": 0.0, "units": 0, "pc": 0.0, "cons": 0.0, "wt": 0.0}
    match_counts = {"sku": 0, "name": 0, "pcat": 0, "category": 0, "blended": 0, "none": 0}
    default_wt = _mean_weight(sku2vol)
    prods: dict[str, dict] = {}
    catg: dict[str, dict] = {}

    # Reconciliation against the efficiency view's revenue cards: summed on the
    # SAME per-row de-duplicated revenue (_rev), so these totals equal the
    # Order value / RTO value / Cancelled value shown there for the same filter.
    recon = {"gross": 0.0, "rto": 0.0, "cancelled": 0.0}
    # How much of the net Selling Price got a real cost vs an estimate, by value.
    val_tier = {"matched": 0.0, "approximate": 0.0, "estimated": 0.0}
    unmatched: dict[str, dict] = {}

    for r in rows:
        rev = r.get("_rev") or 0.0
        outcome = r.get("outcome")
        # -- reconciliation (all outcomes, full order-value basis) --
        recon["gross"] += rev
        if outcome == "RTO":
            recon["rto"] += rev
        elif outcome == "Cancelled":
            recon["cancelled"] += rev

        # Net P&L basis: exclude RTO / Cancelled AND the de-duplicated
        # zero-revenue duplicate lines of multi-package orders (rev == 0). This
        # puts COGS, weight and the unit count on exactly the same basis as the
        # revenue sum, so Sum(SP) == Sum(_rev) over the counted lines.
        if outcome in ("RTO", "Cancelled") or rev <= 0:
            continue
        sp = rev
        entry, level = _match(r, cogs)
        match_counts[level] = match_counts.get(level, 0) + 1
        pc = entry["product_cost"] if entry else 0.0
        cons = entry["consumables"] if entry else 0.0
        wt = _weight_for(r, sku2vol, default_wt)

        cat = r.get("category") or "Unknown"
        # Bucket the value by how the cost was sourced (exact/category/estimated).
        # An SKU, product-name or product-category hit is an exact match.
        if level in ("sku", "name", "pcat"):
            tier = "matched"
        elif level == "category":
            tier = "approximate"
        else:                                   # blended / none
            tier = "estimated"
        val_tier[tier] += sp

        # Non-exact lines are EXCLUDED from the P&L. Record them (with the reason)
        # so the reconciliation can list exactly what to add to the COGS master.
        if tier != "matched":
            key = (r.get("item_name") or r.get("subcategory") or cat or "Unknown")
            u = unmatched.get(key)
            if u is None:
                u = {"product": key, "category": cat, "tier": tier,
                     "value": 0.0, "units": 0}
                unmatched[key] = u
            u["value"] += sp
            u["units"] += 1
            continue

        # ---- exact-match line: include in the P&L, grouped per product --------
        tot["sp"] += sp
        tot["units"] += 1
        tot["pc"] += pc
        tot["cons"] += cons
        tot["wt"] += wt

        # Product identity for the breakdown: the item name for SKU/name matches
        # (the actual product shipped), the product category for pcat matches.
        pkey = (r.get("subcategory") if level == "pcat"
                else (r.get("item_name") or r.get("sku") or r.get("subcategory") or cat))
        pkey = (pkey or "Unknown").strip()
        pr = prods.get(pkey)
        if pr is None:
            gst = entry["gst"] if entry else _gst_for(cat, r.get("subcategory"),
                                                      r.get("item_name"))
            pr = {"product": pkey, "category": cat, "sp": 0.0, "units": 0,
                  "pc": 0.0, "cons": 0.0, "wt": 0.0, "gst": gst}
            prods[pkey] = pr
        pr["sp"] += sp
        pr["units"] += 1
        pr["pc"] += pc
        pr["cons"] += cons
        pr["wt"] += wt

        # Category rollup (same exact-match scope as the product breakdown).
        cg = catg.get(cat)
        if cg is None:
            cg = {"category": cat, "sp": 0.0, "units": 0, "pc": 0.0, "cons": 0.0,
                  "wt": 0.0, "gst_val": {}}
            catg[cat] = cg
        cg["sp"] += sp
        cg["units"] += 1
        cg["pc"] += pc
        cg["cons"] += cons
        cg["wt"] += wt
        gst_here = pr["gst"]
        cg["gst_val"][gst_here] = cg["gst_val"].get(gst_here, 0.0) + sp

    overall = _pnl_from_totals(tot["sp"], tot["units"], tot["pc"], tot["cons"],
                               tot["wt"])
    # Price every product/category against the blended actual EBITDA margin.
    blended_margin = None
    if overall["ebitda_margin_pct"] is not None:
        blended_margin = overall["ebitda_margin_pct"] / 100.0

    cat_rows = []
    for cg in catg.values():
        block = _pnl_from_totals(cg["sp"], cg["units"], cg["pc"], cg["cons"],
                                 cg["wt"], ebitda_margin_for_price=blended_margin)
        block["category"] = cg["category"]
        # Category GST = the rate covering the most value in that category.
        block["gst_rate"] = (max(cg["gst_val"], key=cg["gst_val"].get)
                             if cg["gst_val"] else None)
        cat_rows.append(block)
    cat_rows.sort(key=lambda x: -x["selling_price"])

    prod_rows = []
    for pr in prods.values():
        block = _pnl_from_totals(pr["sp"], pr["units"], pr["pc"], pr["cons"],
                                 pr["wt"], ebitda_margin_for_price=blended_margin)
        block["product"] = pr["product"]
        block["category"] = pr["category"]
        block["gst_rate"] = pr["gst"]
        prod_rows.append(block)
    prod_rows.sort(key=lambda x: -x["selling_price"])
    products_total = len(prod_rows)
    prod_rows = prod_rows[:PRODUCT_LIMIT]

    # Percentages are relative to the full NET order value (all products), so the
    # three coverage tiers add up to 100%.
    net_val = val_tier["matched"] + val_tier["approximate"] + val_tier["estimated"]

    def _vpct(v):
        return round(v / net_val * 100.0, 1) if net_val else None

    top_unmatched = sorted(unmatched.values(), key=lambda u: -u["value"])[:15]
    for u in top_unmatched:
        u["value"] = round(u["value"], 2)
        u["value_pct"] = _vpct(u["value"])

    reconciliation = {
        # Should tie out to the efficiency Revenue cards for the same filter.
        "gross_order_value": round(recon["gross"], 2),
        "rto_value": round(recon["rto"], 2),
        "cancelled_value": round(recon["cancelled"], 2),
        "net_order_value": round(net_val, 2),
        # What the P&L is built on = the exact-match subset only.
        "pnl_selling_price": round(tot["sp"], 2),
        "pnl_included": round(val_tier["matched"], 2),
        "pnl_excluded": round(val_tier["approximate"] + val_tier["estimated"], 2),
        # Share of the net order value whose COGS is exact vs approximate vs
        # estimated (by value); only the exact share feeds the P&L above.
        "value_matched": round(val_tier["matched"], 2),
        "value_approximate": round(val_tier["approximate"], 2),
        "value_estimated": round(val_tier["estimated"], 2),
        "value_matched_pct": _vpct(val_tier["matched"]),
        "value_approximate_pct": _vpct(val_tier["approximate"]),
        "value_estimated_pct": _vpct(val_tier["estimated"]),
        "top_unmatched": top_unmatched,
    }

    return {
        "configured": configured,
        "overall": overall,
        "categories": cat_rows,
        "products": prod_rows,
        "products_total": products_total,
        "product_limit": PRODUCT_LIMIT,
        "match_counts": match_counts,
        "coverage_pct": _vpct(val_tier["matched"]),
        "reconciliation": reconciliation,
        "cogs_rows": cogs["rows"],
        "cogs_sku_count": cogs.get("sku_count", 0),
        "cogs_updated_at": cogs["updated_at"],
        "rates": {
            "direct_labour_pct": DIRECT_LABOUR_PCT * 100,
            "logistics_per_kg": LOGISTICS_RATE_PER_KG,
            "incentive_pct": INCENTIVE_PCT * 100,
            "travel_pct": TRAVEL_PCT * 100,
            "employee_pct": EMPLOYEE_PCT * 100,
            "office_pct": OFFICE_PCT * 100,
            "pg_finance_pct": PG_FINANCE_PCT * 100,
        },
    }


def _weight_for(record, sku2vol, default_wt):
    """Billable weight (kg) for a shipment: item-master volumetric weight keyed by
    SKU, else the master's mean weight so logistics cost is never zero."""
    raw = (record.get("sku") or "").strip()
    if raw and sku2vol:
        for tok in re.split(r"[,;|/]+", raw):
            w = sku2vol.get(tok.strip().upper())
            if w:
                return float(w)
    return default_wt


def _mean_weight(sku2vol):
    if not sku2vol:
        return 0.0
    vals = [v for v in sku2vol.values() if v]
    return (sum(vals) / len(vals)) if vals else 0.0
