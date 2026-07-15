import csv
import datetime
import io
import json
import logging
import os
import re
import urllib.error
import urllib.request

from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from . import auth, bq, invoices, tat
from .kpi import build_report, filter_records, parse_workbook, reclassify

logger = logging.getLogger(__name__)

# Persisted default item master (SKU -> volumetric/billable weight + category)
# used for the weight-dispute comparison, so it doesn't have to be re-uploaded
# each session. Replaceable via the "Update item master" upload (master_config).
_MASTER_PATH = os.path.join(os.path.dirname(__file__), "data", "item_master.xlsx")
_MASTER_CACHE = {"mtime": None, "maps": None}


def _default_master():
    """Parse the persisted item master (cached by file mtime) into
    {awb2cat, sku2cat, sku2vol, sku_count, updated_at}. Returns empty maps when
    no master file is present."""
    empty = {"awb2cat": {}, "sku2cat": {}, "sku2vol": {}, "sku_count": 0, "updated_at": None}
    try:
        st = os.stat(_MASTER_PATH)
    except OSError:
        _MASTER_CACHE["mtime"] = None
        _MASTER_CACHE["maps"] = None
        return empty
    if _MASTER_CACHE["maps"] is not None and _MASTER_CACHE["mtime"] == st.st_mtime:
        return _MASTER_CACHE["maps"]
    maps = dict(empty)
    try:
        with open(_MASTER_PATH, "rb") as fh:
            data = fh.read()
        kind, payload = invoices.ingest(data, os.path.basename(_MASTER_PATH))
        if kind == "master":
            maps["awb2cat"] = payload.get("awb2cat", {})
            maps["sku2cat"] = payload.get("sku2cat", {})
            maps["sku2vol"] = payload.get("sku2vol", {})
    except Exception:  # noqa: BLE001 - a bad master file must not break the report
        logger.exception("Default item master parse failed")
    maps["sku_count"] = len(maps["sku2vol"])
    maps["updated_at"] = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%d %b %Y %H:%M")
    _MASTER_CACHE["mtime"] = st.st_mtime
    _MASTER_CACHE["maps"] = maps
    return maps

def _apply_master_categories(records):
    """Override each shipment's category/subcategory from the item master
    (SKU -> Frido category), which is the authoritative product taxonomy. Falls
    back to the existing name-derived category when the SKU isn't in the master.
    Called once when a fresh record set is loaded (BigQuery or upload)."""
    dm = _default_master()
    sku2cat = dict(dm.get("sku2cat") or {})
    sku2cat.update(_INVOICE_CACHE.get("sku2cat") or {})
    if not sku2cat:
        return
    for r in records:
        raw = (r.get("sku") or "").strip()
        if not raw:
            continue
        # A shipment line can list several SKUs; use the first one the master
        # recognises so the category reflects the actual product shipped.
        for tok in re.split(r"[,;|/]+", raw):
            key = tok.strip().upper()
            hit = sku2cat.get(key)
            if hit:
                cat, name = hit[0], hit[1]
                if cat and cat != "Others":
                    r["category"] = cat
                    r["subcategory"] = name or r.get("subcategory") or cat
                break


# client can re-filter without re-uploading. Single-process dev use.
# "window" is the date range the loaded data covers (BigQuery load window),
# or None for uploaded files; it persists across re-filter requests.
# "daily_orders" is the Unicommerce-sourced "orders placed per day" series for
# the loaded window (None to fall back to the ClickPost-derived daily series,
# e.g. for uploaded files). See _apply_uc_overrides. O2S is made accurate a
# different way — by recomputing each record's o2s on the Unicommerce order time
# at load (see _recompute_o2s_from_uc) — so it flows through every table, not
# just the headline.
_CACHE = {"records": None, "window": None, "daily_orders": None}

# Cache of the latest parsed invoice line items + master enrichment maps
# (Carrier Cost Analysis). awb2cat / sku2cat come from uploaded master files
# (weights / SKU masters) and are used to recover category/SKU on invoices that
# don't carry them.
_INVOICE_CACHE = {"items": [], "files": [], "awb2cat": {}, "sku2cat": {},
                  "version": 0, "report_cache": {},
                  # BigQuery product enrichment fetched by invoice date window
                  # (awb/order -> (category, subcategory, sku, item_name)).
                  "awb2prod": {}, "order2prod": {}, "prod_window": None,
                  # BigQuery declared order value (selling price) by awb/order,
                  # so invoice shipping cost can be shown as a % of item value.
                  "awb2value": {}, "order2value": {}, "awb2order": {},
                  # Item master SKU -> volumetric/billable weight (kg), from an
                  # uploaded master file, for weight over-charge detection.
                  "sku2vol": {}}

# Cached lane maps (AWB/order -> lane), rebuilt only when the shipment record
# set changes — not on every invoice filter/refresh.
_LANE_CACHE = {"records": None, "maps": None}

# Cached product maps (AWB/order -> (category, subcategory, sku, item_name)),
# built from the loaded BigQuery shipment data so invoice AWBs can be enriched
# with sub-category / SKU. Rebuilt only when the record set changes.
_PROD_CACHE = {"records": None, "maps": None}

# Cached value maps (AWB/order -> declared order value = selling price), built
# from the loaded shipment data as a fallback when no invoice-date enrichment
# has been fetched. Rebuilt only when the record set changes.
_VALUE_CACHE = {"records": None, "maps": None}


def _reset_invoice_cache():
    _INVOICE_CACHE["version"] += 1
    _INVOICE_CACHE["report_cache"] = {}
    _INVOICE_CACHE["prod_window"] = None
    for key in ("items", "files"):
        _INVOICE_CACHE[key] = []
    for key in ("awb2cat", "sku2cat", "sku2vol",
                "awb2prod", "order2prod", "awb2value", "order2value", "awb2order"):
        _INVOICE_CACHE[key] = {}


def _refresh_invoice_product_map():
    """Search BigQuery by the uploaded invoices' service month and cache an
    AWB/order -> product (category, sub-category, SKU, item) map for enrichment.

    The window is derived from the invoice dates (service month) padded a little
    each side, so the shipments that the invoices bill are captured regardless of
    any separate BigQuery lookback the user has loaded. No-op when BigQuery isn't
    configured or no invoice carries a service-month/date."""
    from datetime import date as _date, timedelta as _td

    items = _INVOICE_CACHE["items"]
    if not (items and bq.is_configured()):
        return
    keys = sorted({it.get("service_month_key") for it in items
                   if it.get("service_month_key")})
    if not keys:
        return

    def _month_start(k):
        return _date(int(k[:4]), int(k[5:7]), 1)

    def _month_end(k):
        y, m = int(k[:4]), int(k[5:7])
        nxt = _date(y + 1, 1, 1) if m == 12 else _date(y, m + 1, 1)
        return nxt - _td(days=1)

    lo = _month_start(keys[0]) - _td(days=15)
    hi = _month_end(keys[-1]) + _td(days=45)
    # Only the AWBs actually on the invoices need enriching — pass them so the
    # BigQuery loader skips category derivation for every other shipment in the
    # window (most of the per-row cost).
    inv_awbs = {it["awb"] for it in items if it.get("awb")}
    inv_carriers = {it.get("carrier") for it in items if it.get("carrier")}
    try:
        # use_arrow=False -> stream rows instead of materialising the whole window
        # through pyarrow, which was segfaulting the serverless worker.
        awb2prod, order2prod, awb2value, order2value, awb2order = bq.fetch_awb_product_map(
            lo.isoformat(), hi.isoformat(), awbs=inv_awbs, carriers=inv_carriers,
            use_arrow=False)
    except Exception:  # noqa: BLE001 - enrichment is best-effort; never block upload
        logger.exception("Invoice BigQuery enrichment failed")
        return
    _INVOICE_CACHE["awb2prod"] = awb2prod
    _INVOICE_CACHE["order2prod"] = order2prod
    _INVOICE_CACHE["awb2value"] = awb2value
    _INVOICE_CACHE["order2value"] = order2value
    _INVOICE_CACHE["awb2order"] = awb2order
    _INVOICE_CACHE["prod_window"] = {"from": lo.isoformat(), "to": hi.isoformat(),
                                     "awbs": len(awb2prod)}


def _invoice_product_maps():
    """Prefer the BigQuery-by-invoice-date map; fall back to the loaded shipment
    record set when no invoice-date enrichment has been fetched."""
    if _INVOICE_CACHE.get("awb2prod"):
        return _INVOICE_CACHE["awb2prod"], _INVOICE_CACHE.get("order2prod") or {}
    return _product_maps()


def _invoice_value_maps():
    """AWB/order -> declared order value (selling price). Prefer the
    BigQuery-by-invoice-date enrichment (covers the full invoice window); fall
    back to the loaded shipment record set."""
    if _INVOICE_CACHE.get("awb2value"):
        return _INVOICE_CACHE["awb2value"], _INVOICE_CACHE.get("order2value") or {}
    return _value_maps()


def _invoice_report(carrier_filter=None):
    """Build the cost report from the cached line items, optionally restricted
    to a single carrier. The full carrier list is always returned (all_carriers)
    so the frontend dropdown stays populated even while a filter is active."""
    items = _INVOICE_CACHE["items"]
    active = carrier_filter if (carrier_filter and carrier_filter != "all") else ""

    # Cache the built report by (carrier, invoice version, shipment-set id) so a
    # plain refresh or repeated carrier filter is an instant cache hit.
    ckey = (active, _INVOICE_CACHE["version"], id(_CACHE.get("records")))
    cached = _INVOICE_CACHE["report_cache"].get(ckey)
    if cached is not None:
        return cached

    # Full carrier list (by spend desc) — drives the dropdown options.
    spend_by = {}
    for it in items:
        spend_by[it["carrier"]] = spend_by.get(it["carrier"], 0.0) + it["amount"]
    all_carriers = [c for c, _ in sorted(spend_by.items(), key=lambda kv: -kv[1])]

    selected = items
    if active:
        selected = [it for it in items if it["carrier"] == active]

    awb2lane, order2lane = _lane_maps()
    awb2prod, order2prod = _invoice_product_maps()
    awb2value, order2value = _invoice_value_maps()
    # Merge the persisted default item master with any master uploaded this
    # session (session entries override), so weight-dispute comparison works
    # without re-uploading the master every time.
    dm = _default_master()

    # Merge the persisted default master with any session-uploaded master
    # (session entries override). The masters can be large (thousands of SKUs),
    # so only copy when there IS a session override; otherwise reuse the cached
    # default dict directly (build_cost_report treats these maps read-only).
    def _eff_master(key):
        base = dm.get(key) or {}
        override = _INVOICE_CACHE.get(key) or {}
        if not override:
            return base
        merged = dict(base)
        merged.update(override)
        return merged
    eff_awb2cat = _eff_master("awb2cat")
    eff_sku2cat = _eff_master("sku2cat")
    eff_sku2vol = _eff_master("sku2vol")
    report = invoices.build_cost_report(
        selected, _INVOICE_CACHE["files"],
        eff_awb2cat, eff_sku2cat,
        awb2lane=awb2lane, order2lane=order2lane,
        awb2prod=awb2prod, order2prod=order2prod,
        awb2value=awb2value, order2value=order2value,
        awb2order=_INVOICE_CACHE.get("awb2order") or {},
        sku2vol=eff_sku2vol)
    report["master"] = {
        "sku_count": len(eff_sku2vol),
        "updated_at": dm["updated_at"],
        "source": "uploaded" if _INVOICE_CACHE.get("sku2vol") else "default",
    }
    report["all_carriers"] = all_carriers
    report["carrier_filter"] = active
    report["prod_window"] = _INVOICE_CACHE.get("prod_window")
    # Cross-carrier comparison is built from ALL loaded invoices (not the
    # filtered subset) so it stays the same regardless of the carrier dropdown.
    report["carrier_comparison"] = invoices.build_carrier_comparison(items)
    _INVOICE_CACHE["report_cache"][ckey] = report
    return report


def _shipment_maps(cache, value_for):
    """Build (awb_map, order_map) from the loaded shipment records, caching by
    record-set identity so tens of thousands of rows aren't rescanned on every
    invoice filter/refresh. `value_for(r)` returns the value to store, or None to
    skip the row. Uses the invoice parser's AWB normalization so keys line up."""
    recs = _CACHE.get("records") or []
    if cache["records"] is recs and cache["maps"] is not None:
        return cache["maps"]
    awb_map, order_map = {}, {}
    for r in recs:
        val = value_for(r)
        if val is None:
            continue
        awb = r.get("awb")
        if awb:
            awb_map[invoices._norm_awb(awb)] = val
        oid = (r.get("order_id") or "").strip().upper()
        if oid:
            order_map.setdefault(oid, val)
    cache["records"] = recs
    cache["maps"] = (awb_map, order_map)
    return awb_map, order_map


def _lane_maps():
    def value_for(r):
        pin, drop = r.get("pickup_pin") or "", r.get("drop_pin") or ""
        return (pin, drop) if (pin and drop) else None
    return _shipment_maps(_LANE_CACHE, value_for)


def _product_maps():
    def value_for(r):
        sub = (r.get("subcategory") or "").strip()
        sku = (r.get("sku") or "").strip()
        cat = (r.get("category") or "").strip()
        name = (r.get("item_name") or "").strip()
        return (cat, sub, sku, name) if (sub or sku or cat or name) else None
    return _shipment_maps(_PROD_CACHE, value_for)


def _value_maps():
    return _shipment_maps(_VALUE_CACHE, lambda r: r.get("order_value") or None)


def login_view(request):
    """Internal-team sign-in. GET renders the form, POST validates credentials."""
    if auth.is_authenticated(request):
        return redirect("index")

    next_url = request.GET.get("next") or request.POST.get("next") or ""
    # Only allow safe local redirects.
    if not next_url.startswith("/"):
        next_url = ""

    if request.method == "POST":
        ip = auth.client_ip(request)
        if auth.is_locked_out(ip):
            mins = max(1, auth.seconds_until_unlock(ip) // 60)
            return render(request, "dashboard/login.html", {
                "error": f"Too many attempts. Try again in about {mins} minute(s).",
                "next": next_url,
            }, status=429)

        username = request.POST.get("username", "")
        password = request.POST.get("password", "")
        if auth.verify_credentials(username, password):
            auth.clear_failures(ip)
            auth.login_session(request, username.strip())
            return redirect(next_url or "index")

        auth.record_failure(ip)
        return render(request, "dashboard/login.html", {
            "error": "Invalid username or password.",
            "username": username,
            "next": next_url,
        }, status=401)

    return render(request, "dashboard/login.html", {"next": next_url})


def logout_view(request):
    auth.logout_session(request)
    return redirect("login")


@auth.team_required
def index(request):
    return render(request, "dashboard/index.html", {
        "frido_user": request.session.get(auth.SESSION_KEY, ""),
        # Lets the frontend auto-load the default window from BigQuery on entry
        # (and skip the manual upload screen) only when BQ is actually set up.
        "bq_configured": bq.is_configured(),
    })


def _filter_kwargs(request):
    """Read the multi-select filter fields from a POST into build_report kwargs.

    Each categorical filter is a multi-select checkbox dropdown, so each arrives
    as zero or more repeated form fields. getlist() collects them; an empty list
    falls back to the default ("all" = no constraint). Delivery type also
    defaults to "all" so both forward and reverse-pickup shipments are visible
    unless the user narrows it.
    """
    return {
        "delivery_type": request.POST.getlist("delivery_type") or "all",
        "tier": request.POST.getlist("tier") or "all",
        "channel": request.POST.getlist("channel") or "all",
        "zone": request.POST.getlist("zone") or "all",
        "payment": request.POST.getlist("payment") or "all",
        "warehouse": request.POST.getlist("warehouse") or "all",
        "account": request.POST.getlist("account") or "all",
        "weight": request.POST.getlist("weight") or "all",
        "slot": request.POST.getlist("slot") or "all",
        "date_from": request.POST.get("date_from", ""),
        "date_to": request.POST.get("date_to", ""),
    }


# Keys that only powered the (removed) "Top lanes" panel. The full lane list can
# be thousands of rows, so it's stripped from the browser payload to keep the
# dashboard-load response lean. build_report still computes lanes for the AI
# summary (which rebuilds its own report), so nothing downstream breaks.
_CLIENT_DROP_KEYS = ("lanes", "lane_total", "lane_top_n", "lane_min_n")


def _client_report(report):
    """Shallow copy of the report with client-unused heavy keys removed.
    Copies rather than mutating in place because build_report caches and shares
    the dict (the AI-summary path reads the same object)."""
    return {k: v for k, v in report.items() if k not in _CLIENT_DROP_KEYS}


def _apply_uc_overrides(client_report):
    """Swap the Orders-per-day series to the Unicommerce "orders placed per day"
    data when a BigQuery window is loaded (order count complete at placement;
    the ClickPost series undercounts recent unshipped days). This is window-level
    and independent of the categorical filters. Safe to mutate: _client_report
    returns a fresh dict.

    (O2S is handled separately — recomputed onto every record at load time in
    _recompute_o2s_from_uc — so the headline AND the per-carrier/warehouse O2S
    columns are all on the same accurate order-received -> pickup basis.)"""
    daily = _CACHE.get("daily_orders")
    if daily is not None:
        client_report["daily"] = daily
    return client_report


def _o2s_mature_days():
    """Orders younger than this (days) are excluded from O2S — they may not have
    shipped yet, so counting them would bias O2S low. Env BQ_O2S_MATURE_DAYS (4)."""
    try:
        return max(0, int(os.environ.get("BQ_O2S_MATURE_DAYS")))
    except (TypeError, ValueError):
        return 4


def _recompute_o2s_from_uc(records, order_times):
    """Recompute each record's O2S (order->pickup, hours) from the accurate
    Unicommerce order-received timestamp instead of ClickPost's date-only order
    date, so every O2S figure (headline + per-carrier/warehouse tables) shares one
    basis. o2s is set to None — i.e. excluded from the averages — when there's no
    OMS match, no pickup time, a negative gap, or the order is too recent to have
    reliably shipped (younger than the maturity cut)."""
    import datetime as _dt
    cutoff = _dt.date.today() - _dt.timedelta(days=_o2s_mature_days())
    for r in records:
        oid = (r.get("order_id") or "").strip().upper()
        uc_iso = order_times.get(oid)
        pk_iso = r.get("_pickup_ts")
        if not uc_iso or not pk_iso:
            r["o2s"] = None
            continue
        try:
            uc = _dt.datetime.fromisoformat(uc_iso)
            pk = _dt.datetime.fromisoformat(pk_iso)
        except (TypeError, ValueError):
            r["o2s"] = None
            continue
        hrs = (pk - uc).total_seconds() / 3600.0
        r["o2s"] = round(hrs, 2) if (hrs >= 0 and uc.date() <= cutoff) else None


def _report_response(request, empty_msg):
    """Build the report from the cached records using the request's filters."""
    records = _CACHE["records"]
    if not records:
        return JsonResponse({"error": empty_msg}, status=400)
    report = build_report(records, **_filter_kwargs(request))
    # The single authoritative date range for the loaded data (BigQuery load
    # window). None for uploaded files; the frontend falls back to pickup span.
    report["load_window"] = _CACHE.get("window")
    # Orders-per-day and O2S come from Unicommerce (complete/real-time) when a
    # BigQuery window is loaded; otherwise the ClickPost-derived values stand.
    return JsonResponse(_apply_uc_overrides(_client_report(report)))


# ---------------------------------------------------------------------------
# AI business summary (OpenAI) — a short, crisp exec overview of the selected
# date range covering both what's working and what needs attention. The API key
# lives server-side in OPENAI_API_KEY and is never exposed to the browser.
# ---------------------------------------------------------------------------
def _inr(value):
    """Format a rupee amount the Indian way (Cr / L) as a display string, so the
    model never has to convert raw integers into crores itself — it was reliably
    off by 10x (dividing by 1e6 instead of 1e7). 1 Cr = 1e7, 1 L = 1e5.
    """
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    a = abs(v)
    if a >= 1e7:
        return f"₹{v / 1e7:.2f}Cr"
    if a >= 1e5:
        return f"₹{v / 1e5:.2f}L"
    return f"₹{v:,.0f}"


def _score_component(value, floor, ceil, higher_is_better=True):
    """Map a raw metric onto 0-100 using linear anchors.

    ``floor``/``ceil`` are the raw values that map to the worst/best ends of the
    scale. When ``higher_is_better`` is False the direction is inverted (e.g. RTO
    %, where a low value is good). Returns None when ``value`` is missing so the
    component can be dropped and the remaining weights renormalised.
    """
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if higher_is_better:
        pct = (v - floor) / (ceil - floor) if ceil != floor else 0.0
    else:
        pct = (floor - v) / (floor - ceil) if floor != ceil else 0.0
    return round(max(0.0, min(1.0, pct)) * 100, 1)


def _grade(score):
    """A→F letter grade + short verdict for a 0-100 score."""
    if score is None:
        return "N/A", "insufficient data"
    if score >= 85:
        return "A", "excellent"
    if score >= 75:
        return "B", "healthy"
    if score >= 62:
        return "C", "mixed — watch the weak spots"
    if score >= 48:
        return "D", "under strain"
    return "F", "critical — needs intervention"


def _health_score(report):
    """Deterministic, reproducible 0-100 business-health score for the loaded
    slice, so the AI narrative is anchored on the same number every time rather
    than a value the model invents. Blends the five dimensions the dashboard
    actually measures; any dimension with missing inputs is dropped and the
    remaining weights are renormalised so partial data still yields a fair score.
    """
    s = report.get("summary", {}) or {}
    pay = report.get("payment_perf", []) or []

    # COD return risk, if a COD payment group is present.
    cod_rto = None
    for p in pay:
        grp = (p.get("group") or "").lower()
        if "cod" in grp or "cash" in grp:
            cod_rto = p.get("rto_pct")
            break

    # (label, weight, sub-score 0-100). Anchors reflect Indian D2C norms.
    comps = [
        # Delivered / picked. 70% is poor, 95% is best-in-class.
        ("Delivery success", 0.30,
         _score_component(s.get("success_rate"), 70, 95, higher_is_better=True)),
        # RTO rate. 5% is excellent, 30% is alarming for D2C.
        ("RTO / returns control", 0.25,
         _score_component(s.get("rto_pct"), 30, 5, higher_is_better=False)),
        # In-TAT %. 60% is weak, 95% is strong SLA adherence.
        ("On-time / SLA (TAT)", 0.20,
         _score_component(s.get("tat_in_pct"), 60, 95, higher_is_better=True)),
        # Realised revenue = delivered value / booked revenue.
        ("Revenue realisation", 0.15,
         _score_component(
             (s.get("delivered_value") / s.get("revenue") * 100)
             if s.get("revenue") else None,
             55, 90, higher_is_better=True)),
        # COD return exposure. 5% good, 30% risky. Skipped if no COD data.
        ("COD risk", 0.10,
         _score_component(cod_rto, 30, 5, higher_is_better=False)),
    ]

    present = [(label, w, v) for (label, w, v) in comps if v is not None]
    if not present:
        return None
    wsum = sum(w for _, w, _ in present)
    overall = round(sum(w * v for _, w, v in present) / wsum, 1)
    grade, verdict = _grade(overall)
    return {
        "overall": overall,
        "grade": grade,
        "verdict": verdict,
        "scale": "0-100 (higher is better)",
        "components": [
            {"dimension": label, "weight_pct": round(w / wsum * 100),
             "sub_score": v}
            for (label, w, v) in present
        ],
    }


def _ai_metrics(report, window):
    """Compact, model-friendly slice of the report for the summary prompt."""
    s = report.get("summary", {})
    pay = report.get("payment_perf", []) or []
    carriers = report.get("carriers", []) or []
    prods = report.get("products", []) or []
    whs = report.get("warehouses", []) or []
    dests = report.get("destinations", []) or []
    lanes = report.get("lanes", []) or []
    daily = report.get("daily", []) or []
    if window and window.get("from"):
        dr = {"from": window.get("from"), "to": window.get("to")}
    else:
        f = report.get("filters", {}) or {}
        dr = {"from": f.get("date_min"), "to": f.get("date_max")}
    busiest = max(daily, key=lambda d: d["n"], default=None) if daily else None
    lightest = min(daily, key=lambda d: d["n"], default=None) if daily else None

    revenue = s.get("revenue")
    # Carriers already arrive ranked by efficiency score; surface both ends so the
    # model can call out standouts AND laggards, not just the top of the list.
    ranked = [c for c in carriers if c.get("score") is not None]
    weak_carriers = sorted(ranked, key=lambda c: c["score"])[:3] if ranked else []
    # Product categories carrying the most return risk (min volume to be fair).
    risky_prods = sorted(
        [p for p in prods if p.get("rto_pct") is not None and (p.get("n") or 0) >= 20],
        key=lambda p: -p["rto_pct"])[:4]

    # Warehouse & destination breakdowns: surface both the volume leaders AND the
    # weakest performers (by delivery success, above a min volume so a 3-order
    # pincode can't top the "worst" list). Mirrors what those dashboard panels
    # rank on so the summary reflects the same insight without the reader opening
    # the tables. Min-volume gate reuses the panel's own threshold.
    wh_min = report.get("warehouse_min_n") or 20
    dest_min = report.get("warehouse_min_n") or 20
    wh_scorable = [w for w in whs if w.get("success_rate") is not None and (w.get("n") or 0) >= wh_min]
    weak_whs = sorted(wh_scorable, key=lambda w: w["success_rate"])[:3]
    dest_scorable = [d for d in dests if d.get("success_rate") is not None and (d.get("n") or 0) >= dest_min]
    weak_dests = sorted(dest_scorable, key=lambda d: d["success_rate"])[:4]
    # Best-selling individual products (subcategory level) across all categories.
    subs = []
    for p in prods:
        for sub in (p.get("subs") or []):
            subs.append({"product": sub.get("subcategory"), "category": p.get("category"),
                         "revenue": sub.get("revenue"), "orders": sub.get("n"),
                         "rto_pct": sub.get("rto_pct")})
    top_products = sorted([x for x in subs if x.get("revenue") is not None],
                          key=lambda x: -x["revenue"])[:5]

    return {
        "currency": "INR",
        "date_range": dr,
        # Deterministic health score the narrative must be built around.
        "health_score": _health_score(report),
        # ---- Volume & revenue -------------------------------------------------
        # NOTE: all rupee amounts are PRE-FORMATTED strings (₹X.XCr / ₹X.XL). Quote
        # them verbatim — do not reconvert. Percentages are separate numeric fields.
        "orders": s.get("total"),
        "delivered_orders": s.get("delivered"),
        "delivery_success_pct": s.get("success_rate"),
        "revenue": _inr(revenue),
        "avg_order_value": _inr(s.get("aov")),
        "delivered_value": _inr(s.get("delivered_value")),
        "delivered_value_pct_of_revenue": (
            round(s.get("delivered_value") / revenue * 100, 1)
            if revenue and s.get("delivered_value") is not None else None),
        # ---- Revenue leakage (value stuck / lost) -----------------------------
        "rto_orders": s.get("rto"), "rto_pct": s.get("rto_pct"),
        "rto_value": _inr(s.get("rto_value")), "rto_value_pct": s.get("rto_value_pct"),
        "forward_pending_orders": s.get("pending_orders"),
        "forward_pending_value": _inr(s.get("pending_value")),
        "forward_pending_value_pct": s.get("pending_value_pct"),
        "cancelled_orders": s.get("cancelled_orders"),
        "cancelled_value": _inr(s.get("cancelled_value")),
        "cancelled_value_pct": s.get("cancelled_value_pct"),
        # ---- Speed & SLA ------------------------------------------------------
        "in_tat_pct": s.get("tat_in_pct"), "out_of_tat_pct": s.get("tat_out_pct"),
        "out_of_tat_delivered_late": s.get("tat_out_delivered"),
        "out_of_tat_pending_breach": s.get("tat_out_pending"),
        "avg_pickup_to_delivery_hrs": s.get("avg_p2d"),
        "avg_pickup_to_order_hrs": s.get("avg_p2o"),
        "avg_order_to_ship_hrs": s.get("avg_o2s"),
        "ndd_orders": s.get("ndd_orders"), "ndd_pct": s.get("ndd_pct"),
        # ---- Mix --------------------------------------------------------------
        "carrier_count": s.get("carriers"), "warehouse_count": s.get("warehouses"),
        "city_tier_orders": s.get("tiers"),
        "city_tier_revenue": {k: _inr(v) for k, v in (s.get("tier_revenue") or {}).items()},
        "payment_modes": [
            {"mode": p.get("group"), "orders": p.get("n"), "revenue": _inr(p.get("revenue")),
             "success_pct": p.get("success_rate"), "rto_pct": p.get("rto_pct"),
             "cod_cash_in_transit": _inr(p.get("cod_exposure"))}
            for p in pay
        ],
        # ---- Partners & regions ----------------------------------------------
        "top_carriers": [
            {"carrier": c.get("carrier"), "score": c.get("score"), "orders": c.get("n"),
             "success_pct": c.get("success_rate"), "rto_pct": c.get("rto_pct"),
             "pickup_to_delivery_hrs": c.get("p2d")}
            for c in carriers[:6]
        ],
        "weakest_carriers": [
            {"carrier": c.get("carrier"), "score": c.get("score"), "orders": c.get("n"),
             "success_pct": c.get("success_rate"), "rto_pct": c.get("rto_pct")}
            for c in weak_carriers
        ],
        # ---- Product breakdown (categories ranked by revenue; RTO = return risk)
        "product_categories_count": len([p for p in prods if p.get("category")]),
        "top_product_categories": [
            {"category": p.get("category"), "revenue": _inr(p.get("revenue")),
             "orders": p.get("n"), "rto_pct": p.get("rto_pct")}
            for p in prods[:6]
        ],
        "best_selling_products": [
            {"product": x["product"], "category": x["category"],
             "revenue": _inr(x["revenue"]), "orders": x["orders"], "rto_pct": x["rto_pct"]}
            for x in top_products
        ],
        "highest_return_categories": [
            {"category": p.get("category"), "orders": p.get("n"), "rto_pct": p.get("rto_pct")}
            for p in risky_prods
        ],
        # ---- Warehouse breakdown (pickup points; score blends success/speed/attempts)
        "top_warehouses": [
            {"warehouse": w.get("warehouse") or w.get("pickup_pin"), "orders": w.get("n"),
             "score": w.get("score"), "success_pct": w.get("success_rate"),
             "first_attempt_pct": w.get("first_attempt_rate"),
             "pickup_to_delivery_hrs": w.get("p2d")}
            for w in whs[:5]
        ],
        "weakest_warehouses": [
            {"warehouse": w.get("warehouse") or w.get("pickup_pin"), "orders": w.get("n"),
             "score": w.get("score"), "success_pct": w.get("success_rate")}
            for w in weak_whs
        ],
        # ---- Destination rollup (by drop city; where orders land & how they fare)
        "top_destinations": [
            {"city": d.get("drop_city"), "orders": d.get("n"),
             "success_pct": d.get("success_rate"),
             "first_attempt_pct": d.get("first_attempt_rate")}
            for d in dests[:5]
        ],
        "weakest_destinations": [
            {"city": d.get("drop_city"), "orders": d.get("n"),
             "success_pct": d.get("success_rate")}
            for d in weak_dests
        ],
        "top_lanes": [
            {"lane": ln.get("lane"), "orders": ln.get("n"),
             "success_pct": ln.get("success_rate"), "rto_pct": ln.get("rto_pct")}
            for ln in lanes[:5]
        ],
        # ---- Tempo ------------------------------------------------------------
        "busiest_day": ({"date": busiest["date"], "orders": busiest["n"]} if busiest else None),
        "lightest_day": ({"date": lightest["date"], "orders": lightest["n"]} if lightest else None),
    }


_AI_SYSTEM = (
    "You are the operations analyst for Frido, an Indian D2C e-commerce brand. You "
    "receive KPI metrics (JSON) for a selected date range from a logistics "
    "dashboard. Your job: write a self-contained executive briefing so a manager "
    "who has NOT scrolled the dashboard fully understands how the business is "
    "performing and why. Assume the reader sees ONLY your summary.\n\n"
    "IMPORTANT — money: every rupee amount in the JSON is ALREADY formatted as a "
    "string (e.g. '₹4.59Cr', '₹77.85L', '₹1,819'). Quote these strings EXACTLY as "
    "given. Never recompute, rescale, round or convert them, and never turn a raw "
    "number into Cr/L yourself. Percentages are given as separate numeric fields — "
    "pair a % with its amount. Use ONLY values present in the JSON; never invent or "
    "estimate figures. Omit a point if its data is null.\n\n"
    "Structure your answer in this exact order using markdown:\n"
    "1. '**Overall health score: <overall>/100 (<grade> — <verdict>)**' on its own "
    "line. Copy the score, grade and verdict verbatim from the provided "
    "`health_score` object; do NOT compute your own.\n"
    "2. '**Bottom line:**' — 2-3 sentences summarising the state of the business: "
    "scale (orders + revenue), how much of that revenue is actually being realised "
    "vs stuck/lost, and the single biggest strength and biggest risk.\n"
    "3. '## Score drivers' — one '- ' bullet per component in `health_score."
    "components`, each as '**<dimension>: <sub_score>/100** — <one-line reason "
    "citing the underlying metric>'. This explains WHY the score is what it is.\n"
    "4. '## What's working' — 3-4 bullets on genuine strengths (strong carriers, "
    "high-success regions, healthy TAT, low-return categories, revenue realised).\n"
    "5. '## What needs attention' — 3-4 bullets on the real risks, ordered by "
    "rupee/impact: RTO value, COD cash exposure, TAT breaches, weak carriers, "
    "high-return categories, pending/cancelled value.\n"
    "6. '## Product, warehouse & destination insights' — exactly 3 bullets, one "
    "each: (a) PRODUCTS — the top revenue category/product ("
    "`top_product_categories`/`best_selling_products`) vs the highest-return "
    "category (`highest_return_categories`), with revenue and RTO%; (b) WAREHOUSES "
    "— the best vs weakest pickup point (`top_warehouses` by score/success vs "
    "`weakest_warehouses`), citing success% and volume; (c) DESTINATIONS — the "
    "biggest drop cities (`top_destinations`) and any weak-delivery cities "
    "(`weakest_destinations`), citing orders and success%. Skip a clause only if "
    "its data is empty.\n\n"
    "Be specific, quantified and decision-useful. No preamble, no generic advice, "
    "no closing pleasantries. Aim for 220-320 words."
)


def _openai_summary(api_key, metrics, model="gpt-4o-mini"):
    """Call OpenAI Chat Completions via stdlib urllib (no extra dependency)."""
    body = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": _AI_SYSTEM},
            {"role": "user", "content": "KPI metrics JSON:\n" + json.dumps(metrics, ensure_ascii=False)},
        ],
        "temperature": 0.3,
        "max_tokens": 1000,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=body,
        headers={"Authorization": "Bearer " + api_key, "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "replace")[:300]
        raise RuntimeError(f"OpenAI API error {exc.code}: {detail}") from exc
    return (data["choices"][0]["message"]["content"] or "").strip()


@auth.team_required
@require_POST
def ai_summary(request):
    """Generate an AI executive summary of the current (filtered) date range."""
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return JsonResponse(
            {"error": "OpenAI API key not configured. Add OPENAI_API_KEY to your .env and restart."},
            status=400,
        )
    records = _CACHE["records"]
    if not records:
        return JsonResponse({"error": "Load data first, then generate a summary."}, status=400)
    report = build_report(records, **_filter_kwargs(request))
    metrics = _ai_metrics(report, _CACHE.get("window"))
    try:
        summary = _openai_summary(api_key, metrics)
    except Exception as exc:  # noqa: BLE001 - surface any API/network failure cleanly
        logger.exception("AI summary failed")
        return JsonResponse({"error": f"AI summary failed: {exc}"}, status=502)
    return JsonResponse({"summary": summary})


# Columns written to the shipment-level CSV export, in order: (record key, header).
_EXPORT_FIELDS = [
    ("awb", "AWB"), ("order_id", "Order ID"),
    ("carrier", "Carrier"), ("account", "Account"),
    ("delivery_type", "Delivery type"),
    ("pickup_pin", "Pickup pincode"), ("warehouse", "Warehouse"), ("city", "Pickup city"),
    ("drop_pin", "Drop pincode"), ("drop_city", "Drop city"), ("tier", "City tier"),
    ("payment", "Payment"), ("pickup_date", "Pickup date"),
    ("status", "Latest status"), ("outcome", "Outcome"), ("delivered", "Delivered"),
    ("o2s", "Order->Pickup O2S (hrs)"),
    ("p2o", "Pickup->OFD1 (hrs)"), ("p2d", "Pickup->Delivery (hrs)"),
    ("promised_tat", "Promised TAT"), ("tat_status", "TAT status"), ("tat_margin", "TAT margin"),
]


@auth.team_required
@require_POST
def export_shipments(request):
    """Download the currently-filtered shipments as CSV so the summary cards
    (Shipments / In TAT / Out of TAT) can be traced to the underlying rows.

    Honours every active filter (same as the report). An optional `tat_status`
    multi-value field limits the export to one or more compliance buckets
    ("In TAT", "Out of TAT", "No rule", "Pending")."""
    records = _CACHE["records"]
    if not records:
        return JsonResponse({"error": "Load data first, then export."}, status=400)

    rows = filter_records(records, **_filter_kwargs(request))
    status_sel = set(request.POST.getlist("tat_status"))
    if status_sel:
        rows = [r for r in rows if (r.get("tat_status") or "No rule") in status_sel]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in _EXPORT_FIELDS])
    for r in rows:
        writer.writerow([r.get(key, "") for key, _ in _EXPORT_FIELDS])

    suffix = ("_" + "_".join(sorted(status_sel)).replace(" ", "")) if status_sel else ""
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="shipments{suffix}.csv"'
    return resp


def _sla_header_index(header):
    """Locate the Warehouse-pincode / (destination) Pincode / TAT columns by
    (loose) header name. The origin column is the warehouse/pickup pincode."""
    h = [str(c or "").strip().lower() for c in (header or [])]

    def find(names, exclude=()):
        for i, c in enumerate(h):
            if i in exclude:
                continue
            if any(n in c for n in names):
                return i
        return None

    origin = find(["warehouse pin", "pickup pin", "origin pin",
                   "warehouse", "origin", "hub"])
    excl = {origin} if origin is not None else set()
    dest = find(["destination pin", "drop pin", "delivery pin", "dest pin"], exclude=excl)
    if dest is None:
        dest = find(["pincode", "pin code", "pin"], exclude=excl)
    return {
        "origin_pin": origin,
        "pincode": dest,
        "tat": find(["tat", "days", "threshold", "sla"]),
    }


def _parse_sla_rows(upload):
    """Parse an uploaded SLA file into [{warehouse, pincode, tat}] rows.

    Accepts .xlsx/.xlsm/.csv/.tsv with columns Warehouse pincode, Pincode, TAT
    (warehouse pincode optional -> defaults to 'ANY')."""
    name = (upload.name or "").lower()
    data = upload.read()
    records = []

    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        it = wb[wb.sheetnames[0]].iter_rows(values_only=True)
        rows = list(it)
    elif name.endswith((".csv", ".tsv")):
        text = data.decode("utf-8", "replace")
        delim = "\t" if name.endswith(".tsv") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    else:
        raise ValueError("Upload a .xlsx or .csv with Warehouse, Pincode, TAT columns.")

    if not rows:
        raise ValueError("The file is empty.")
    idx = _sla_header_index(rows[0])
    if idx["pincode"] is None or idx["tat"] is None:
        raise ValueError("Could not find 'Pincode' and 'TAT' columns in the header.")

    def cell(r, key):
        i = idx.get(key)
        return r[i] if (i is not None and i < len(r)) else None

    for r in rows[1:]:
        if not r:
            continue
        pin, tat = cell(r, "pincode"), cell(r, "tat")
        if pin is None or tat is None:
            continue
        records.append({"origin_pin": cell(r, "origin_pin") or "ANY",
                        "pincode": pin, "tat": tat})
    if not records:
        raise ValueError("No valid rows found (need Pincode + TAT in each row).")
    return records


@auth.team_required
@require_POST
def sla_config(request):
    """View/replace a carrier's promised-TAT (SLA) table.

    actions: list | get | save (rows JSON) | upload (file) | clear.
    A save/upload/clear re-scores the loaded data and returns a fresh report."""
    action = request.POST.get("action", "get")
    carrier = (request.POST.get("carrier") or "").strip()

    if action == "list":
        return JsonResponse({"carriers": tat.carriers_meta()})
    if action == "get":
        return JsonResponse({"carrier": carrier, "carriers": tat.carriers_meta(),
                             "rows": tat.get_override_rows(carrier)})

    try:
        if action == "clear":
            tat.clear_override(carrier)
        elif action == "save":
            rows = json.loads(request.POST.get("rows", "[]"))
            tat.set_override(carrier, rows)
        elif action == "upload":
            upload = request.FILES.get("file")
            if upload is None:
                return JsonResponse({"error": "No file uploaded."}, status=400)
            tat.set_override(carrier, _parse_sla_rows(upload))
        else:
            return JsonResponse({"error": "Unknown action."}, status=400)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    out = {"ok": True, "carrier": carrier, "carriers": tat.carriers_meta(),
           "rows": tat.get_override_rows(carrier)}
    records = _CACHE["records"]
    if records:
        reclassify(records)
        report = build_report(records, **_filter_kwargs(request))
        report["load_window"] = _CACHE.get("window")
        out["report"] = _apply_uc_overrides(_client_report(report))
    return JsonResponse(out)


@auth.team_required
@require_POST
def process_upload(request):
    """Accept a new file upload, or a re-filter request on cached data."""
    upload = request.FILES.get("file")
    if upload is not None:
        name = upload.name.lower()
        if not name.endswith((".xlsx", ".xlsm", ".csv", ".tsv")):
            return JsonResponse(
                {"error": "Please upload a .xlsx or .csv file in the standard export format."},
                status=400,
            )
        try:
            records = parse_workbook(upload, filename=upload.name)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except Exception as exc:  # noqa: BLE001 - surface any parse failure cleanly
            return JsonResponse({"error": f"Could not read file: {exc}"}, status=400)

        if not records:
            return JsonResponse(
                {"error": "No data rows found in the file."}, status=400
            )
        _apply_master_categories(records)
        _CACHE["records"] = records
        _CACHE["window"] = None  # uploaded file has no BigQuery load window
        # Uploaded files use their own (ClickPost) per-day series and O2S (there's
        # no OMS order time to recompute against), so drop any Unicommerce daily
        # override left over from a previous BigQuery load.
        _CACHE["daily_orders"] = None

    return _report_response(
        request, "No data loaded yet. Load from BigQuery or upload a file first."
    )


def _ingest_into_cache(name, data, errors):
    """Parse one invoice/master file's bytes and merge it into _INVOICE_CACHE.
    Shared by the browser upload and the Google Drive import. Appends a message
    to `errors` on failure; returns True when something was ingested."""
    if not str(name).lower().endswith((".xlsx", ".xlsm", ".xlsb", ".csv", ".tsv")):
        errors.append(f"{name}: unsupported file type")
        return None
    try:
        kind, payload = invoices.ingest(data or b"", name)
    except ValueError as exc:
        errors.append(f"{name}: {exc}")
        return None
    except Exception as exc:  # noqa: BLE001
        errors.append(f"{name}: could not read ({exc})")
        return None

    if kind == "master":
        _INVOICE_CACHE["awb2cat"].update(payload.get("awb2cat", {}))
        _INVOICE_CACHE["sku2cat"].update(payload.get("sku2cat", {}))
        _INVOICE_CACHE["sku2vol"].update(payload.get("sku2vol", {}))
        entries = (len(payload.get("awb2cat", {})) + len(payload.get("sku2cat", {}))
                   + len(payload.get("sku2vol", {})))
        _INVOICE_CACHE["files"].append(
            {"name": name, "carrier": "Reference / master", "lines": entries,
             "spend": 0, "kind": "master"})
    else:  # invoice
        spend = round(sum(i["amount"] for i in payload), 1)
        carrier = payload[0]["carrier"] if payload else "Unknown"
        # Each file is one billing period; tag every line with the month derived
        # from its file name so the cost report can compare periods.
        mkey, mlabel = invoices.month_from_filename(name)
        for it in payload:
            it["month"] = mkey
            it["month_label"] = mlabel
        _INVOICE_CACHE["items"].extend(payload)
        _INVOICE_CACHE["files"].append(
            {"name": name, "carrier": carrier, "lines": len(payload),
             "spend": spend, "kind": "invoice", "month": mlabel})
    return kind


def _sorted_months(buckets):
    """Order month buckets for display: real calendar months (YYYY-MM) newest
    first, then month-only buckets, then undated files (named without a parseable
    month) last."""
    buckets = list(buckets)

    def _is_dated(m):
        k = m["key"]
        return len(k) >= 7 and k[:2] == "20" and k[4] == "-"

    dated = sorted((m for m in buckets if _is_dated(m)),
                   key=lambda m: m["key"], reverse=True)
    other = sorted((m for m in buckets if not _is_dated(m)),
                   key=lambda m: m["key"])
    return dated + other


@auth.team_required
@require_POST
def import_from_drive(request):
    """Import invoice/master files from a shared Google Drive folder, parse them
    server-side, and return the report. Bypasses request-body size limits (the
    server downloads the files) and supports monthly auto-import.

    A month must be selected before importing: the file's billing month is
    inferred from its name (same rule the invoice parser uses), and only files
    for the chosen month are downloaded and parsed.

    Params: `folder` (optional folder ID/link; else GDRIVE_INVOICE_FOLDER env),
    `action="months"` to list the months present in the folder (no download),
    `month` (REQUIRED to import; the "YYYY-MM"-style key from the months list),
    `append=1` to add to the current set instead of replacing."""
    from . import gdrive, invoices
    logger.info("drive import: start (action=%s)", request.POST.get("action") or "import")
    if not gdrive.is_configured():
        return JsonResponse(
            {"error": "Google Drive isn't configured (no service-account "
                      "credentials). See setup notes."}, status=400)
    fid = gdrive.folder_id(request.POST.get("folder"))
    if not fid:
        return JsonResponse(
            {"error": "No Drive folder set. Pass a folder link or set "
                      "GDRIVE_INVOICE_FOLDER."}, status=400)
    try:
        files = gdrive.list_files(fid)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Drive listing failed")
        return JsonResponse(
            {"error": f"Could not read the Drive folder (is it shared with the "
                      f"service account, and the Drive API enabled?): {exc}"}, status=502)

    invoice_files = [f for f in files if gdrive.supported(f.get("name"))]
    logger.info("drive import: listed %d file(s) from Drive, %d supported",
                len(files), len(invoice_files))
    if not invoice_files:
        return JsonResponse(
            {"error": "No .xlsx / .xlsb / .csv files found in that Drive folder."},
            status=400)

    # Each file's billing month is inferred from its name using the same rule the
    # invoice parser uses to bucket months, so the client can offer (and require)
    # only the months that actually exist in the folder — no download needed here.
    def _month_key(f):
        return invoices.month_from_filename(f.get("name") or "")[0]

    # action=months: report the months present in the folder (newest first) and
    # return without downloading anything.
    if request.POST.get("action") == "months":
        buckets = {}
        for f in invoice_files:
            key, label = invoices.month_from_filename(f.get("name") or "")
            b = buckets.get(key)
            if b is None:
                b = buckets[key] = {"key": key, "label": label, "count": 0}
            b["count"] += 1
        return JsonResponse({"months": _sorted_months(buckets.values())})

    # Import path: a month must be selected, and only that month's files load.
    month = (request.POST.get("month") or "").strip()
    if not month:
        return JsonResponse({"error": "Select a month to import."}, status=400)
    invoice_files = [f for f in invoice_files if _month_key(f) == month]
    if not invoice_files:
        return JsonResponse(
            {"error": "No files for the selected month were found in the Drive "
                      "folder."}, status=400)

    if request.POST.get("append") != "1":
        _reset_invoice_cache()

    errors = []
    master_saved = False
    # Build the Drive client once and reuse it for every file — rebuilding it per
    # file (new credentials + discovery fetch) was the main slowdown. Downloads
    # stay sequential: parallel threads segfaulted the worker on Vercel's
    # serverless runtime, and one-file-at-a-time also keeps peak memory low.
    logger.info("drive import: month=%s, downloading %d file(s)", month, len(invoice_files))
    svc = gdrive.new_service()
    for f in invoice_files:
        name = f.get("name")
        try:
            data = gdrive.download(f["id"], f.get("mimeType"), svc=svc)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{name}: download failed ({exc})")
            continue
        logger.info("drive import: downloaded %s (%d bytes), parsing", name, len(data))
        kind = _ingest_into_cache(name, data, errors)
        # A master file found in Drive becomes the persisted default item master
        # (so it's remembered across restarts, like the "Update item master"
        # upload). Best-effort: a read-only filesystem just skips this.
        if kind == "master":
            try:
                os.makedirs(os.path.dirname(_MASTER_PATH), exist_ok=True)
                with open(_MASTER_PATH, "wb") as fh:
                    fh.write(data)
                _MASTER_CACHE["mtime"] = None
                _MASTER_CACHE["maps"] = None
                master_saved = True
            except OSError:
                pass

    if not _INVOICE_CACHE["items"]:
        if master_saved:
            dm = _default_master()
            return JsonResponse({"ok": True, "master_only": True,
                                 "master": {"sku_count": dm["sku_count"],
                                            "updated_at": dm["updated_at"]},
                                 "warnings": errors})
        msg = ("No billed invoice lines found in the Drive files. "
               + (" · ".join(errors) if errors else ""))
        return JsonResponse({"error": msg}, status=400)

    _INVOICE_CACHE["version"] += 1
    _INVOICE_CACHE["report_cache"] = {}
    # BigQuery product enrichment maps each invoice AWB -> product / category / SKU
    # (and feeds the weight-dispute check). It now streams rows instead of
    # materialising the window through pyarrow (see _refresh_invoice_product_map ->
    # use_arrow=False), so it no longer segfaults the serverless worker. On by
    # default; set INVOICE_SKIP_BQ_ENRICH=1 to turn it off if ever needed.
    if os.environ.get("INVOICE_SKIP_BQ_ENRICH") in ("1", "true", "True"):
        logger.info("drive import: parsed %d item(s); BQ enrichment skipped (env)",
                    len(_INVOICE_CACHE["items"]))
    else:
        logger.info("drive import: parsed %d item(s); starting BQ product enrichment",
                    len(_INVOICE_CACHE["items"]))
        _refresh_invoice_product_map()
        logger.info("drive import: BQ product enrichment done")
    report = _invoice_report()
    report["imported_from_drive"] = len(invoice_files)
    report["imported_month"] = month
    if errors:
        report["warnings"] = errors
    logger.info("drive import: complete, returning report")
    return JsonResponse(report)


@auth.team_required
@require_POST
def process_invoices(request):
    """Parse one or more carrier invoice (or master) files and return the
    cost-analysis report.

    Each file is auto-detected: an invoice's billed lines are aggregated into
    spend by carrier x category x SKU; a master/reference file (weights, SKU
    master) feeds AWB/SKU -> category maps used to enrich invoices that don't
    carry product info. `reset=1` (no files) clears everything; `append=1` adds
    to the current set instead of replacing it.
    """
    if request.POST.get("reset") == "1" and not request.FILES.getlist("files"):
        _reset_invoice_cache()
        return JsonResponse(_invoice_report())

    uploads = request.FILES.getlist("files") or request.FILES.getlist("file")
    if not uploads:
        # No new files: a plain refresh or a carrier-filter request on the
        # already-cached invoice lines.
        return JsonResponse(_invoice_report(request.POST.get("carrier")))

    if request.POST.get("append") != "1":
        _reset_invoice_cache()

    errors = []
    for up in uploads:
        try:
            up.seek(0)
        except Exception:  # noqa: BLE001
            pass
        _ingest_into_cache(up.name, up.read() or b"", errors)

    if not _INVOICE_CACHE["items"]:
        msg = ("No billed invoice lines found. "
               + (" · ".join(errors) if errors else
                  "Uploaded file(s) had no amounts — add an invoice with charges."))
        return JsonResponse({"error": msg}, status=400)

    # New data ingested (incl. append): invalidate the cached reports.
    _INVOICE_CACHE["version"] += 1
    _INVOICE_CACHE["report_cache"] = {}

    # Search BigQuery by the invoices' service month and cache the AWB -> SKU /
    # sub-category map, so the sub-category analysis is populated automatically.
    _refresh_invoice_product_map()

    # A fresh upload always shows the full (unfiltered) picture; the frontend
    # resets its carrier dropdown to "All carriers" to match.
    report = _invoice_report()
    if errors:
        report["warnings"] = errors
    return JsonResponse(report)


@auth.team_required
@require_POST
def export_invoice_awbs(request):
    """Download every billed AWB line as CSV, enriched with the SKU /
    sub-category joined from the loaded BigQuery shipment data. This is the full
    per-shipment detail behind the per-invoice reconciliation rows (the UI shows
    only counts and the sub-category rollup)."""
    items = _INVOICE_CACHE["items"]
    if not items:
        return JsonResponse({"error": "Upload invoices first, then export."}, status=400)

    # Enrich in place using the same map the report uses (BigQuery searched by
    # invoice date, falling back to the loaded shipment records).
    awb2prod, order2prod = _invoice_product_maps()
    invoices._attach_products(items, awb2prod, order2prod)

    carrier = (request.POST.get("carrier") or "").strip()
    rows = items
    if carrier and carrier != "all":
        rows = [it for it in items if it["carrier"] == carrier]

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["AWB", "Invoice Number", "Service Month", "Carrier", "Direction",
                "SKU", "Sub-category", "Category", "Item Name",
                "Charged Wt (kg)", "Amount with GST", "Amount ex GST",
                "Matched from BigQuery"])
    for it in rows:
        w.writerow([
            it.get("awb", ""), it.get("invoice_number", ""), it.get("service_month", ""),
            it.get("carrier", ""), it.get("direction", ""),
            it.get("sku", ""), it.get("subcategory", ""), it.get("category", ""),
            it.get("product") or it.get("item_name") or it.get("sku_name") or "",
            it.get("weight_kg") or "",
            round(it.get("amount") or 0.0, 2), round(it.get("amount_ex_gst") or 0.0, 2),
            "yes" if it.get("prod_matched") else "no",
        ])

    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="invoice_awb_detail.csv"'
    return resp


@auth.team_required
@require_POST
def master_config(request):
    """View or replace the persisted item master (SKU -> volumetric weight) used
    for the weight-dispute comparison.

    actions: get (status) | upload (replace with a new .xlsx/.csv). The uploaded
    file is validated as a recognised master (must yield SKU volumetric weights)
    before it replaces the stored one, and the invoice report cache is cleared so
    the new weights apply immediately."""
    action = request.POST.get("action", "get")
    if action == "upload":
        upload = request.FILES.get("file")
        if upload is None:
            return JsonResponse({"error": "No file uploaded."}, status=400)
        if not upload.name.lower().endswith((".xlsx", ".xlsm", ".csv", ".tsv")):
            return JsonResponse({"error": "Upload a .xlsx or .csv item master."}, status=400)
        data = upload.read() or b""
        try:
            kind, payload = invoices.ingest(data, upload.name)
        except Exception as exc:  # noqa: BLE001
            return JsonResponse({"error": f"Could not read the file: {exc}"}, status=400)
        if kind != "master" or not payload.get("sku2vol"):
            return JsonResponse(
                {"error": "This file isn't a recognised item master — it needs a "
                          "SKU column and a volumetric/billable weight column."},
                status=400)
        try:
            os.makedirs(os.path.dirname(_MASTER_PATH), exist_ok=True)
            with open(_MASTER_PATH, "wb") as fh:
                fh.write(data)
        except OSError as exc:
            return JsonResponse({"error": f"Could not save the master: {exc}"}, status=500)
        _MASTER_CACHE["mtime"] = None
        _MASTER_CACHE["maps"] = None
        # New weights -> invalidate cached invoice reports.
        _INVOICE_CACHE["version"] += 1
        _INVOICE_CACHE["report_cache"] = {}

    dm = _default_master()
    return JsonResponse({"ok": True, "loaded": dm["sku_count"] > 0,
                         "sku_count": dm["sku_count"], "updated_at": dm["updated_at"]})


@auth.team_required
@require_POST
def load_bigquery(request):
    """Fetch a lookback window from BigQuery into the cache, then return the report."""
    if not bq.is_configured():
        return JsonResponse(
            {"error": "BigQuery is not configured on the server "
                      "(set BQ_PROJECT, BQ_DATASET and BQ_TABLE)."},
            status=400,
        )
    # Either an explicit calendar range (win_from/win_to, from the date pickers)
    # or a lookback window (how many days back to pull). The range takes
    # precedence when present.
    win_from = (request.POST.get("win_from") or "").strip()
    win_to = (request.POST.get("win_to") or "").strip()
    use_range = bool(win_from or win_to)

    lookback = request.POST.get("lookback_days")
    lookback_days = int(lookback) if (lookback or "").isdigit() else None

    try:
        records = bq.fetch_records(
            lookback_days=lookback_days,
            date_from=win_from or None,
            date_to=win_to or None,
        )
    except ValueError as exc:  # bad date input -> client error, not a 502
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:  # noqa: BLE001 - surface any BQ/auth failure cleanly
        # Log the full traceback to the server console so the real cause is
        # visible (the client only sees the short message below).
        logger.exception("BigQuery load failed")
        return JsonResponse({"error": f"BigQuery load failed: {exc}"}, status=502)

    if not records:
        return JsonResponse(
            {"error": "BigQuery returned no rows for the selected date range."},
            status=400,
        )
    _apply_master_categories(records)
    _CACHE["records"] = records
    # The authoritative window the UI shows: the picked range, or the lookback.
    if use_range:
        _CACHE["window"] = {"from": win_from or None, "to": win_to or None, "days": None}
    else:
        _CACHE["window"] = bq.lookback_window(lookback_days)

    # Orders-per-day chart: pull "orders placed per day" from Unicommerce for the
    # SAME window. ClickPost only holds shipped orders, so its recent days look
    # like a false decline; Unicommerce is complete at placement time. Best-effort
    # — on any failure (table missing, auth) fall back to the ClickPost series.
    uc_from = win_from or None if use_range else None
    uc_to = win_to or None if use_range else None
    uc_lookback = None if use_range else lookback_days
    try:
        _CACHE["daily_orders"] = bq.fetch_d2c_orders_per_day(
            date_from=uc_from, date_to=uc_to, lookback_days=uc_lookback) or None
    except Exception:  # noqa: BLE001 - never block the load on the orders series
        logger.exception("Unicommerce orders-per-day fetch failed")
        _CACHE["daily_orders"] = None
    # Accurate O2S: recompute each record's order->pickup on the Unicommerce
    # order-received time (falls back to the ClickPost value on failure), so the
    # headline card and every per-carrier/warehouse O2S column share one basis.
    try:
        order_times = bq.fetch_d2c_order_times(
            date_from=uc_from, date_to=uc_to, lookback_days=uc_lookback)
        if order_times:
            _recompute_o2s_from_uc(records, order_times)
    except Exception:  # noqa: BLE001 - never block the load on the O2S refinement
        logger.exception("Unicommerce O2S recompute failed")

    return _report_response(request, "No data loaded.")