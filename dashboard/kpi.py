"""
KPI computation engine for carrier partner efficiency.

Reads a shipment workbook (same column layout as the standard export) and
produces per-carrier KPIs plus a weighted 0-100 efficiency score, along with
overall summary numbers and business-mix breakdowns.

This version adds WAREHOUSE-LEVEL CARRIER TRACKING: in addition to the global
per-carrier scoreboard and the per-warehouse scoreboard, it computes a
warehouse x carrier matrix so you can see how each partner performs out of a
specific warehouse (pickup pincode). Warehouses are labelled "City - Pincode"
using a pincode-prefix lookup so the UI reads naturally.

All timestamp columns may arrive either as real datetimes (openpyxl converts
Excel serials automatically) or as raw numeric serials; both are handled.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from functools import lru_cache
from io import BytesIO

from openpyxl import load_workbook

from . import tat as tat_rules

# Logical field -> list of candidate header names (first match wins). Matching
# is case/space-insensitive, so minor header drift won't break parsing, and any
# export that is a SUBSET of the full 151-column standard format works: missing
# optional columns simply yield blank/None for that field.
COLUMNS = {
    "carrier": ["Carrier Partner Name"],
    "account": ["Carrier Partner Account Name"],
    "weight": ["Shipment Weight"],
    "payment": ["Payment Mode"],
    "pickup_pin": ["Pickup Pincode"],
    "drop_pin": ["Drop Pin Code", "Drop Pincode"],
    # Destination city as shipped (free text). Optional; we fall back to a
    # pincode-derived region when it's blank.
    "drop_city": ["Drop City", "Destination City"],
    "pickup_ts": ["Pickup Timestamp"],
    # Order-received time, used for the Order->Pickup processing time (O2S).
    # In the standard export this is the customer order/creation date.
    "order_ts": ["Order Date", "Order Created At", "Order Creation Date",
                 "Order Placed", "Order Timestamp", "Created At"],
    "delivery_ts": ["Delivery Timestamp"],
    "ofd1_ts": ["OFD1 Timestamp"],
    # The standard export uses "Shipment Zone"; older trimmed exports used "Zone".
    "zone": ["Zone", "Shipment Zone", "Pricing Zone"],
    # Sales channel the order came from (Shopify, Amazon, etc.). Optional.
    "channel": ["Channel Name", "Channel", "channel_name"],
    "delivery_type": ["Delivery Type"],
    "status": ["Latest Status"],
    "attempts": ["Number of Delivery Attempts"],
    # Product info (used to derive category / subcategory). Optional.
    "item_names": ["Item Names", "Item Name", "Product Name"],
    "sku": ["Product SKU Codes", "SKU", "Product SKU Code"],
    # Declared order/invoice value in ₹ (the sale value of the goods = revenue).
    # Used for the revenue panel and per-tier revenue. Optional in uploads.
    "order_value": ["Invoice Value", "Order Value", "Invoice Amount",
                    "Order Amount", "Invoice Amt", "Total Value"],
    # COD collectable amount in ₹ (0 for prepaid). Used for COD cash exposure.
    "cod_value": ["COD Value", "COD Amount", "Collectable Amount", "COD Amt"],
    # Shipment identifiers — optional, used to join carrier invoices back to the
    # shipment master (recover SKU/category for invoices that lack them).
    "awb": ["AWB", "AWB Number", "AWB No", "AWB No.", "Waybill Number", "Tracking Id", "awb_number"],
    "order_id": ["Order Number", "Order Id", "Client Order Id", "client_order_id", "Reference Number"],
}

# Fields that are required; if none of their candidate headers are present the
# file is rejected with a clear error. Everything else is optional.
REQUIRED_FIELDS = ["carrier"]

# Weights for the composite efficiency score (must sum to 1.0).
WEIGHTS = {
    "p2o": 0.30,   # Pickup -> OFD1 TAT       (lower is better)
    "p2d": 0.25,   # Pickup -> Delivery TAT   (lower is better)
    "succ": 0.20,  # Delivery success rate    (higher is better)
    "fa": 0.15,    # First-attempt strike     (higher is better)
    "att": 0.10,   # Avg delivery attempts    (lower is better)
}

EXCEL_EPOCH = datetime(1899, 12, 30)

# ---------------------------------------------------------------------------
# Pincode -> city resolution.
#
# The export has no pickup-city column (Drop City is the destination, not the
# warehouse), so we derive a readable warehouse name from the pickup pincode.
# We use the first 3 digits of the Indian PIN, which maps to a sorting
# district / region. The table below covers the high-volume warehouses seen in
# the data plus common metro prefixes; anything unmatched falls back to the
# 2-digit postal-circle name, and finally to "PIN <prefix>xx".
# ---------------------------------------------------------------------------

# 3-digit prefix -> city/region label (highest priority).
PIN3_CITY = {
    "412": "Pune (Maval)",
    "411": "Pune",
    "413": "Solapur",
    "122": "Gurugram",
    "121": "Faridabad",
    "110": "Delhi",
    "562": "Bengaluru (Rural)",
    "560": "Bengaluru",
    "561": "Bengaluru (Rural)",
    "712": "Howrah",
    "711": "Howrah",
    "700": "Kolkata",
    "501": "Hyderabad (RR)",
    "500": "Hyderabad",
    "421": "Thane",
    "400": "Mumbai",
    "401": "Palghar",
    "440": "Nagpur",
    "302": "Jaipur",
    "380": "Ahmedabad",
    "600": "Chennai",
    "632": "Vellore",
    "590": "Belagavi",
    "682": "Kochi",
    "457": "Ratlam",
    "470": "Sagar",
    "532": "Srikakulam",
    "232": "Ghazipur",
}

# 2-digit prefix -> postal circle / broad region (fallback).
PIN2_CIRCLE = {
    "11": "Delhi", "12": "Haryana", "13": "Punjab", "14": "Punjab",
    "16": "Chandigarh", "17": "Himachal", "18": "J&K", "19": "J&K",
    "20": "UP (West)", "21": "UP", "22": "UP", "23": "UP", "24": "UP",
    "25": "UP", "26": "UP", "27": "UP", "28": "UP",
    "30": "Rajasthan", "31": "Rajasthan", "32": "Rajasthan", "33": "Rajasthan",
    "34": "Rajasthan",
    "36": "Gujarat", "37": "Gujarat", "38": "Gujarat", "39": "Gujarat",
    "40": "Mumbai/MH", "41": "Maharashtra", "42": "Maharashtra",
    "43": "Maharashtra", "44": "Maharashtra",
    "45": "MP", "46": "MP", "47": "MP", "48": "MP",
    "49": "Chhattisgarh",
    "50": "Telangana", "51": "Andhra", "52": "Andhra", "53": "Andhra",
    "56": "Karnataka", "57": "Karnataka", "58": "Karnataka", "59": "Karnataka",
    "60": "Chennai/TN", "61": "Tamil Nadu", "62": "Tamil Nadu",
    "63": "Tamil Nadu", "64": "Tamil Nadu",
    "67": "Kerala", "68": "Kerala", "69": "Kerala",
    "70": "Kolkata/WB", "71": "West Bengal", "72": "West Bengal",
    "73": "West Bengal", "74": "West Bengal",
    "75": "Odisha", "76": "Odisha", "77": "Odisha",
    "78": "Assam", "79": "North East",
    "80": "Bihar", "81": "Bihar", "82": "Jharkhand", "83": "Jharkhand",
    "84": "Bihar", "85": "Bihar",
}


@lru_cache(maxsize=None)
def city_for_pincode(pin: str) -> str:
    """Return a readable city/region for a (string) pincode, or '' if blank."""
    if not pin:
        return ""
    digits = "".join(ch for ch in str(pin) if ch.isdigit())
    if len(digits) < 2:
        return ""
    p3 = digits[:3]
    if p3 in PIN3_CITY:
        return PIN3_CITY[p3]
    p2 = digits[:2]
    if p2 in PIN2_CIRCLE:
        return PIN2_CIRCLE[p2]
    return "PIN " + p3 + "xx"


@lru_cache(maxsize=None)
def warehouse_label(pin: str) -> str:
    """'City - Pincode' label used as the warehouse display name."""
    if not pin:
        return ""
    city = city_for_pincode(pin)
    return (city + " \u00b7 " + str(pin)) if city else str(pin)


# ---------------------------------------------------------------------------
# City-tier classification (by DESTINATION drop pincode).
#
# Standard e-commerce style India tiering keyed on the 3-digit PIN prefix:
#   Tier 1 : the 8 major metros + their metro regions (Mumbai MMR, Delhi-NCR,
#            Bengaluru, Hyderabad, Chennai, Kolkata, Pune, Ahmedabad).
#   Tier 2 : state capitals & other large cities.
#   Tier 3 : everywhere else with a resolvable pincode.
# A blank/too-short pincode is "Unknown".
# ---------------------------------------------------------------------------
TIER_LEVELS = ["Tier 1", "Tier 2", "Tier 3"]

TIER1_PIN3 = {
    "400", "401", "410", "421",          # Mumbai MMR (Mumbai, Mira-Bhayandar, Navi Mumbai, Thane)
    "110", "201", "122", "121",          # Delhi NCR (Delhi, Noida/Ghaziabad, Gurugram, Faridabad)
    "560", "561", "562",                 # Bengaluru
    "500", "501",                        # Hyderabad
    "600", "601", "602", "603",          # Chennai
    "700", "711", "712",                 # Kolkata / Howrah
    "411", "412",                        # Pune
    "380", "382",                        # Ahmedabad / Gandhinagar
}

TIER2_PIN3 = {
    "302", "303", "305", "313", "324", "342",                       # Rajasthan
    "390", "391", "360", "361", "364", "394", "395",                # Gujarat
    "226", "227", "208", "282", "221", "250", "211", "243", "202", "273",  # UP
    "452", "453", "462", "463", "482", "474", "456",                # MP
    "440", "441", "422", "423", "431", "416", "413", "444",         # Maharashtra (non-metro)
    "530", "531", "520", "521", "522", "517", "524", "506",         # AP / Telangana
    "682", "683", "695", "673", "680", "691", "670",                # Kerala
    "641", "642", "625", "620", "621", "636", "627", "638", "632", "605",  # TN / Puducherry
    "160", "140", "134", "141", "142", "143", "144", "147",         # Chandigarh / Punjab
    "492", "493", "490",                                            # Chhattisgarh
    "800", "801", "823", "812",                                     # Bihar
    "834", "831", "826", "827",                                     # Jharkhand
    "751", "753", "769",                                            # Odisha
    "781",                                                          # Assam (Guwahati)
    "248", "249",                                                   # Uttarakhand
    "570", "571", "575", "580", "590", "591", "577",                # Karnataka (non-metro)
    "403",                                                          # Goa
    "734", "713",                                                   # West Bengal (Siliguri, Durgapur/Asansol)
    "180", "190", "171",                                            # J&K / Himachal
}
@lru_cache(maxsize=100_000)
def tier_for_pincode(pin: str) -> str:
    """Tier 1/2/3 (or 'Unknown') for a destination pincode, by 3-digit prefix."""
    if not pin:
        return "Unknown"
    digits = "".join(ch for ch in str(pin) if ch.isdigit())
    if len(digits) < 3:
        return "Unknown"
    p3 = digits[:3]
    if p3 in TIER1_PIN3:
        return "Tier 1"
    if p3 in TIER2_PIN3:
        return "Tier 2"
    return "Tier 3"


@lru_cache(maxsize=100_000)
def _clean_city(text: str) -> str:
    """Normalize a free-text city name so casing/spacing variants merge.

    The export's Drop City is raw uppercase free text (e.g. 'NEW DELHI',
    ' mumbai '); collapsing whitespace and title-casing folds the obvious
    duplicates together for a cleaner destination rollup.
    """
    if not text:
        return ""
    return " ".join(str(text).split()).title()


def _norm_pincode(raw) -> str:
    """Normalize a pincode cell to a clean digit string ('560037.0' -> '560037')."""
    if raw is None or str(raw).strip() == "":
        return ""
    try:
        return str(int(float(raw)))
    except (ValueError, TypeError):
        return str(raw).strip()


def _norm_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _to_datetime(value):
    """Coerce a cell value to datetime, or None if not parseable/blank."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return EXCEL_EPOCH + timedelta(days=float(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s:
        return None
    # ISO first (real datetimes / clean exports).
    try:
        return datetime.fromisoformat(s)
    except (ValueError, TypeError):
        pass
    # CSV exports of this dataset use US-style M/D/YYYY with an optional time,
    # e.g. "6/2/2026 23:58" or "6/2/2026". Try the common explicit formats.
    for fmt in (
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%m/%d/%y %H:%M:%S", "%m/%d/%y %H:%M", "%m/%d/%y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _hours_between(start, end):
    """Positive elapsed hours, or None if either end is missing/negative."""
    if start is None or end is None:
        return None
    delta = (end - start).total_seconds() / 3600.0
    return delta if delta >= 0 else None


def _to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _weight_bucket(grams):
    if grams is None:
        return "Unknown"
    if grams <= 500:
        return "0-500g"
    if grams <= 1000:
        return "500g-1kg"
    if grams <= 5000:
        return "1-5kg"
    return "5kg+"


# Coarse 3-way weight class used for the filter dropdown and its breakdown.
# Weights in the export are in GRAMS, so the kg thresholds are *1000.
#   Light  : 0 - 3 kg
#   Medium : 3 - 10 kg
#   Heavy  : 10 kg and above
WEIGHT_LIGHT_MAX = 3000     # <= 3 kg
WEIGHT_MEDIUM_MAX = 10000   # 3 kg - 10 kg


@lru_cache(maxsize=4096)
def _weight_class(grams):
    if grams is None:
        return "Unknown"
    if grams <= WEIGHT_LIGHT_MAX:
        return "Light"
    if grams <= WEIGHT_MEDIUM_MAX:
        return "Medium"
    return "Heavy"


# Time-of-day slots for the pickup timestamp. Hourly between 12pm and 5pm,
# with open buckets before noon and after 5pm:
#   Before 12pm : hour < 12
#   12-1pm      : hour == 12
#   1-2pm       : hour == 13
#   2-3pm       : hour == 14
#   3-4pm       : hour == 15
#   4-5pm       : hour == 16
#   After 5pm   : hour >= 17
# The slot labels double as the filter values (kept stable for the frontend).
PICKUP_SLOTS = [
    "Before 12pm", "12\u20131pm", "1\u20132pm", "2\u20133pm",
    "3\u20134pm", "4\u20135pm", "After 5pm",
]

# Map the hour (12..16) to its label for the hourly band.
_HOURLY_SLOT = {
    12: "12\u20131pm", 13: "1\u20132pm", 14: "2\u20133pm",
    15: "3\u20134pm", 16: "4\u20135pm",
}


def _pickup_slot(dt):
    if dt is None:
        return "Unknown"
    h = dt.hour
    if h < 12:
        return "Before 12pm"
    if h >= 17:
        return "After 5pm"
    return _HOURLY_SLOT[h]


# ---------------------------------------------------------------------------
# Product categorization (derived from the Item Names text, since the export
# has no category column). Rules are ordered: the FIRST matching (category,
# subcategory) whose keyword appears in the item name wins, so put more
# specific rules before broader ones. Matching is case-insensitive on whole
# words/substrings of the product name.
# ---------------------------------------------------------------------------
# Each rule: (category, subcategory, [keywords]). First keyword hit wins, so
# ORDER MATTERS - more specific categories come before broader ones (a
# "pregnancy pillow" must hit Maternity before the generic Pillows rule; an
# "arch support insole" must hit Insoles before Footwears/Orthotics).
#
# Top-level categories are consolidated into 6 themed groups (Comfort & Sleep,
# Footwear, Orthopedic & Wellness, Mobility & Furniture, Maternity & Baby Care,
# Accessories); the subcategory (2nd field) preserves the finer catalogue
# grouping and can be tuned freely. Anything unmatched falls to
# ("Others", "Other"). Keywords match as case-insensitive substrings.
PRODUCT_RULES = [
    # --- Maternity & Baby Care (FIRST, so a pregnancy/maternity/feeding pillow
    # lands here instead of under Comfort & Sleep) ---
    ("Maternity & Baby Care", "Pregnancy Pillow", ["pregnancy", "maternity"]),
    ("Maternity & Baby Care", "Baby Care", ["baby", "infant", "nursing", "feeding pillow", "kids"]),

    # --- Footwear :: Barefoot (Frido's barefoot shoes/socks line — before the
    # broader Footwear/Socks rules so a "barefoot sock shoe" lands here) ---
    ("Footwear", "Barefoot", ["barefoot", "sock shoe", "skinners"]),

    # --- Mobility & Furniture :: Mobility Devices (wheelchairs, scooters,
    # transfer/bathroom aids). Early so a wheelchair mentioning "footrest"/"seat"
    # isn't read as Foot Care/Cushion. ---
    ("Mobility & Furniture", "Wheelchair", ["wheelchair", "wheel chair"]),
    ("Mobility & Furniture", "Mobility Scooter", ["mobility scooter", "scooter"]),
    ("Mobility & Furniture", "Transfer & Lift Aids", ["transfer lift", "patient lift", "transfer aid", "hoist", "walker", "rollator", "crutch"]),
    ("Mobility & Furniture", "Commode", ["commode"]),
    ("Mobility & Furniture", "Bathroom Safety", ["bath mat", "anti-slip", "anti slip", "grab bar", "bed rail", "shower stool", "shower chair", "ramp"]),

    # --- Mobility & Furniture :: Chairs (ergonomic / office / gaming, recliners) ---
    ("Mobility & Furniture", "Ergonomic Chair", ["ergonomic chair", "ergo chair", "office chair", "gaming chair", "study chair"]),
    ("Mobility & Furniture", "Recliner", ["recliner"]),
    ("Mobility & Furniture", "Chair", ["chair"]),

    # --- Mobility & Furniture :: Workspace (standing desks, laptop tables/stands) ---
    ("Mobility & Furniture", "Standing Desk", ["standing desk", "height desk", "adjustable desk"]),
    ("Mobility & Furniture", "Desk & Table", ["laptop table", "study table", "work table", "desk", "workstation"]),
    ("Mobility & Furniture", "Stand & Mount", ["laptop stand", "monitor stand", "monitor arm", "monitor mount", "laptop mount", "laptop holder", "monitor", "foot rest", "footrest"]),

    # --- Footwear :: Insoles (before the broader Footwear/Orthopedic rules) ---
    ("Footwear", "Insoles", ["insole", "arch support", "shoe insert", "foot insert"]),

    # --- Footwear :: Socks ---
    ("Footwear", "Socks", ["sock"]),

    # --- Orthopedic & Wellness :: Face Mask (anti-pollution; sleep/eye masks below) ---
    ("Orthopedic & Wellness", "Face Mask", ["face mask", "n95", "n-95", "anti pollution", "anti-pollution", "pollution mask", "surgical mask"]),

    # --- Comfort & Sleep :: Mattress Topper / Protector ---
    ("Comfort & Sleep", "Topper", ["mattress topper", "topper"]),
    ("Comfort & Sleep", "Protector", ["mattress protector", "mattress"]),

    # --- Comfort & Sleep :: Covers (before Pillows/Cushions rules so "pillow
    # cover" and "cushion cover" land here) ---
    ("Comfort & Sleep", "Covers", ["cuddle cover", "wedge cover", "pillow cover", "cushion cover", "seat cover", "replacement cover", "cover"]),

    # --- Comfort & Sleep :: Pillows ---
    ("Comfort & Sleep", "Neck Pillow", ["neck pillow", "cervical", "neck contour", "travel pillow"]),
    ("Comfort & Sleep", "Wedge Pillow", ["wedge"]),
    ("Comfort & Sleep", "Sleep Pillow", ["sleep pillow", "cozy pillow", "memory foam pillow", "bed pillow", "pillow"]),

    # --- Comfort & Sleep :: Cushions ---
    ("Comfort & Sleep", "Seat Cushion", ["seat cushion", "donut", "coccyx", "seat"]),
    ("Comfort & Sleep", "Backrest", ["backrest", "back rest", "lumbar cushion", "lumbar"]),
    ("Comfort & Sleep", "Cushion", ["cushion"]),

    # --- Footwear (before Orthopedic; "footwear"/"foot" would otherwise hit Foot Care) ---
    ("Footwear", "Sandals", ["sandal"]),
    ("Footwear", "Slippers", ["slipper", "flip flop", "flipflop", "clog", "chappal"]),
    ("Footwear", "Shoes", ["shoe", "sneaker", "footwear"]),

    # --- Orthopedic & Wellness :: Orthotics (posture, braces, joint & foot supports) ---
    ("Orthopedic & Wellness", "Posture Corrector", ["posture"]),
    ("Orthopedic & Wellness", "Knee & Joint Support", ["knee", "ankle", "elbow", "wrist", "shoulder"]),
    ("Orthopedic & Wellness", "Braces & Wraps", ["brace", "wrap", "lumbo sacral", "sacral", "compression", "support belt", "belt", "support"]),
    ("Orthopedic & Wellness", "Foot Care", ["bunion", "heel", "plantar", "toe", "arch", "orthotic", "foot"]),

    # --- Orthopedic & Wellness :: Personal Care (therapy, sleep/eye masks, nasal, massage) ---
    ("Orthopedic & Wellness", "Sleep & Eye Mask", ["eye mask", "sleep mask", "mask"]),
    ("Orthopedic & Wellness", "Hot/Cold Therapy", ["therapy", "hot & cold", "cold & hot", "heating pad", "heat pad"]),
    ("Orthopedic & Wellness", "Nasal Care", ["nasal", "nose"]),
    ("Orthopedic & Wellness", "Massage & Relief", ["massager", "massage", "roller", "pain relief", "pain-relief", "kinesiology", "tape"]),
    ("Orthopedic & Wellness", "Gloves", ["glove"]),

    # --- Accessories (wallets, straps, bags, combos, spare parts) ---
    ("Accessories", "Wallet", ["wallet", "card holder", "cardholder"]),
    ("Accessories", "Spare Parts", ["spare part", "sparepart", "castor wheel", "castor", "joystick"]),
    ("Accessories", "Accessories", ["strap", "pouch", "bag", "cap", "combo", "accessor"]),
]


@lru_cache(maxsize=100_000)
def _product_category(item_name):
    """Return (category, subcategory) for a product name, or ('Others','Other')."""
    if not item_name:
        return ("Unknown", "Unknown")
    text = item_name.lower()
    for category, subcategory, keywords in PRODUCT_RULES:
        for kw in keywords:
            if kw in text:
                return (category, subcategory)
    return ("Others", "Other")


# ---------------------------------------------------------------------------
# Consolidated category groups. Shipments are tagged with fine-grained catalogue
# categories (from the item master's Category column, invoice enrichment, or the
# name rules above). For the Product breakdown we roll those up into 6 legible
# groups. This is a DISPLAY rollup only — the fine category is left on the record
# in case other consumers need the finer taxonomy.
# ---------------------------------------------------------------------------
CATEGORY_GROUP = {
    # Comfort & Sleep
    "Pillows": "Comfort & Sleep",
    "Cushions": "Comfort & Sleep",
    "Mattress Topper Protector": "Comfort & Sleep",
    "Covers": "Comfort & Sleep",
    # Footwear
    "Footwear": "Footwear",
    "Barefoot": "Footwear",
    "Insoles": "Footwear",
    "Socks": "Footwear",
    # Orthopedic & Wellness
    "Orthotics": "Orthopedic & Wellness",
    "Personal Care": "Orthopedic & Wellness",
    "Masks": "Orthopedic & Wellness",
    # Mobility & Furniture
    "Mobility Devices": "Mobility & Furniture",
    "Chairs": "Mobility & Furniture",
    "Workspace": "Mobility & Furniture",
    # Unchanged
    "Maternity & Baby Care": "Maternity & Baby Care",
    "Accessories": "Accessories",
    # Fallbacks
    "Others": "Others",
    "Unknown": "Others",
}


def canonical_category(cat):
    """Roll a fine-grained catalogue category up to one of the 6 consolidated
    groups. Values already in group form (or unknown labels) pass through."""
    if not cat:
        return cat
    return CATEGORY_GROUP.get(cat, cat)


# ---------------------------------------------------------------------------
# Delivery-outcome categorization (drives the status matrix).
#
# Latest Status is bucketed into four mutually exclusive outcomes that mirror
# the ops pivot: Delivered / FWD Pendency / RTO / Cancelled. FWD Pendency is
# the catch-all for shipments still moving through the forward pipeline (not
# yet delivered, not returned, not cancelled). The four outcome buckets are a
# strict partition, so per-carrier shares sum to 100%.
# ---------------------------------------------------------------------------
OUTCOMES = ["Delivered", "FWD Pendency", "RTO", "Cancelled"]

# A still-"pending" order that was never picked up and is at least this many days
# old (measured from its order-placed date) is presumed lost and reclassified from
# FWD Pendency to Cancelled. Orders younger than this may simply be awaiting
# pickup, so they are left unchanged.
LOST_ORDER_AGE_DAYS = 30

# Latest Status values treated as terminal cancellations (not pendency).
_CANCELLED_STATUSES = {
    "cancelled", "notserviceable", "lost", "damaged", "nostatusexist",
}


@lru_cache(maxsize=None)
def _outcome(status):
    s = (status or "").strip()
    low = s.lower()
    if low == "delivered":
        return "Delivered"
    # Any RTO-* / RTO ... status is a return-to-origin outcome.
    if low.startswith("rto-") or low.startswith("rto "):
        return "RTO"
    if low in _CANCELLED_STATUSES:
        return "Cancelled"
    return "FWD Pendency"


# The FWD-Pendency sub-states we surface as columns in the pendency matrix.
# These are the Latest Status values that fall under FWD Pendency, matching the
# operational checklist. Any pendency status not in this list rolls into
# "Other".
PENDENCY_STATES = [
    "InTransit", "OutForDelivery", "FailedDelivery", "ShipmentDelayed",
    "OutForPickup", "PickupFailed", "PickupPending", "ShipmentHeld",
    "DestinationHubIn", "OriginCityIn", "OriginCityOut", "PickedUp",
    "ContactCustomerCare", "OrderPlaced", "Awb Registered",
]
_PENDENCY_SET = {s.lower() for s in PENDENCY_STATES}


@lru_cache(maxsize=None)
def _pendency_state(status):
    """Sub-bucket for a FWD-Pendency shipment's Latest Status."""
    s = (status or "").strip()
    if s.lower() in _PENDENCY_SET:
        # Return the canonical spelling from PENDENCY_STATES.
        for canon in PENDENCY_STATES:
            if canon.lower() == s.lower():
                return canon
    return "Other"


def _read_all_bytes(file_obj) -> bytes:
    """Return the full byte content of an upload, regardless of stream state.

    Django's UploadedFile may have already had its position moved (by a size
    check, middleware, or a prior read), which would make a plain .read()
    return empty bytes and openpyxl raise 'File is not a zip file'. We rewind
    when possible and fall back to .chunks() for large/temporary uploads.
    """
    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass

    data = b""
    chunks = getattr(file_obj, "chunks", None)
    if callable(chunks):
        try:
            data = b"".join(chunks())
        except Exception:  # noqa: BLE001 - fall through to .read()
            data = b""
    if not data:
        try:
            file_obj.seek(0)
        except (AttributeError, OSError):
            pass
        data = file_obj.read() or b""

    if isinstance(data, str):
        data = data.encode("utf-8", "ignore")
    return data


def _rows_from_bytes(data: bytes, filename: str = ""):
    """Yield rows (tuples) from either an .xlsx or a .csv byte payload.

    Returns an iterator over rows where the first row is the header. Raises
    ValueError with a clear message if the bytes are neither.
    """
    name = (filename or "").lower()
    is_xlsx = data[:2] == b"PK"

    if is_xlsx:
        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return ws.iter_rows(values_only=True)

    # Not a ZIP/xlsx. Old .xls (OLE) can't be read here.
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValueError(
            "This looks like an old-format .xls file. Please re-save it as .xlsx "
            "or export as CSV, then upload again."
        )

    # Treat as CSV/TSV text. Decode tolerantly (handles the BOM and stray bytes).
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1", "replace")

    # If it looks like HTML, reject with a clear message.
    stripped = text.lstrip()
    if stripped[:1] == "<" or stripped[:5].lower() == "<?xml":
        raise ValueError(
            "This file looks like HTML/XML, not a workbook or CSV. Open it in "
            "Excel and Save As 'CSV' or 'Excel Workbook (.xlsx)', then upload."
        )

    import csv as _csv
    import io as _io
    # Sniff the delimiter (comma vs tab vs semicolon); default to comma.
    sample = text[:4096]
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delim = dialect.delimiter
    except _csv.Error:
        delim = "\t" if name.endswith(".tsv") else ","
    reader = _csv.reader(_io.StringIO(text), delimiter=delim)
    return reader


def build_record(get):
    """Build one normalized record dict from a raw-field getter.

    `get(logical_key)` returns the raw cell value for a logical field (a key of
    COLUMNS) or None. This is the single source of truth for row normalization,
    shared by the CSV/XLSX parser and the BigQuery loader (dashboard/bq.py) so
    both produce byte-for-byte identical records. Returns None for rows with no
    carrier (caller should skip).
    """
    carrier = get("carrier")
    if carrier is None or str(carrier).strip() == "":
        return None

    pickup = _to_datetime(get("pickup_ts"))
    order = _to_datetime(get("order_ts"))   # order-received time, for O2S
    win = _to_datetime(get("window_ts"))    # load-window/partition date (BQ only)
    ofd1 = _to_datetime(get("ofd1_ts"))
    delivery = _to_datetime(get("delivery_ts"))
    edd = _to_datetime(get("edd_ts"))   # carrier's committed expected delivery date
    attempts = _to_float(get("attempts"))
    weight = _to_float(get("weight"))
    order_value = _to_float(get("order_value"))  # declared ₹ value of the order
    cod_value = _to_float(get("cod_value"))       # ₹ to collect on delivery (COD)
    status = str(get("status") or "").strip()
    payment = str(get("payment") or "").strip().upper()

    # Pincodes can arrive as ints or floats (560037.0) - normalize to a clean
    # string without decimals.
    pickup_pin = _norm_pincode(get("pickup_pin"))
    drop_pin = _norm_pincode(get("drop_pin"))

    pickup_city = city_for_pincode(pickup_pin)
    # Destination: prefer the explicit Drop City; fall back to a region derived
    # from the drop pincode, then to the bare pincode.
    drop_city = _clean_city(str(get("drop_city") or "").strip())
    if not drop_city:
        drop_city = city_for_pincode(drop_pin) or drop_pin
    # Lane label only when both endpoints resolve to something readable.
    lane = (pickup_city + " → " + drop_city) if (pickup_city and drop_city) else ""

    item_name = str(get("item_names") or "").strip()
    category, subcategory = _product_category(item_name)

    carrier_name = str(carrier).strip()
    account_name = str(get("account") or "").strip()
    delivered_flag = status.lower() == "delivered"
    outcome = _outcome(status)
    p2d = _hours_between(pickup, delivery)
    # Calendar days pickup->delivery (date diff), used by day-based SLAs.
    transit_days = (delivery.date() - pickup.date()).days if (pickup and delivery) else None
    if transit_days is not None and transit_days < 0:
        transit_days = None
    # Elapsed AGE of an undelivered shipment (pickup -> now), so an active
    # forward-pendency shipment that has already blown past its promised TAT can
    # be flagged Out of TAT before it is ever delivered. Same units as transit:
    # hours for hour-based carriers, calendar days for day-based ones.
    now = datetime.now()
    age_hours = _hours_between(pickup, now)
    age_days = (now.date() - pickup.date()).days if pickup else None
    if age_days is not None and age_days < 0:
        age_days = None
    # Presumed-lost reclassification: an order still in FWD Pendency that was
    # never picked up and is at least LOST_ORDER_AGE_DAYS old (by order-placed
    # date) will never move, so treat it as Cancelled / Lost instead of pending.
    # This shifts its order value out of the FWD Pendency revenue bucket into
    # Cancelled / Lost. Younger, not-yet-picked orders are left as pending.
    if outcome == "FWD Pendency" and pickup is None:
        order_age_days = (now.date() - order.date()).days if order else None
        if order_age_days is not None and order_age_days >= LOST_ORDER_AGE_DAYS:
            outcome = "Cancelled"
    forward_pending = outcome == "FWD Pendency"
    # First out-for-delivery attempt timing (pickup->OFD1), used to score RTO
    # shipments: they were attempted but returned, so judge the carrier on when
    # it first tried, not on a delivery that never happened.
    is_rto = outcome == "RTO"
    p2o = _hours_between(pickup, ofd1)
    ofd1_days = (ofd1.date() - pickup.date()).days if (pickup and ofd1) else None
    if ofd1_days is not None and ofd1_days < 0:
        ofd1_days = None
    # Carrier's committed transit window: days from pickup to its expected
    # delivery date. The universal SLA scores delivery vs this (delivered <= EDD).
    edd_days = (edd.date() - pickup.date()).days if (pickup and edd) else None
    if edd_days is not None and edd_days < 0:
        edd_days = None
    # Promised-TAT (SLA) compliance: compare actual transit against the carrier's
    # promised TAT for this lane (hours for Blue Dart, calendar days for Delhivery).
    # Late forward pendency past its SLA is counted as Out of TAT for all carriers;
    # RTO is scored on pickup->OFD1 vs the same promise.
    tat_status, promised_tat, tat_margin = tat_rules.classify(
        carrier_name, account_name, pickup_pin, payment, drop_pin,
        p2d, transit_days, delivered_flag,
        age_hours=age_hours, age_days=age_days, forward_pending=forward_pending,
        rto=is_rto, ofd1_hours=p2o, ofd1_days=ofd1_days,
        pickup_city=pickup_city, drop_city=drop_city, edd_days=edd_days,
    )

    return {
        "carrier": carrier_name,
        "account": account_name,
        "delivery_type": str(get("delivery_type") or "").strip(),
        "zone": str(get("zone") or "").strip(),
        "channel": str(get("channel") or "").strip(),
        "pickup_pin": pickup_pin,
        "warehouse": warehouse_label(pickup_pin),
        "city": pickup_city,
        "drop_pin": drop_pin,
        "drop_city": drop_city,
        "tier": tier_for_pincode(drop_pin),
        "lane": lane,
        "payment": payment,
        "weight": weight,
        "weight_class": _weight_class(weight),
        "order_value": order_value,
        "cod_value": cod_value,
        "pickup_date": pickup.date().isoformat() if pickup else "",
        # Order-placed date (every order has one, even before pickup). Kept for
        # the O2S metric and as the per-day chart fallback for uploaded files.
        "order_date": order.date().isoformat() if order else "",
        # Load-window/partition date — the same day the load window and the
        # headline KPI cards count on. Drives the orders-per-day chart so it
        # reconciles with the "Shipments" total. Empty for uploaded files.
        "window_date": win.date().isoformat() if win else "",
        "pickup_slot": _pickup_slot(pickup),
        "status": status,
        "outcome": outcome,
        "pendency_state": _pendency_state(status),
        "item_name": item_name,
        "sku": str(get("sku") or "").strip(),
        "category": category,
        "subcategory": subcategory,
        "awb": str(get("awb") or "").strip(),
        "order_id": str(get("order_id") or "").strip(),
        "picked": pickup is not None,
        "delivered": delivered_flag,
        "attempts": attempts,
        "p2o": p2o,
        "p2d": p2d,
        # O2S: order-received -> pickup processing time, in hours. Order date is
        # often date-only, so this is effectively (pickup - order midnight).
        "o2s": _hours_between(order, pickup),
        "promised_tat": promised_tat,
        "tat_status": tat_status,
        "tat_margin": tat_margin,
        # Raw inputs kept so TAT can be recomputed in place (reclassify) when the
        # user changes a carrier's SLA, without re-querying BigQuery.
        "transit_days": transit_days,
        "age_hours": age_hours,
        "age_days": age_days,
        "ofd1_days": ofd1_days,
        "edd_days": edd_days,
    }


def parse_workbook(file_obj, filename: str = "") -> list[dict]:
    """Read the first worksheet/CSV into a list of normalized row dicts.

    Accepts .xlsx/.xlsm workbooks and .csv/.tsv files (the standard export is
    a 151-column CSV). 'filename' is an optional hint used only for delimiter
    defaults; content sniffing is primary.
    """
    data = _read_all_bytes(file_obj)
    if not data:
        raise ValueError(
            "The uploaded file was empty (no bytes received). Please re-select "
            "the file and try again."
        )

    rows_iter = _rows_from_bytes(data, filename)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    # Map our logical names to actual column indices. Each field has a list of
    # candidate header names; the first one present in this file wins. This is
    # what lets any subset of the standard format work.
    lookup = {_norm_header(h): i for i, h in enumerate(header)}
    idx = {}
    for key, candidates in COLUMNS.items():
        for name in candidates:
            i = lookup.get(_norm_header(name))
            if i is not None:
                idx[key] = i
                break

    missing_required = [k for k in REQUIRED_FIELDS if k not in idx]
    if missing_required:
        wanted = ", ".join(COLUMNS[k][0] for k in missing_required)
        raise ValueError(
            "The file is missing required column(s): " + wanted + ". "
            "Check that the uploaded file matches the standard export format."
        )

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    records = []
    for row in rows_iter:
        if row is None:
            continue
        rec = build_record(lambda key, _row=row: cell(_row, key))
        if rec is not None:
            records.append(rec)
    return records


def _mean(values):
    clean = [v for v in values if v is not None]
    return sum(clean) / len(clean) if clean else None


# Carrier accounts treated as Next-Day-Delivery (NDD) partners, matched as
# case-insensitive substrings of the account code:
#   "ndd"        -> Swift_NDD, Delhivery NDD, Elasticrun_sdd&ndd, Shadofax_NDD_Large
#   "skye"       -> SKYE Air
#   "urbanbolt"  -> Urbanbolt (Urbane Bolt)
#   "elasticrun" -> ElasticRun (NDD wholesale)
#   "shadofax"/"shadowfax" -> Shadowfax (NDD wholesale)
# Note: Swift and Delhivery are NOT matched wholesale (they also run reverse/
# heavy/mobility accounts) — only their explicit "NDD" accounts qualify.
NDD_ACCOUNT_KEYWORDS = (
    "ndd", "skye", "urbanbolt", "urbane", "elasticrun", "shadofax", "shadowfax",
)

# Accounts classified as NDD by EXACT (case-insensitive) name, used where a
# substring match would be too broad — e.g. the partner literally named "Test"
# (matching the substring "test" would also catch words like "fastest").
NDD_ACCOUNT_EXACT = {"test"}


@lru_cache(maxsize=8192)
def _is_ndd(account) -> bool:
    a = (account or "").strip().lower()
    if a in NDD_ACCOUNT_EXACT:
        return True
    return any(kw in a for kw in NDD_ACCOUNT_KEYWORDS)


def _filter_set(val):
    """Normalize a filter argument to a set of accepted values, or None for 'all'.

    Each categorical filter may arrive as 'all'/''/None (no filtering), a single
    string value, or a list/tuple/set of values (multi-select). Returning None
    means 'accept everything'; a set means 'accept only these'.
    """
    if val is None or val == "" or val == "all":
        return None
    if isinstance(val, (list, tuple, set)):
        picked = {v for v in val if v not in (None, "", "all")}
        return picked or None
    return {val}


def filter_records(records, delivery_type="all", zone="all", payment="all",
                   warehouse="all", account="all", weight="all",
                   slot="all", date_from="", date_to="", tier="all",
                   channel="all"):
    # Every categorical filter supports single value, list (multi-select) or
    # "all". An empty selection (None) means no constraint on that field.
    dt_set = _filter_set(delivery_type)
    zone_set = _filter_set(zone)
    pay_set = _filter_set(payment)
    wh_set = _filter_set(warehouse)
    acct_set = _filter_set(account)
    wt_set = _filter_set(weight)
    slot_set = _filter_set(slot)
    tier_set = _filter_set(tier)
    channel_set = _filter_set(channel)

    # Fast path: nothing is actually constrained (the common initial/unfiltered
    # load), so skip a full scan + list copy and hand back the record set as-is.
    if (dt_set is None and zone_set is None and pay_set is None and wh_set is None
            and acct_set is None and wt_set is None and slot_set is None
            and tier_set is None and channel_set is None
            and not date_from and not date_to):
        return records

    out = []
    for r in records:
        if dt_set is not None and r["delivery_type"] not in dt_set:
            continue
        if tier_set is not None and r["tier"] not in tier_set:
            continue
        if channel_set is not None and r["channel"] not in channel_set:
            continue
        if zone_set is not None and r["zone"] not in zone_set:
            continue
        if pay_set is not None and r["payment"] not in pay_set:
            continue
        if wh_set is not None and r["pickup_pin"] not in wh_set:
            continue
        if acct_set is not None and r["account"] not in acct_set:
            continue
        if wt_set is not None and r["weight_class"] not in wt_set:
            continue
        if slot_set is not None and r["pickup_slot"] not in slot_set:
            continue
        # Date range is inclusive on both ends and is applied to the ORDER date
        # (the day the customer placed the order) so the selected range matches
        # order-date-based sales reports. Rows with no order date are dropped
        # only when a bound is set (they can't be attributed to a calendar day).
        if date_from:
            if not r["order_date"] or r["order_date"] < date_from:
                continue
        if date_to:
            if not r["order_date"] or r["order_date"] > date_to:
                continue
        out.append(r)
    return out


def _kpi_accumulate(groups, r, key, label_field, extra_fields):
    """Fold one record `r` into its KPI `groups` bucket under `key` (creating the
    bucket on first sight). Shared by aggregate_by and the single-pass collector
    so both build byte-for-byte identical group state."""
    g = groups.get(key)
    if g is None:
        g = {
            label_field: key, "n": 0, "picked": 0, "delivered": 0,
            "first_attempt": 0, "o2s": [], "p2o": [], "p2d": [], "att": [],
        }
        for out_name, rec_key in extra_fields:
            g[out_name] = r.get(rec_key, "")
        groups[key] = g
    g["n"] += 1
    if r["picked"]:
        g["picked"] += 1
    if r["delivered"]:
        g["delivered"] += 1
        if r["attempts"] == 1:
            g["first_attempt"] += 1
    if r["o2s"] is not None:
        g["o2s"].append(r["o2s"])
    if r["p2o"] is not None:
        g["p2o"].append(r["p2o"])
    if r["p2d"] is not None:
        g["p2d"].append(r["p2d"])
    if r["attempts"] is not None:
        g["att"].append(r["attempts"])


def _kpi_finalize(groups, label_field, extra_fields) -> list[dict]:
    """Turn accumulated KPI `groups` into the standard per-group result rows."""
    result = []
    for g in groups.values():
        row = {
            label_field: g[label_field],
            "n": g["n"],
            "picked": g["picked"],
            "delivered": g["delivered"],
            "success_rate": (g["delivered"] / g["picked"] * 100) if g["picked"] else None,
            "first_attempt_rate": (g["first_attempt"] / g["delivered"] * 100) if g["delivered"] else None,
            "o2s": _mean(g["o2s"]),
            "p2o": _mean(g["p2o"]),
            "p2d": _mean(g["p2d"]),
            "avg_attempts": _mean(g["att"]),
        }
        for out_name, _ in extra_fields:
            row[out_name] = g.get(out_name, "")
        result.append(row)
    return result


def aggregate_by(records, key_field, label_field="group", extra_fields=None) -> list[dict]:
    """Group records by any field and compute the standard KPI block.

    extra_fields: optional list of (out_name, record_key) to carry the first
    seen value of a record field onto each group row (e.g. carry the city
    label onto a warehouse group).
    """
    extra_fields = extra_fields or []
    groups: dict[str, dict] = {}
    for r in records:
        key = r.get(key_field) or ""
        if key == "":
            continue
        _kpi_accumulate(groups, r, key, label_field, extra_fields)
    return _kpi_finalize(groups, label_field, extra_fields)


def aggregate_by_carrier(records) -> list[dict]:
    return aggregate_by(records, "carrier", "carrier")


def _score_metric(value, all_values, lower_is_better):
    """Min-max normalize one metric to 0-100 across the carrier set."""
    pool = [v for v in all_values if v is not None]
    if value is None or len(pool) < 2:
        return None
    lo, hi = min(pool), max(pool)
    if hi == lo:
        return 100.0
    s = (value - lo) / (hi - lo)
    if lower_is_better:
        s = 1 - s
    return s * 100.0


def attach_scores(agg: list[dict], min_n: int = 0) -> list[dict]:
    # Only rows meeting the volume threshold participate in the normalization
    # pool and receive a score - this keeps a 3-shipment warehouse from
    # producing a meaningless rank.
    scorable = [a for a in agg if a["n"] >= min_n]
    pools = {
        "p2o": [a["p2o"] for a in scorable],
        "p2d": [a["p2d"] for a in scorable],
        "succ": [a["success_rate"] for a in scorable],
        "fa": [a["first_attempt_rate"] for a in scorable],
        "att": [a["avg_attempts"] for a in scorable],
    }
    lower_better = {"p2o": True, "p2d": True, "succ": False, "fa": False, "att": True}

    for a in agg:
        if a["n"] < min_n:
            a["subscores"] = {k: None for k in WEIGHTS}
            a["score"] = None
            continue
        raw = {
            "p2o": a["p2o"], "p2d": a["p2d"], "succ": a["success_rate"],
            "fa": a["first_attempt_rate"], "att": a["avg_attempts"],
        }
        sub = {k: _score_metric(raw[k], pools[k], lower_better[k]) for k in WEIGHTS}
        a["subscores"] = {k: (round(v, 1) if v is not None else None) for k, v in sub.items()}

        total_w, acc = 0.0, 0.0
        for k, w in WEIGHTS.items():
            if sub[k] is not None:
                acc += sub[k] * w
                total_w += w
        a["score"] = round(acc / total_w, 1) if total_w > 0 else None
    return agg


def _round(v, dp=1):
    return round(v, dp) if v is not None else None


def _round_metrics(agg: list[dict]) -> None:
    """Round the numeric KPI fields in place for JSON transport."""
    for a in agg:
        a["success_rate"] = _round(a["success_rate"])
        a["first_attempt_rate"] = _round(a["first_attempt_rate"])
        a["o2s"] = _round(a["o2s"])
        a["p2o"] = _round(a["p2o"])
        a["p2d"] = _round(a["p2d"])
        a["avg_attempts"] = _round(a["avg_attempts"], 2)


# Warehouses with fewer than this many shipments still appear in the table but
# are not assigned an efficiency score (too little data to rank fairly).
WAREHOUSE_MIN_N = 20

# Warehouse x carrier cells need fewer shipments to be meaningful than a whole
# warehouse, but we still suppress scores for very thin cells.
WH_CARRIER_MIN_N = 10

# Only the top-N warehouses by shipment volume are shown in the breakdown table
# (scores are still computed across ALL warehouses, so each rank is fair; the
# table is just truncated). Set to None to show every warehouse.
TOP_N_WAREHOUSES = 10

# Destination (drop city/region) and lane (pickup -> drop) breakdowns. Like the
# warehouse table, scores are computed across ALL rows then the table is
# truncated to the top-N by volume so each rank stays fair.
TOP_N_DESTINATIONS = 12
TOP_N_LANES = 15
# Destinations reuse the warehouse volume threshold; lanes are finer-grained so
# they get a slightly lower bar before a score is assigned.
DESTINATION_MIN_N = WAREHOUSE_MIN_N
LANE_MIN_N = 15


def aggregate_warehouse_carrier(records, by_wh=None) -> list[dict]:
    """Warehouse x carrier-account matrix.

    One row per (pickup pincode, carrier account) pair. Each row carries the
    warehouse label, city, carrier account code and the standard KPI block.
    Scores are computed PER WAREHOUSE: within each warehouse the accounts are
    normalized against each other, so the score answers 'which carrier account
    is best out of THIS warehouse', which is exactly the comparison an ops lead
    wants.
    """
    rows = []
    # Bucket records by warehouse first (or reuse a bucket dict already built by
    # the single-pass collector so we don't scan the rows a second time).
    if by_wh is None:
        by_wh = {}
        for r in records:
            if not r["pickup_pin"]:
                continue
            by_wh.setdefault(r["pickup_pin"], []).append(r)

    for pin, recs in by_wh.items():
        agg = aggregate_by(
            recs, "account", "account",
            extra_fields=[("warehouse", "warehouse"), ("city", "city")],
        )
        # Score carriers relative to each other within this warehouse.
        attach_scores(agg, min_n=WH_CARRIER_MIN_N)
        _round_metrics(agg)
        wh_total = sum(a["n"] for a in agg)
        for a in agg:
            a["pickup_pin"] = pin
            a["wh_total"] = wh_total
            a["wh_share"] = _round(a["n"] / wh_total * 100 if wh_total else None)
        agg.sort(key=lambda a: (a["score"] is None, -(a["score"] or 0), -a["n"]))
        rows.extend(agg)
    return rows


def aggregate_status_matrix(records, key_field="account") -> list[dict]:
    """Per-carrier-account outcome matrix (mirrors the ops pivot).

    One row per carrier account, with the count and % share of each of the
    four outcomes (Delivered / FWD Pendency / RTO / Cancelled). Shares are of
    that account's own total, so each row's four percentages sum to 100.
    """
    groups: dict[str, dict] = {}
    for r in records:
        key = r.get(key_field) or ""
        if key == "":
            continue
        _status_accumulate(groups, r, key)
    return _status_finalize(groups)


def _status_accumulate(groups, r, key):
    g = groups.get(key)
    if g is None:
        g = {"key": key, "n": 0, "counts": {o: 0 for o in OUTCOMES}}
        groups[key] = g
    g["n"] += 1
    g["counts"][r["outcome"]] += 1


def _status_finalize(groups) -> list[dict]:
    result = []
    for g in groups.values():
        n = g["n"]
        row = {"account": g["key"], "n": n, "counts": dict(g["counts"]), "pct": {}}
        for o in OUTCOMES:
            row["pct"][o] = _round(g["counts"][o] / n * 100 if n else None)
        result.append(row)
    result.sort(key=lambda a: -a["n"])
    return result


def aggregate_pendency_matrix(records, key_field="account") -> dict:
    """FWD-Pendency sub-state matrix, per carrier account.

    Only shipments whose outcome is FWD Pendency are counted. Columns are the
    PENDENCY_STATES (plus "Other"); each row's percentages are of that
    account's pendency total. Returns the rows plus the ordered column list so
    the frontend can render a stable header.
    """
    groups: dict[str, dict] = {}
    for r in records:
        if r["outcome"] != "FWD Pendency":
            continue
        key = r.get(key_field) or ""
        if key == "":
            continue
        _pendency_accumulate(groups, r, key)
    return _pendency_finalize(groups)


_PENDENCY_COLS = PENDENCY_STATES + ["Other"]


def _pendency_accumulate(groups, r, key):
    g = groups.get(key)
    if g is None:
        g = {"key": key, "n": 0, "counts": {c: 0 for c in _PENDENCY_COLS}}
        groups[key] = g
    g["n"] += 1
    g["counts"][r["pendency_state"]] += 1


def _pendency_finalize(groups) -> dict:
    cols = _PENDENCY_COLS
    rows = []
    for g in groups.values():
        n = g["n"]
        row = {"account": g["key"], "n": n, "counts": dict(g["counts"]), "pct": {}}
        for c in cols:
            row["pct"][c] = _round(g["counts"][c] / n * 100 if n else None)
        rows.append(row)
    rows.sort(key=lambda a: -a["n"])
    # Drop columns that are entirely zero across all accounts to keep the table
    # readable (e.g. RiderAllocated rarely appears).
    active_cols = [c for c in cols
                   if any(rw["counts"].get(c, 0) > 0 for rw in rows)]
    return {"columns": active_cols, "rows": rows}


# ---------------------------------------------------------------------------
# Promised-TAT (SLA) compliance: In TAT / Out of TAT breakdown.
#
# Each record already carries tat_status / promised_tat / tat_margin (set in
# build_record via dashboard/tat.py). Here we roll those up per carrier account
# and per warehouse, plus an overall summary. The headline metric is:
#
#   In-TAT %     = In TAT / (In TAT + Out of TAT)   -> of shipments with a
#                                                      determinable SLA outcome
#   coverage %   = (In + Out + Pending) / total      -> share of all shipments
#                                                      that HAVE a promised TAT
#
# "Out of TAT" now also includes active forward-pendency shipments whose age
# has already passed the promised TAT (a confirmed breach before delivery).
# "No rule" = no promised TAT for the lane (unmapped origin warehouse, or
# destination pincode not in the carrier master). "Pending" = has a rule, not
# delivered yet, and still inside the promised window.
# ---------------------------------------------------------------------------
def _tat_block(counts, delivered, margins):
    """Shared metric block from raw status counts."""
    evaluable = counts["In TAT"] + counts["Out of TAT"]
    has_rule = evaluable + counts["Pending"]   # everything except "No rule"
    total = has_rule + counts["No rule"]
    return {
        "counts": dict(counts),
        "delivered": delivered,
        "evaluable": evaluable,
        "in_tat": counts["In TAT"],
        "out_tat": counts["Out of TAT"],
        "no_rule": counts["No rule"],
        "pending": counts["Pending"],
        "in_tat_pct": _round(counts["In TAT"] / evaluable * 100 if evaluable else None),
        "out_tat_pct": _round(counts["Out of TAT"] / evaluable * 100 if evaluable else None),
        # Share of all shipments that carry an SLA rule (master coverage of
        # volume); independent of delivery, so it stays <=100% even though
        # Out of TAT can now include undelivered breaches.
        "coverage_pct": _round(has_rule / total * 100 if total else None),
        "avg_margin": _round(_mean(margins)),
    }


def aggregate_tat_matrix(records, key_field="account") -> list[dict]:
    """Per-key (account/warehouse) promised-TAT compliance rows, ranked by volume."""
    statuses = tat_rules.TAT_STATUSES
    groups: dict[str, dict] = {}
    for r in records:
        key = r.get(key_field) or ""
        if key == "":
            continue
        g = groups.get(key)
        if g is None:
            g = {"key": key, "delivered": 0,
                 "counts": {s: 0 for s in statuses}, "margins": [],
                 "warehouse": r.get("warehouse", "")}
            groups[key] = g
        if r["delivered"]:
            g["delivered"] += 1
        g["counts"][r.get("tat_status") or "No rule"] += 1
        if r.get("tat_margin") is not None:
            g["margins"].append(r["tat_margin"])

    rows = []
    for g in groups.values():
        block = _tat_block(g["counts"], g["delivered"], g["margins"])
        block[key_field] = g["key"]
        block["n"] = sum(g["counts"].values())
        block["warehouse"] = g["warehouse"]
        rows.append(block)
    rows.sort(key=lambda a: -a["n"])
    return rows


def _tat_new_state() -> dict:
    return {"counts": {s: 0 for s in tat_rules.TAT_STATUSES},
            "delivered": 0, "out_delivered": 0, "out_pending": 0, "margins": []}


def _tat_accumulate(state, r):
    # Out of TAT splits into two kinds: shipments that were DELIVERED but later
    # than promised ("delivered late"), and active forward-pendency shipments
    # not yet delivered whose age already passed the promise ("pending breach").
    st = r.get("tat_status") or "No rule"
    state["counts"][st] += 1
    if st == "Out of TAT":
        if r["delivered"]:
            state["out_delivered"] += 1
        else:
            state["out_pending"] += 1
    if r["delivered"]:
        state["delivered"] += 1
    if r.get("tat_margin") is not None:
        state["margins"].append(r["tat_margin"])


def _tat_finalize(state) -> dict:
    block = _tat_block(state["counts"], state["delivered"], state["margins"])
    block["statuses"] = tat_rules.TAT_STATUSES
    block["out_delivered"] = state["out_delivered"]
    block["out_pending"] = state["out_pending"]
    return block


def tat_summary(records) -> dict:
    """Overall promised-TAT compliance across the filtered record set."""
    state = _tat_new_state()
    for r in records:
        _tat_accumulate(state, r)
    return _tat_finalize(state)


# Filter-bar option lists (zones / accounts / warehouses / date bounds) depend
# only on the loaded record set, not on the active filters — identical on
# every re-filter. Cache them keyed on the records list identity to skip four
# full-dataset passes per refresh. The strong reference pins the list's identity
# (so id reuse can't cause a stale hit) and is replaced when new data loads.
_FILTER_OPTIONS_CACHE = {"records": None, "value": None}


def _filter_options(records) -> dict:
    if _FILTER_OPTIONS_CACHE["records"] is records:
        return _FILTER_OPTIONS_CACHE["value"]

    zones = sorted({r["zone"] for r in records if r["zone"]})
    channel_counts: dict[str, int] = {}
    acct_counts: dict[str, int] = {}
    wh_counts: dict[str, int] = {}
    wh_labels: dict[str, str] = {}
    pickup_dates_min = None
    pickup_dates_max = None
    for r in records:
        ch = r["channel"]
        if ch:
            channel_counts[ch] = channel_counts.get(ch, 0) + 1
        a = r["account"]
        if a:
            acct_counts[a] = acct_counts.get(a, 0) + 1
        pin = r["pickup_pin"]
        if pin:
            wh_counts[pin] = wh_counts.get(pin, 0) + 1
            wh_labels[pin] = r["warehouse"]
        pd = r["pickup_date"]
        if pd:
            if pickup_dates_min is None or pd < pickup_dates_min:
                pickup_dates_min = pd
            if pickup_dates_max is None or pd > pickup_dates_max:
                pickup_dates_max = pd

    channels = [
        {"value": c, "label": c, "n": cnt}
        for c, cnt in sorted(channel_counts.items(), key=lambda kv: -kv[1])
    ]
    accounts = [
        {"value": a, "label": a, "n": cnt}
        for a, cnt in sorted(acct_counts.items(), key=lambda kv: -kv[1])
    ]
    warehouses_opts = [
        {"value": pin, "label": wh_labels[pin], "n": cnt}
        for pin, cnt in sorted(wh_counts.items(), key=lambda kv: -kv[1])
    ]
    if TOP_N_WAREHOUSES is not None:
        warehouses_opts = warehouses_opts[:TOP_N_WAREHOUSES]

    value = {
        "zones": zones,
        "channels": channels,
        "accounts": accounts,
        "warehouses": warehouses_opts,
        "weight_classes": ["Light", "Medium", "Heavy"],
        "pickup_slots": PICKUP_SLOTS,
        "date_min": pickup_dates_min or "",
        "date_max": pickup_dates_max or "",
    }
    _FILTER_OPTIONS_CACHE["records"] = records
    _FILTER_OPTIONS_CACHE["value"] = value
    return value


def reclassify(records) -> None:
    """Recompute tat_status / promised_tat / tat_margin in place for every
    record, using the inputs stored on each row. Called after the user changes
    a carrier's SLA (override) so the dashboard re-scores without re-querying."""
    # TAT scoring changes in place -> invalidate any cached reports for this set.
    global _REPORT_GEN
    _REPORT_GEN += 1
    for r in records:
        outcome = r.get("outcome")
        status, promised, margin = tat_rules.classify(
            r.get("carrier"), r.get("account"), r.get("pickup_pin"),
            r.get("payment"), r.get("drop_pin"),
            r.get("p2d"), r.get("transit_days"), r.get("delivered"),
            age_hours=r.get("age_hours"), age_days=r.get("age_days"),
            forward_pending=(outcome == "FWD Pendency"),
            rto=(outcome == "RTO"),
            ofd1_hours=r.get("p2o"), ofd1_days=r.get("ofd1_days"),
            pickup_city=r.get("city"), drop_city=r.get("drop_city"),
            edd_days=r.get("edd_days"),
        )
        r["tat_status"], r["promised_tat"], r["tat_margin"] = status, promised, margin


# In-process cache of the built report, keyed on the loaded record-set identity
# plus the active filter arguments. Toggling a filter and returning to a prior
# state (or a plain reload with the same filters) then becomes an instant cache
# hit instead of a full recompute. Invalidated automatically when a new record
# set is loaded (its list identity changes) or when reclassify() re-scores the
# TAT status in place (the generation counter is bumped).
_REPORT_GEN = 0
_REPORT_CACHE = {"records": None, "gen": None, "entries": {}}


def _report_cache_key(delivery_type, zone, payment, warehouse, account, weight,
                      slot, date_from, date_to, tier, channel):
    """Hashable, order-independent key for a filter selection (multi-select
    values arrive as lists, so sort them; two selections that filter to the same
    rows map to the same key)."""
    def norm(v):
        if isinstance(v, (list, tuple, set)):
            return tuple(sorted(str(x) for x in v))
        return v
    return tuple(norm(v) for v in (
        delivery_type, zone, payment, warehouse, account, weight, slot,
        date_from, date_to, tier, channel))


def _report_cache_get(records, key):
    c = _REPORT_CACHE
    if c["records"] is not records or c["gen"] != _REPORT_GEN:
        return None
    return c["entries"].get(key)


def _report_cache_put(records, key, report):
    c = _REPORT_CACHE
    if c["records"] is not records or c["gen"] != _REPORT_GEN:
        c["records"] = records
        c["gen"] = _REPORT_GEN
        c["entries"] = {}
    c["entries"][key] = report


def _aggregate_all(rows):
    """ONE pass over the filtered rows building every row-grouping the report
    needs: the carrier / warehouse / destination / lane KPI groups, the
    per-account status and FWD-pendency matrices, the overall TAT summary, and
    the warehouse->rows buckets for the warehouse x carrier matrix. Replaces
    eight separate scans of `rows` with a single scan.

    The logic here is deliberately identical to _kpi_accumulate /
    _status_accumulate / etc.; _kpi_finalize / _status_finalize /
    _pendency_finalize / _tat_finalize turn the accumulated state into the exact
    same result the standalone aggregate_* functions produce."""
    def fold(g, picked, delivered, att, o2s, p2o, p2d):
        g["n"] += 1
        if picked:
            g["picked"] += 1
        if delivered:
            g["delivered"] += 1
            if att == 1:
                g["first_attempt"] += 1
        if o2s is not None:
            g["o2s"].append(o2s)
        if p2o is not None:
            g["p2o"].append(p2o)
        if p2d is not None:
            g["p2d"].append(p2d)
        if att is not None:
            g["att"].append(att)

    carrier_g: dict = {}
    wh_g: dict = {}
    dest_g: dict = {}
    lane_g: dict = {}
    status_g: dict = {}
    pend_g: dict = {}
    by_wh: dict = {}
    tat_counts = {s: 0 for s in tat_rules.TAT_STATUSES}
    tat_delivered = tat_out_delivered = tat_out_pending = 0
    tat_margins: list = []

    for r in rows:
        picked = r["picked"]
        delivered = r["delivered"]
        att = r["attempts"]
        o2s = r["o2s"]
        p2o = r["p2o"]
        p2d = r["p2d"]
        outcome = r["outcome"]

        # --- carrier / warehouse / destination / lane KPI groups ---
        ck = r["carrier"]
        if ck:
            g = carrier_g.get(ck)
            if g is None:
                g = {"carrier": ck, "n": 0, "picked": 0, "delivered": 0,
                     "first_attempt": 0, "o2s": [], "p2o": [], "p2d": [], "att": []}
                carrier_g[ck] = g
            fold(g, picked, delivered, att, o2s, p2o, p2d)

        wk = r["pickup_pin"]
        if wk:
            g = wh_g.get(wk)
            if g is None:
                g = {"pickup_pin": wk, "n": 0, "picked": 0, "delivered": 0,
                     "first_attempt": 0, "o2s": [], "p2o": [], "p2d": [], "att": [],
                     "warehouse": r.get("warehouse", ""), "city": r.get("city", "")}
                wh_g[wk] = g
            fold(g, picked, delivered, att, o2s, p2o, p2d)
            bucket = by_wh.get(wk)
            if bucket is None:
                by_wh[wk] = [r]
            else:
                bucket.append(r)

        dk = r["drop_city"]
        if dk:
            g = dest_g.get(dk)
            if g is None:
                g = {"drop_city": dk, "n": 0, "picked": 0, "delivered": 0,
                     "first_attempt": 0, "o2s": [], "p2o": [], "p2d": [], "att": []}
                dest_g[dk] = g
            fold(g, picked, delivered, att, o2s, p2o, p2d)

        lk = r["lane"]
        if lk:
            g = lane_g.get(lk)
            if g is None:
                g = {"lane": lk, "n": 0, "picked": 0, "delivered": 0,
                     "first_attempt": 0, "o2s": [], "p2o": [], "p2d": [], "att": []}
                lane_g[lk] = g
            fold(g, picked, delivered, att, o2s, p2o, p2d)

        # --- per-account status matrix (+ FWD-pendency sub-states) ---
        ak = r.get("account") or ""
        if ak:
            g = status_g.get(ak)
            if g is None:
                g = {"key": ak, "n": 0, "counts": {o: 0 for o in OUTCOMES}}
                status_g[ak] = g
            g["n"] += 1
            g["counts"][outcome] += 1
            if outcome == "FWD Pendency":
                gp = pend_g.get(ak)
                if gp is None:
                    gp = {"key": ak, "n": 0,
                          "counts": {c: 0 for c in _PENDENCY_COLS}}
                    pend_g[ak] = gp
                gp["n"] += 1
                gp["counts"][r["pendency_state"]] += 1

        # --- overall TAT summary ---
        st = r.get("tat_status") or "No rule"
        tat_counts[st] += 1
        if st == "Out of TAT":
            if delivered:
                tat_out_delivered += 1
            else:
                tat_out_pending += 1
        if delivered:
            tat_delivered += 1
        if r.get("tat_margin") is not None:
            tat_margins.append(r["tat_margin"])

    tat_state = {"counts": tat_counts, "delivered": tat_delivered,
                 "out_delivered": tat_out_delivered,
                 "out_pending": tat_out_pending, "margins": tat_margins}
    return {"carrier": carrier_g, "warehouse": wh_g, "dest": dest_g,
            "lane": lane_g, "status": status_g, "pend": pend_g,
            "by_wh": by_wh, "tat": tat_state}


def build_report(records, delivery_type="all", zone="all", payment="all",
                 warehouse="all", account="all", weight="all",
                 slot="all", date_from="", date_to="", tier="all",
                 channel="all") -> dict:
    """Top-level entry: filter, aggregate, score, and assemble the payload."""
    _ckey = _report_cache_key(delivery_type, zone, payment, warehouse, account,
                              weight, slot, date_from, date_to, tier, channel)
    _cached = _report_cache_get(records, _ckey)
    if _cached is not None:
        return _cached

    rows = filter_records(records, delivery_type, zone, payment,
                          warehouse, account, weight,
                          slot, date_from, date_to, tier=tier, channel=channel)

    # Every per-row grouping the report needs is collected in a single scan.
    acc = _aggregate_all(rows)

    agg = attach_scores(_kpi_finalize(acc["carrier"], "carrier", ()))
    agg.sort(key=lambda a: (a["score"] is None, -(a["score"] or 0)))

    # Warehouse-wise breakdown, keyed on pickup pincode but labelled by city.
    # Same KPI block and scoring, with a volume threshold so the long tail of
    # tiny pincodes doesn't generate noisy ranks.
    wh = attach_scores(
        _kpi_finalize(acc["warehouse"], "pickup_pin",
                      (("warehouse", "warehouse"), ("city", "city"))),
        min_n=WAREHOUSE_MIN_N,
    )
    wh.sort(key=lambda a: -a["n"])  # warehouses ranked by volume
    _round_metrics(wh)
    wh_total_count = len(wh)
    if TOP_N_WAREHOUSES is not None:
        wh = wh[:TOP_N_WAREHOUSES]

    # Warehouse x carrier matrix (reusing the buckets built in the scan above).
    wh_carrier = aggregate_warehouse_carrier(rows, by_wh=acc["by_wh"])

    # Status-outcome matrix per carrier account (the ops pivot), plus the
    # FWD-Pendency sub-state breakdown.
    status_matrix = _status_finalize(acc["status"])
    pendency_matrix = _pendency_finalize(acc["pend"])

    # Promised-TAT (SLA) compliance counts for the current filter slice. This is
    # computed over the already-filtered rows, so it reflects the SELECTED date
    # range (not the whole loaded window). Surfaced as two summary cards
    # (In TAT / Out of TAT); only lanes with a mapped rule (Blue Dart) count.
    tat_overall = _tat_finalize(acc["tat"])

    # Destination rollup: same KPI block + scoring, grouped by the drop city
    # (where the parcel is going), ranked by volume and truncated to top-N.
    dest_agg = attach_scores(
        _kpi_finalize(acc["dest"], "drop_city", ()), min_n=DESTINATION_MIN_N
    )
    dest_agg.sort(key=lambda a: -a["n"])
    _round_metrics(dest_agg)
    dest_total = len(dest_agg)

    # Lane rollup: pickup city -> drop city pairs. Highest-volume lanes first.
    lane_agg = attach_scores(
        _kpi_finalize(acc["lane"], "lane", ()), min_n=LANE_MIN_N
    )
    lane_agg.sort(key=lambda a: -a["n"])
    _round_metrics(lane_agg)
    lane_total = len(lane_agg)

    # The FULL destination/lane lists are sent (not truncated) so the client
    # can search across every entry; the UI shows the top-N by volume until the
    # user types in the panel's search box. subscores aren't shown in these
    # tables, so drop them to keep the payload lean (these can be 1000s of rows).
    for _row in dest_agg:
        _row.pop("subscores", None)
    for _row in lane_agg:
        _row.pop("subscores", None)

    # Single pass over the filtered rows for every count/value reduction (order
    # counts, the four outcome value buckets, NDD, and the destination city-tier
    # count+revenue). This replaces ~12 separate generator scans and runs on
    # every re-filter. The four outcome buckets partition revenue exactly, so
    # Delivered + RTO + Pending + Cancelled == total (both value and order count).
    # ---- Revenue de-duplication --------------------------------------------
    # invoice_value is an ORDER-level figure that the source stamps onto every
    # AWB of the order, so summing it per shipment double-counts multi-package
    # orders. Count each distinct (order_id, value) once: identical values on the
    # same order collapse to a single contribution, while genuinely different
    # per-AWB values still add up (real multi-item orders). Rows with no order id
    # can't be de-duped, so each keeps its own value. `_rev` is the per-row
    # revenue contribution every ₹ rollup below uses instead of raw order_value.
    _seen_rev = set()
    _order_ids = set()
    _rev_no_oid = 0

    total = len(rows)
    delivered = picked = ndd_orders = rto = 0
    pending_orders = cancelled_orders = 0
    revenue = rto_value = delivered_value = pending_value = cancelled_value = 0.0
    tier_counts = {t: 0 for t in TIER_LEVELS + ["Unknown"]}
    tier_revenue = {t: 0.0 for t in TIER_LEVELS + ["Unknown"]}
    # Averages accumulated inline as (sum, count of non-None) instead of building
    # three throwaway lists and scanning the rows three more times via _mean().
    _p2o_sum = _p2d_sum = _o2s_sum = 0.0
    _p2o_n = _p2d_n = _o2s_n = 0
    # ONE pass now computes each row's de-duplicated revenue (_rev) AND every
    # count/value reduction below AND the three timing averages — folding what
    # used to be the revenue-dedup loop, the summary loop and three _mean() list
    # comprehensions (five full scans of `rows`) into a single scan.
    for r in rows:
        oid = r.get("order_id") or ""
        v = r.get("order_value") or 0.0
        if oid:
            _order_ids.add(oid)
            key = (oid, v)
            if key in _seen_rev:
                val = 0.0
            else:
                _seen_rev.add(key)
                val = v
        else:
            _rev_no_oid += 1
            val = v
        r["_rev"] = val

        revenue += val
        outcome = r.get("outcome")
        if r["delivered"]:
            delivered += 1
            delivered_value += val
        if r["picked"]:
            picked += 1
        if _is_ndd(r["account"]):
            ndd_orders += 1
        if outcome == "RTO":
            rto += 1
            rto_value += val
        elif outcome == "FWD Pendency":
            pending_orders += 1
            pending_value += val
        elif outcome == "Cancelled":
            cancelled_orders += 1
            cancelled_value += val
        t = r.get("tier") or "Unknown"
        tier_counts[t] = tier_counts.get(t, 0) + 1
        tier_revenue[t] = tier_revenue.get(t, 0.0) + val

        p2o_v = r["p2o"]
        if p2o_v is not None:
            _p2o_sum += p2o_v
            _p2o_n += 1
        p2d_v = r["p2d"]
        if p2d_v is not None:
            _p2d_sum += p2d_v
            _p2d_n += 1
        o2s_v = r["o2s"]
        if o2s_v is not None:
            _o2s_sum += o2s_v
            _o2s_n += 1
    # Distinct order units (for a true average ORDER value, not per-shipment).
    order_units = len(_order_ids) + _rev_no_oid
    aov = (revenue / order_units) if order_units else None
    avg_p2o = (_p2o_sum / _p2o_n) if _p2o_n else None
    avg_p2d = (_p2d_sum / _p2d_n) if _p2d_n else None
    avg_o2s = (_o2s_sum / _o2s_n) if _o2s_n else None

    # ---- Payment-mode performance -------------------------------------------
    # Management KPI block grouped by payment mode. `cod_exposure` (COD only) is
    # the collectable ₹ on COD orders still in transit — cash owed to the
    # business that hasn't been collected yet.
    def _perf_rows(groups):
        out = []
        for g in groups.values():
            n = g["n"]
            out.append({
                "group": g["group"],
                "n": n,
                "n_pct": _round(n / total * 100 if total else None),
                "revenue": _round(g["revenue"]),
                "revenue_pct": _round(g["revenue"] / revenue * 100 if revenue else None),
                "aov": _round(g["revenue"] / n if n else None),
                "delivered": g["delivered"],
                "success_rate": _round(g["delivered"] / g["picked"] * 100 if g["picked"] else None),
                "rto": g["rto"],
                "rto_pct": _round(g["rto"] / n * 100 if n else None),
                "cod_exposure": _round(g["cod_exposure"]),
            })
        out.sort(key=lambda a: -a["n"])
        return out

    # Payment-mode performance, the per-day order series and the product-category
    # tree are all built together in ONE pass over `rows` (previously three
    # separate scans):
    #   * pay_groups   -> payment-mode KPI block; cod_exposure = collectable ₹ on
    #                     COD orders still in transit (cash owed, not yet collected)
    #   * _daily       -> per-day order volume + revenue, keyed on the ORDER
    #                     date (falls back to pickup, then window date for
    #                     uploaded files) so every bar sits inside the selected
    #                     order-date range and the chart totals reconcile with
    #                     the KPI cards
    #   * product_tree -> category -> subcategory economics (volume, revenue, RTO)
    #                     plus which carrier accounts ship each category
    pay_groups: dict[str, dict] = {}
    _daily: dict[str, dict] = {}
    product_tree: dict[str, dict] = {}
    has_products = False
    for r in rows:
        val = r.get("_rev") or 0.0   # de-duplicated revenue (see revenue block)
        outcome = r.get("outcome")
        is_rto = outcome == "RTO"
        deliv = r["delivered"]
        pick = r["picked"]

        # -- payment-mode performance (key is never blank -> no rows skipped) --
        pk = r["payment"] or "(unknown)"
        g = pay_groups.get(pk)
        if g is None:
            g = {"group": pk, "n": 0, "picked": 0, "delivered": 0, "rto": 0,
                 "revenue": 0.0, "cod_exposure": 0.0}
            pay_groups[pk] = g
        g["n"] += 1
        g["revenue"] += val
        if pick:
            g["picked"] += 1
        if deliv:
            g["delivered"] += 1
        if is_rto:
            g["rto"] += 1
        if outcome == "FWD Pendency":
            g["cod_exposure"] += r.get("cod_value") or 0.0

        # -- per-day volume + revenue (rows with no usable date are omitted) --
        d = r.get("order_date") or r.get("pickup_date") or r.get("window_date")
        if d:
            cell = _daily.get(d)
            if cell is None:
                cell = {"date": d, "n": 0, "delivered": 0, "revenue": 0.0}
                _daily[d] = cell
            cell["n"] += 1
            cell["revenue"] += val
            if deliv:
                cell["delivered"] += 1

        # -- product economics tree --
        # Roll the fine catalogue category up to one of the 6 display groups
        # (the record keeps its fine category on the row).
        cat = canonical_category(r.get("category") or "Unknown")
        sub = r.get("subcategory") or "Unknown"
        acct = r.get("account") or "(unknown)"
        if r.get("item_name"):
            has_products = True
        c = product_tree.get(cat)
        if c is None:
            c = {"category": cat, "n": 0, "revenue": 0.0, "rto": 0,
                 "delivered": 0, "picked": 0, "subs": {}, "accts": {}}
            product_tree[cat] = c
        c["n"] += 1
        c["revenue"] += val
        if is_rto:
            c["rto"] += 1
        if deliv:
            c["delivered"] += 1
        if pick:
            c["picked"] += 1
        s = c["subs"].get(sub)
        if s is None:
            s = {"n": 0, "revenue": 0.0, "rto": 0}
            c["subs"][sub] = s
        s["n"] += 1
        s["revenue"] += val
        if is_rto:
            s["rto"] += 1
        c["accts"][acct] = c["accts"].get(acct, 0) + 1

    payment_perf = _perf_rows(pay_groups)
    daily = [{**_daily[d], "revenue": _round(_daily[d]["revenue"])} for d in sorted(_daily)]
    products = []
    # Ranked by revenue so the biggest-money categories lead the view.
    for c in sorted(product_tree.values(), key=lambda x: -x["revenue"]):
        subs = [
            {"subcategory": s, "n": sd["n"], "revenue": _round(sd["revenue"]),
             "rto_pct": _round(sd["rto"] / sd["n"] * 100 if sd["n"] else None)}
            for s, sd in sorted(c["subs"].items(), key=lambda kv: -kv[1]["revenue"])
        ]
        accts = [
            {"account": a, "n": n}
            for a, n in sorted(c["accts"].items(), key=lambda kv: -kv[1])
        ]
        products.append({
            "category": c["category"], "n": c["n"],
            "revenue": _round(c["revenue"]),
            "revenue_pct": _round(c["revenue"] / revenue * 100 if revenue else None),
            "rto_pct": _round(c["rto"] / c["n"] * 100 if c["n"] else None),
            "aov": _round(c["revenue"] / c["n"] if c["n"] else None),
            "success_rate": _round(c["delivered"] / c["picked"] * 100 if c["picked"] else None),
            "subs": subs, "accounts": accts,
        })

    _round_metrics(agg)

    # Filter option lists are built from the FULL record set (not the filtered
    # rows) so selecting one value doesn't empty out the other dropdowns. They
    # don't change between filters, so this is cached per loaded dataset.
    filter_opts = _filter_options(records)

    report = {
        "summary": {
            "total": total,
            "picked": picked,
            "delivered": delivered,
            "success_rate": _round(delivered / picked * 100 if picked else None),
            "avg_p2o": _round(avg_p2o),
            "avg_p2d": _round(avg_p2d),
            "avg_o2s": _round(avg_o2s),
            "carriers": len(agg),
            "warehouses": wh_total_count,
            "ndd_orders": ndd_orders,
            "ndd_pct": _round(ndd_orders / total * 100 if total else None),
            "tiers": tier_counts,
            "tier_revenue": {k: _round(v) for k, v in tier_revenue.items()},
            "revenue": _round(revenue),
            "rto_value": _round(rto_value),
            "rto_value_pct": _round(rto_value / revenue * 100 if revenue else None),
            "delivered_value": _round(delivered_value),
            "pending_value": _round(pending_value),
            "pending_value_pct": _round(pending_value / revenue * 100 if revenue else None),
            "pending_orders": pending_orders,
            "cancelled_value": _round(cancelled_value),
            "cancelled_value_pct": _round(cancelled_value / revenue * 100 if revenue else None),
            "cancelled_orders": cancelled_orders,
            "aov": _round(aov),
            "tat_in": tat_overall["in_tat"],
            "tat_out": tat_overall["out_tat"],
            "tat_in_pct": tat_overall["in_tat_pct"],
            "tat_out_pct": tat_overall["out_tat_pct"],
            # Out of TAT split: delivered-but-late vs not-yet-delivered breach.
            "tat_out_delivered": tat_overall["out_delivered"],
            "tat_out_pending": tat_overall["out_pending"],
            "rto": rto,
            "rto_pct": _round(rto / total * 100 if total else None),
        },
        "carriers": agg,
        "products": products,
        "has_products": has_products,
        "warehouses": wh,
        "warehouse_total": wh_total_count,
        "warehouse_top_n": TOP_N_WAREHOUSES,
        "warehouse_carrier": wh_carrier,
        "status_matrix": status_matrix,
        "pendency_matrix": pendency_matrix,
        "outcomes": OUTCOMES,
        "destinations": dest_agg,
        "destination_total": dest_total,
        "destination_top_n": TOP_N_DESTINATIONS,
        "lanes": lane_agg,
        "lane_total": lane_total,
        "lane_top_n": TOP_N_LANES,
        "lane_min_n": LANE_MIN_N,
        "warehouse_min_n": WAREHOUSE_MIN_N,
        "wh_carrier_min_n": WH_CARRIER_MIN_N,
        "payment_perf": payment_perf,
        "daily": daily,
        "filters": filter_opts,
        "weights": {k: int(v * 100) for k, v in WEIGHTS.items()},
    }
    _report_cache_put(records, _ckey, report)
    return report